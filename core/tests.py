from decimal import Decimal
from io import BytesIO
import os
from pathlib import Path
import tempfile
from unittest.mock import patch

from django.contrib.auth.models import Group, Permission, User
from django.contrib.admin.models import LogEntry
from django.contrib.auth import authenticate
from django.conf import settings
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import IntegrityError, connections
from django.test import TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from .permissions import sync_system_admin_role
from .models import (
    Account,
    Budget,
    Client,
    DepartmentChoices,
    Deployment,
    Employee,
    GuardSchedule,
    Contract,
    ContractSiteRequirement,
    Attendance,
    AttendanceDevice,
    AttendanceDeviceEvent,
    Invoice,
    JobOffer,
    JournalEntry,
    JournalLine,
    Payment,
    Position,
    RecruitmentApplication,
    RecruitmentInterview,
    RecruitmentRequisition,
    Role,
    RosterAttendance,
    Salary,
    Shift,
    Site,
    StatusChoices,
    SubscriptionPlan,
    TenantMembership,
    TenantOrganization,
    Training,
    Zone,
    ZoneEmployeeAllocation,
    ZoneSiteAllocation,
)
from .accounting import ensure_default_accounts, post_all_accounting, post_invoice, post_salary
from .crud import MODEL_REGISTRY
from .db_runtime import ensure_writable_sqlite_database
from .forms import ContractForm, ContractSiteRequirementForm, InvoiceForm
import security_management.settings as project_settings


class DatabaseRuntimeTests(TestCase):
    @override_settings(
        DATABASES={
            "default": {
                "ENGINE": "django.db.backends.sqlite3",
                "NAME": "/var/task/db.sqlite3",
            }
        }
    )
    def test_vercel_sqlite_path_is_switched_to_tmp(self):
        with override_settings(BASE_DIR="/var/task"):
            with self.settings():
                os.environ["VERCEL"] = "1"
                try:
                    changed = ensure_writable_sqlite_database()
                finally:
                    os.environ.pop("VERCEL", None)
                    connections["default"].close()

        self.assertTrue(changed)
        self.assertEqual(settings.DATABASES["default"]["NAME"].replace("\\", "/"), "/tmp/erp.sqlite3")

    def test_vercel_sqlite_database_url_is_switched_to_tmp(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            bundled_db = Path(temp_dir) / "db.sqlite3"
            writable_db = Path(temp_dir) / "erp.sqlite3"
            bundled_db.write_bytes(b"")
            with (
                patch.object(project_settings, "IS_VERCEL", True),
                patch.object(project_settings, "BASE_DIR", Path(temp_dir)),
                patch.dict(os.environ, {"DJANGO_SQLITE_TMP_NAME": str(writable_db)}, clear=False),
            ):
                config = project_settings.database_from_url(f"sqlite:///{bundled_db.as_posix()}")

        self.assertEqual(config["ENGINE"], "django.db.backends.sqlite3")
        self.assertEqual(Path(config["NAME"]), writable_db)


class EnvSuperuserBackendTests(TestCase):
    def test_env_superuser_is_created_when_credentials_match(self):
        with patch.dict(
            os.environ,
            {
                "DJANGO_SUPERUSER_USERNAME": "deploy-admin",
                "DJANGO_SUPERUSER_PASSWORD": "temporary-pass",
                "DJANGO_SUPERUSER_EMAIL": "admin@example.com",
            },
            clear=False,
        ):
            user = authenticate(username="deploy-admin", password="temporary-pass")

        self.assertIsNotNone(user)
        self.assertTrue(user.is_active)
        self.assertTrue(user.is_staff)
        self.assertTrue(user.is_superuser)
        self.assertTrue(user.check_password("temporary-pass"))

    def test_env_superuser_can_login_with_configured_email(self):
        with patch.dict(
            os.environ,
            {
                "DJANGO_SUPERUSER_USERNAME": "deploy-admin",
                "DJANGO_SUPERUSER_PASSWORD": "temporary-pass",
                "DJANGO_SUPERUSER_EMAIL": "admin@example.com",
            },
            clear=False,
        ):
            user = authenticate(username="admin@example.com", password="temporary-pass")

        self.assertIsNotNone(user)
        self.assertEqual(user.username, "deploy-admin")

    @override_settings(DISABLE_LAST_LOGIN_UPDATE=True, SESSION_ENGINE="django.contrib.sessions.backends.signed_cookies")
    def test_login_view_succeeds_when_database_login_writes_are_disabled(self):
        with patch.dict(
            os.environ,
            {
                "DJANGO_SUPERUSER_USERNAME": "deploy-admin",
                "DJANGO_SUPERUSER_PASSWORD": "temporary-pass",
            },
            clear=False,
        ):
            response = self.client.post(
                reverse("login"),
                {"username": "deploy-admin", "password": "temporary-pass"},
            )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], settings.LOGIN_REDIRECT_URL)
        self.assertIn("sessionid", self.client.cookies)


class PermanentLoginMiddlewareTests(TestCase):
    @override_settings(
        ERP_PERMANENT_LOGIN=True,
        ERP_PERMANENT_LOGIN_USERNAME="always-admin",
        ERP_PERMANENT_LOGIN_PASSWORD="",
        ERP_PERMANENT_LOGIN_EMAIL="admin@example.com",
        ERP_PERMANENT_LOGIN_AGE=315360000,
    )
    def test_protected_pages_open_without_credentials(self):
        response = self.client.get(reverse("core:dashboard"))

        self.assertEqual(response.status_code, 200)
        user = User.objects.get(username="always-admin")
        self.assertTrue(user.is_superuser)
        self.assertTrue(user.is_staff)
        self.assertGreaterEqual(int(self.client.session.get_expiry_age()), settings.ERP_PERMANENT_LOGIN_AGE - 5)

    @override_settings(
        ERP_PERMANENT_LOGIN=True,
        ERP_PERMANENT_LOGIN_USERNAME="always-admin",
        ERP_PERMANENT_LOGIN_PASSWORD="",
        ERP_PERMANENT_LOGIN_EMAIL="admin@example.com",
    )
    def test_login_page_redirects_when_permanent_login_is_active(self):
        response = self.client.get(reverse("login"))

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], settings.LOGIN_REDIRECT_URL)


class SaaSTenantMiddlewareTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="tenant-owner", password="pass", is_staff=True)
        self.plan = SubscriptionPlan.objects.create(name="Standard", slug="standard")
        self.organization = TenantOrganization.objects.create(
            name="Tenant One",
            slug="tenant-one",
            primary_domain="tenant.testserver",
            owner=self.user,
            plan=self.plan,
            status=TenantOrganization.Status.ACTIVE,
        )
        TenantMembership.objects.create(
            organization=self.organization,
            user=self.user,
            role=TenantMembership.Role.OWNER,
        )

    @override_settings(ALLOWED_HOSTS=["tenant.testserver", "testserver"])
    def test_request_resolves_current_tenant_from_domain(self):
        self.client.login(username="tenant-owner", password="pass")
        response = self.client.get(reverse("core:dashboard"), HTTP_HOST="tenant.testserver")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.wsgi_request.tenant, self.organization)

    @override_settings(SAAS_ENFORCE_TENANT_ACCESS=True)
    def test_tenant_enforcement_blocks_staff_without_membership(self):
        outsider = User.objects.create_user(username="outsider", password="pass", is_staff=True)
        self.client.login(username=outsider.username, password="pass")

        response = self.client.get(reverse("core:dashboard"))

        self.assertEqual(response.status_code, 302)


class AdminLoginRedirectTests(TestCase):
    def test_admin_login_redirects_to_dashboard_after_success(self):
        User.objects.create_user(
            username="staff-admin",
            password="temporary-pass",
            is_active=True,
            is_staff=True,
        )

        response = self.client.post(
            "/admin/login/?next=/admin/",
            {"username": "staff-admin", "password": "temporary-pass", "next": "/admin/"},
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], settings.LOGIN_REDIRECT_URL)
        self.assertIn("sessionid", self.client.cookies)


class ActivityLogAdminTests(TestCase):
    def setUp(self):
        self.admin_user = User.objects.create_superuser(
            username="activity-admin",
            email="activity@example.com",
            password="temporary-pass",
        )
        self.client.login(username="activity-admin", password="temporary-pass")

    def test_admin_activity_logs_record_and_display_admin_actions(self):
        response = self.client.post(
            "/admin/core/client/add/",
            {
                "client_name": "Logged Client",
                "contact_person": "Amina",
                "phone_number": "0700000200",
                "email": "",
                "address": "",
                "contract_start_date": "2026-05-01",
                "contract_end_date": "",
                "contract_status": StatusChoices.ACTIVE,
                "_save": "Save",
            },
        )

        self.assertEqual(response.status_code, 302)
        log_entry = LogEntry.objects.select_related("user", "content_type").get(object_repr="Logged Client")
        self.assertEqual(log_entry.user, self.admin_user)
        self.assertEqual(log_entry.content_type.model, "client")

        index = self.client.get("/admin/")
        changelist = self.client.get("/admin/admin/logentry/")

        self.assertContains(index, "Activity logs")
        self.assertEqual(changelist.status_code, 200)
        self.assertContains(changelist, "Activity logs")
        self.assertContains(changelist, "Logged Client")
        self.assertContains(changelist, "activity-admin")
        self.assertContains(changelist, "Added")

    def test_admin_activity_logs_are_read_only(self):
        self.assertEqual(self.client.get("/admin/admin/logentry/add/").status_code, 403)


class ContractDeliverableFormTests(TestCase):
    def setUp(self):
        self.client_record = Client.objects.create(
            client_name="Deliverable Client",
            contact_person="A",
            phone_number="0700000100",
            contract_start_date="2026-01-01",
        )
        self.base_data = {
            "client": self.client_record.pk,
            "start_date": "2026-01-01",
            "billing_rate": "1000.00",
            "status": StatusChoices.ACTIVE,
        }

    def test_other_contract_service_stores_deliverables(self):
        form = ContractForm(
            data={
                **self.base_data,
                "contract_number": "OTH-DEL-001",
                "service_type": Contract.ServiceType.OTHERS,
                "dog_count": "2",
                "dog_rate": "100.00",
                "metal_detector_count": "3",
                "metal_detector_rate": "200.00",
                "walk_through_detector_count": "1",
                "walk_through_detector_rate": "300.00",
                "panic_baton_count": "4",
                "panic_baton_rate": "50.00",
                "handcuffs_count": "5",
                "handcuffs_rate": "25.00",
            }
        )

        self.assertTrue(form.is_valid(), form.errors)
        contract = form.save()

        self.assertEqual(
            contract.other_deliverables,
            "Dogs: 2 @ 100.00, Metal detectors: 3 @ 200.00, Walk through detectors: 1 @ 300.00, Panic batons: 4 @ 50.00, Handcuffs: 5 @ 25.00",
        )

    def test_non_other_contract_service_zeros_deliverables(self):
        form = ContractForm(
            data={
                **self.base_data,
                "contract_number": "MANNED-DEL-001",
                "service_type": Contract.ServiceType.MANNED_GUARDING,
                "dog_count": "9",
                "dog_rate": "100.00",
            }
        )

        self.assertTrue(form.is_valid(), form.errors)
        contract = form.save()

        self.assertEqual(contract.dog_count, 0)
        self.assertEqual(contract.dog_rate, 0)
        self.assertEqual(contract.other_deliverables, "-")


class ContractSiteRequirementFormTests(TestCase):
    def test_requirement_creates_site_code_and_pulls_contract_deliverables(self):
        client = Client.objects.create(
            client_name="Requirement Client",
            contact_person="A",
            phone_number="0700000101",
            contract_start_date="2026-01-01",
        )
        contract = Contract.objects.create(
            client=client,
            contract_number="REQ-DATA-001",
            service_type=Contract.ServiceType.OTHERS,
            start_date="2026-02-01",
            end_date="2026-12-31",
            billing_rate=Decimal("150000.00"),
            dog_count=2,
            dog_rate=Decimal("10000.00"),
            metal_detector_count=3,
            metal_detector_rate=Decimal("20000.00"),
            walk_through_detector_count=1,
            walk_through_detector_rate=Decimal("30000.00"),
            panic_baton_count=4,
            panic_baton_rate=Decimal("5000.00"),
            handcuffs_count=5,
            handcuffs_rate=Decimal("2500.00"),
        )

        form = ContractSiteRequirementForm(
            data={
                "client": client.pk,
                "contract": contract.pk,
                "site_name": "Main Gate",
                "site_address": "Plot 1",
                "city": "Kampala",
                "required_guards": "6",
                "status": StatusChoices.ACTIVE,
            }
        )

        self.assertTrue(form.is_valid(), form.errors)
        requirement = form.save()

        self.assertEqual(requirement.site.site_code, "RCXXS0001")
        self.assertEqual(requirement.rate_per_guard, Decimal("150000.00"))
        self.assertEqual(requirement.start_date.isoformat(), "2026-02-01")
        self.assertEqual(requirement.end_date.isoformat(), "2026-12-31")
        self.assertEqual(requirement.dog_count, 2)
        self.assertEqual(requirement.dog_rate, Decimal("10000.00"))
        self.assertEqual(requirement.metal_detector_count, 3)
        self.assertEqual(requirement.walk_through_machine_count, 1)
        self.assertEqual(requirement.panic_baton_count, 4)
        self.assertEqual(requirement.handcuffs_count, 5)


class AdminRolePermissionTests(TestCase):
    def setUp(self):
        self.admin_user = User.objects.create_user(
            username="staff-admin",
            password="pass",
            is_staff=True,
            is_superuser=False,
        )
        sync_system_admin_role(assign_active_staff=True)
        self.admin_user.refresh_from_db()
        self.client.force_login(self.admin_user)

    def test_system_admin_role_can_manage_users_without_superuser(self):
        self.assertFalse(self.admin_user.is_superuser)
        self.assertTrue(self.admin_user.has_perm("auth.add_user"))
        self.assertTrue(self.admin_user.has_perm("auth.change_user"))
        self.assertTrue(self.admin_user.has_perm("auth.view_group"))

        response = self.client.get(reverse("admin:auth_user_add"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Create user")
        self.assertContains(response, "Role assignment")

    def test_user_change_form_allows_roles_and_direct_permissions(self):
        target = User.objects.create_user(username="new-operator", password="pass", is_staff=True)

        response = self.client.get(reverse("admin:auth_user_change", args=[target.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Role assignment")
        self.assertContains(response, "Direct permissions")
        self.assertContains(response, "User permissions")
        self.assertContains(response, "Superuser status")


class FinanceCalculationTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="finance-manager", password="pass")
        manager_group, _created = Group.objects.get_or_create(name="Manager")
        self.user.groups.add(manager_group)
        self.client.login(username="finance-manager", password="pass")
        self.role = Role.objects.create(role_name="Guard", department=DepartmentChoices.OPERATIONS)
        self.position = Position.objects.create(
            position_title="Security Guard",
            department=DepartmentChoices.OPERATIONS,
            salary_range_min=Decimal("300000.00"),
            salary_range_max=Decimal("600000.00"),
        )
        self.employee = Employee.objects.create(
            first_name="Alex",
            last_name="Okello",
            phone_number="0700000000",
            email="alex@example.com",
            national_id="NIN001",
            role=self.role,
            position=self.position,
        )

    def test_salary_calculates_net_salary(self):
        salary = Salary.objects.create(
            employee=self.employee,
            pay_period_start="2026-05-01",
            pay_period_end="2026-05-31",
            basic_salary=Decimal("500000.00"),
            allowances=Decimal("50000.00"),
            deductions=Decimal("25000.00"),
            overtime_pay=Decimal("10000.00"),
            bonus=Decimal("15000.00"),
        )

        self.assertEqual(salary.gross_pay, Decimal("575000.00"))
        self.assertEqual(salary.nssf_employee, Decimal("28750.00"))
        self.assertEqual(salary.nssf_employer, Decimal("57500.00"))
        self.assertEqual(salary.total_deductions, Decimal("53750.00"))
        self.assertEqual(salary.net_salary, Decimal("521250.00"))

    def test_shift_splits_basic_and_overtime_hours_under_uganda_law(self):
        shift = Shift.objects.create(
            shift_name="Twelve Hour Day",
            code="T12",
            start_time="08:00",
            end_time="20:00",
        )
        shift.refresh_from_db()

        self.assertEqual(shift.duration_hours, Decimal("12.00"))
        self.assertEqual(shift.basic_hours, Decimal("8.00"))
        self.assertEqual(shift.daily_overtime_hours, Decimal("4.00"))
        self.assertEqual(shift.normal_day_overtime_multiplier, Decimal("1.50"))
        self.assertEqual(shift.public_holiday_overtime_multiplier, Decimal("2.00"))

    def test_invoice_balance_and_paid_status_are_calculated(self):
        client = Client.objects.create(
            client_name="Acme Mall",
            contact_person="Jane",
            phone_number="0711111111",
            contract_start_date="2026-01-01",
        )
        invoice = Invoice.objects.create(
            client=client,
            invoice_number="INV-001",
            due_date="2026-06-15",
            total_amount=Decimal("1000000.00"),
            paid_amount=Decimal("1000000.00"),
        )

        self.assertEqual(invoice.balance_amount, Decimal("0.00"))
        self.assertEqual(invoice.status, StatusChoices.PAID)

    def test_invoice_pdf_is_downloadable(self):
        client = Client.objects.create(
            client_name="Acme Mall",
            contact_person="Jane",
            phone_number="0711111111",
            email="jane@example.com",
            address="Plot 10 Kampala",
            contract_start_date="2026-01-01",
        )
        invoice = Invoice.objects.create(
            client=client,
            invoice_number="INV-PDF-001",
            due_date="2026-06-15",
            guard_count=2,
            rate_per_guard=Decimal("300000.00"),
            paid_amount=Decimal("100000.00"),
        )

        response = self.client.get(f"/invoices/{invoice.pk}/pdf/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertTrue(response.content.startswith(b"%PDF"))

    def test_finance_aging_and_reconciliation_reports_render(self):
        client = Client.objects.create(
            client_name="Aging Client",
            contact_person="Jane",
            phone_number="0711111111",
            contract_start_date="2026-01-01",
        )
        posted_invoice = Invoice.objects.create(
            client=client,
            invoice_number="INV-POSTED-001",
            invoice_date="2026-05-01",
            due_date="2026-05-15",
            guard_count=1,
            rate_per_guard=Decimal("100000.00"),
        )
        missing_invoice = Invoice.objects.create(
            client=client,
            invoice_number="INV-MISSING-001",
            invoice_date="2026-05-01",
            due_date="2026-04-15",
            guard_count=1,
            rate_per_guard=Decimal("200000.00"),
        )
        post_invoice(posted_invoice)

        aging_response = self.client.get("/accounting/receivables-aging/", {"as_of": "2026-05-24"})
        reconciliation_response = self.client.get("/accounting/reconciliation/")
        payroll_reconciliation_response = self.client.get("/accounting/reconciliation/payroll/")
        expense_reconciliation_response = self.client.get("/accounting/reconciliation/expenses/")
        payment_reconciliation_response = self.client.get("/accounting/reconciliation/payments/")

        self.assertEqual(aging_response.status_code, 200)
        self.assertContains(aging_response, "Receivables Aging Report")
        self.assertContains(aging_response, "Customer Code")
        self.assertContains(aging_response, "Debt Collector")
        self.assertContains(aging_response, "Balance Due")
        self.assertContains(aging_response, "Aging Client")
        self.assertEqual(reconciliation_response.status_code, 200)
        self.assertContains(reconciliation_response, "INV-POSTED-001")
        self.assertContains(reconciliation_response, "Matched")
        self.assertContains(reconciliation_response, "INV-MISSING-001")
        self.assertContains(reconciliation_response, "Missing")
        self.assertEqual(payroll_reconciliation_response.status_code, 200)
        self.assertContains(payroll_reconciliation_response, "Payroll Reconciliation")
        self.assertEqual(expense_reconciliation_response.status_code, 200)
        self.assertContains(expense_reconciliation_response, "Expense Reconciliation")
        self.assertEqual(payment_reconciliation_response.status_code, 200)
        self.assertContains(payment_reconciliation_response, "Payment Reconciliation")

    def test_budget_remaining_amount_is_calculated(self):
        budget = Budget.objects.create(
            year=2026,
            department=DepartmentChoices.OPERATIONS,
            category="Equipment",
            allocated_amount=Decimal("2000000.00"),
            spent_amount=Decimal("750000.00"),
        )

        self.assertEqual(budget.remaining_amount, Decimal("1250000.00"))

    def test_default_accounts_are_created(self):
        accounts = ensure_default_accounts()

        self.assertIn("1000", accounts)
        self.assertEqual(Account.objects.count(), 12)

    def test_invoice_posts_balanced_journal_entry(self):
        client = Client.objects.create(
            client_name="Ledger Client",
            contact_person="Jane",
            phone_number="0711111112",
            contract_start_date="2026-01-01",
        )
        invoice = Invoice.objects.create(
            client=client,
            invoice_number="INV-LEDGER-001",
            due_date="2026-06-15",
            total_amount=Decimal("750000.00"),
        )

        entry = post_invoice(invoice)

        self.assertTrue(entry.is_balanced)
        self.assertEqual(entry.total_debit, Decimal("750000.00"))
        self.assertEqual(entry.total_credit, Decimal("750000.00"))
        self.assertTrue(JournalLine.objects.filter(account__account_code="2300").exists())

    def test_invoice_calculates_from_contract_site_equipment_and_vat(self):
        client = Client.objects.create(
            client_name="Auto Billing Client",
            contact_person="Jane",
            phone_number="0711111114",
            contract_start_date="2026-01-01",
        )
        site = Site.objects.create(
            client=client,
            site_name="Auto Billing Site",
            site_address="Plot 2",
            city="Kampala",
            required_guards_per_shift=3,
        )
        contract = Contract.objects.create(
            client=client,
            contract_number="AUTO-BILL-001",
            start_date="2026-01-01",
            billing_rate=Decimal("100000.00"),
            status=StatusChoices.ACTIVE,
        )
        ContractSiteRequirement.objects.create(
            contract=contract,
            site=site,
            required_guards=4,
            start_date="2026-01-01",
            status=StatusChoices.ACTIVE,
        )

        invoice = Invoice.objects.create(
            client=client,
            contract=contract,
            site=site,
            billing_month="2026-05-01",
            due_date="2026-06-15",
            gun_count=2,
            gun_rate=Decimal("50000.00"),
            radio_count=1,
            radio_rate=Decimal("25000.00"),
            metal_detector_count=1,
            metal_detector_rate=Decimal("30000.00"),
            walk_through_machine_count=1,
            walk_through_machine_rate=Decimal("70000.00"),
            dog_count=1,
            dog_rate=Decimal("90000.00"),
        )

        self.assertTrue(invoice.invoice_number.startswith("INV-2026"))
        self.assertEqual(invoice.guard_count, 4)
        self.assertEqual(invoice.rate_per_guard, Decimal("100000.00"))
        self.assertEqual(invoice.subtotal_amount, Decimal("715000.00"))
        self.assertEqual(invoice.vat_amount, Decimal("128700.00"))
        self.assertEqual(invoice.total_amount, Decimal("843700.00"))
        self.assertEqual(invoice.balance_amount, Decimal("843700.00"))

    def test_invoice_can_bill_all_sites_under_contract(self):
        client = Client.objects.create(
            client_name="Multi Site Client",
            contact_person="Ruth",
            phone_number="0711111120",
            email="ruth@example.com",
            address="Plot 45 Kampala",
            contract_start_date="2026-01-01",
        )
        first_site = Site.objects.create(client=client, site_name="Gate A", site_address="Plot A", city="Kampala")
        second_site = Site.objects.create(client=client, site_name="Warehouse", site_address="Plot B", city="Kampala")
        contract = Contract.objects.create(
            client=client,
            contract_number="MULTI-001",
            start_date="2026-01-01",
            billing_rate=Decimal("100000.00"),
            status=StatusChoices.ACTIVE,
        )
        ContractSiteRequirement.objects.create(
            contract=contract,
            site=first_site,
            required_guards=2,
            rate_per_guard=Decimal("120000.00"),
            gun_count=1,
            gun_rate=Decimal("50000.00"),
            start_date="2026-01-01",
            status=StatusChoices.ACTIVE,
        )
        ContractSiteRequirement.objects.create(
            contract=contract,
            site=second_site,
            required_guards=3,
            radio_count=2,
            radio_rate=Decimal("25000.00"),
            start_date="2026-01-01",
            status=StatusChoices.ACTIVE,
        )

        invoice = Invoice.objects.create(
            client=client,
            contract=contract,
            billing_scope=Invoice.BillingScope.CONTRACT,
            billing_month="2026-05-01",
            due_date="2026-06-15",
        )

        self.assertIsNone(invoice.site)
        self.assertEqual(invoice.client_name, "Multi Site Client")
        self.assertEqual(invoice.client_email, "ruth@example.com")
        self.assertEqual(invoice.guard_count, 5)
        self.assertEqual(invoice.subtotal_amount, Decimal("640000.00"))
        self.assertEqual(invoice.total_amount, Decimal("755200.00"))

    def test_invoice_can_bill_selected_sites_under_one_contract(self):
        client = Client.objects.create(
            client_name="Selected Sites Client",
            contact_person="Sarah",
            phone_number="0711111188",
            contract_start_date="2026-01-01",
        )
        first_site = Site.objects.create(client=client, site_name="Main Gate", site_address="Plot A", city="Kampala")
        second_site = Site.objects.create(client=client, site_name="Warehouse", site_address="Plot B", city="Kampala")
        third_site = Site.objects.create(client=client, site_name="Admin Block", site_address="Plot C", city="Kampala")
        contract = Contract.objects.create(
            client=client,
            contract_number="SELECTED-001",
            start_date="2026-01-01",
            billing_rate=Decimal("100000.00"),
            status=StatusChoices.ACTIVE,
        )
        ContractSiteRequirement.objects.create(
            contract=contract,
            site=first_site,
            required_guards=2,
            rate_per_guard=Decimal("120000.00"),
            start_date="2026-01-01",
            status=StatusChoices.ACTIVE,
        )
        ContractSiteRequirement.objects.create(
            contract=contract,
            site=second_site,
            required_guards=3,
            rate_per_guard=Decimal("100000.00"),
            radio_count=1,
            radio_rate=Decimal("50000.00"),
            start_date="2026-01-01",
            status=StatusChoices.ACTIVE,
        )
        ContractSiteRequirement.objects.create(
            contract=contract,
            site=third_site,
            required_guards=5,
            rate_per_guard=Decimal("100000.00"),
            start_date="2026-01-01",
            status=StatusChoices.ACTIVE,
        )

        invoice = Invoice.objects.create(
            client=client,
            contract=contract,
            billing_scope=Invoice.BillingScope.MULTIPLE_SITES,
            billing_month="2026-05-01",
            due_date="2026-06-15",
        )
        invoice.selected_sites.set([first_site, second_site])
        invoice.save()

        self.assertIsNone(invoice.site)
        self.assertEqual(invoice.guard_count, 5)
        self.assertEqual(invoice.radio_count, 1)
        self.assertEqual(invoice.subtotal_amount, Decimal("590000.00"))
        self.assertEqual(invoice.total_amount, Decimal("696200.00"))

    def test_invoice_can_bill_one_site_under_contract(self):
        client = Client.objects.create(
            client_name="Single Site Client",
            contact_person="Mark",
            phone_number="0711111121",
            contract_start_date="2026-01-01",
        )
        first_site = Site.objects.create(client=client, site_name="Main", site_address="Plot A", city="Kampala")
        second_site = Site.objects.create(client=client, site_name="Branch", site_address="Plot B", city="Kampala")
        contract = Contract.objects.create(
            client=client,
            contract_number="ONE-001",
            start_date="2026-01-01",
            billing_rate=Decimal("80000.00"),
            status=StatusChoices.ACTIVE,
        )
        ContractSiteRequirement.objects.create(
            contract=contract,
            site=first_site,
            required_guards=2,
            start_date="2026-01-01",
            status=StatusChoices.ACTIVE,
        )
        ContractSiteRequirement.objects.create(
            contract=contract,
            site=second_site,
            required_guards=4,
            start_date="2026-01-01",
            status=StatusChoices.ACTIVE,
        )

        invoice = Invoice.objects.create(
            client=client,
            contract=contract,
            billing_scope=Invoice.BillingScope.SITE,
            site=second_site,
            billing_month="2026-05-01",
            due_date="2026-06-15",
        )

        self.assertEqual(invoice.site, second_site)
        self.assertEqual(invoice.guard_count, 4)
        self.assertEqual(invoice.subtotal_amount, Decimal("320000.00"))

    def test_contract_requirement_defaults_guard_rate_from_contract(self):
        client = Client.objects.create(
            client_name="Requirement Rate Client",
            contact_person="Rose",
            phone_number="0711111123",
            contract_start_date="2026-01-01",
        )
        site = Site.objects.create(client=client, site_name="Rate Site", site_address="Plot R", city="Kampala")
        contract = Contract.objects.create(
            client=client,
            contract_number="RATE-001",
            start_date="2026-01-01",
            billing_rate=Decimal("175000.00"),
            status=StatusChoices.ACTIVE,
        )

        requirement = ContractSiteRequirement.objects.create(
            contract=contract,
            site=site,
            required_guards=2,
            start_date="2026-01-01",
            status=StatusChoices.ACTIVE,
        )

        self.assertEqual(requirement.rate_per_guard, Decimal("175000.00"))
        self.assertEqual(requirement.billing_rate, Decimal("175000.00"))
        self.assertEqual(requirement.billable_total, Decimal("350000.00"))

    def test_contract_requirement_can_override_guard_rate(self):
        client = Client.objects.create(
            client_name="Override Rate Client",
            contact_person="Rose",
            phone_number="0711111124",
            contract_start_date="2026-01-01",
        )
        site = Site.objects.create(client=client, site_name="Override Site", site_address="Plot O", city="Kampala")
        contract = Contract.objects.create(
            client=client,
            contract_number="RATE-002",
            start_date="2026-01-01",
            billing_rate=Decimal("175000.00"),
            status=StatusChoices.ACTIVE,
        )

        requirement = ContractSiteRequirement.objects.create(
            contract=contract,
            site=site,
            required_guards=2,
            rate_per_guard=Decimal("200000.00"),
            start_date="2026-01-01",
            status=StatusChoices.ACTIVE,
        )

        self.assertEqual(requirement.billing_rate, Decimal("200000.00"))
        self.assertEqual(requirement.billable_total, Decimal("400000.00"))

    def test_contract_invoice_data_returns_client_and_sites(self):
        client = Client.objects.create(
            client_name="Endpoint Client",
            contact_person="Sarah",
            phone_number="0711111122",
            email="sarah@example.com",
            address="Plot 90",
            contract_start_date="2026-01-01",
        )
        site = Site.objects.create(client=client, site_name="Endpoint Site", site_address="Plot E", city="Kampala")
        contract = Contract.objects.create(
            client=client,
            contract_number="END-001",
            start_date="2026-01-01",
            billing_rate=Decimal("75000.00"),
            status=StatusChoices.ACTIVE,
        )
        ContractSiteRequirement.objects.create(
            contract=contract,
            site=site,
            required_guards=3,
            start_date="2026-01-01",
            status=StatusChoices.ACTIVE,
        )

        response = self.client.get(f"/contracts/{contract.id}/invoice-data/?billing_month=2026-05-01")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["client"]["name"], "Endpoint Client")
        self.assertEqual(payload["client"]["email"], "sarah@example.com")
        self.assertEqual(payload["sites"][0]["id"], site.id)
        self.assertEqual(payload["sites"][0]["guards"], 3)

    def test_invoice_add_form_uses_billing_form_and_hides_manual_totals(self):
        response = self.client.get("/records/invoices/add/")
        current_month = timezone.localdate().replace(day=1).isoformat()

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Billing Month")
        self.assertContains(response, f'id="id_billing_month"')
        self.assertContains(response, f'value="{current_month}"')
        self.assertContains(response, "Number of Guards")
        self.assertContains(response, "Rate Per Guard")
        self.assertContains(response, "18% VAT")
        self.assertNotContains(response, "Total amount")

    def test_invoice_form_defaults_and_normalizes_billing_month(self):
        form = InvoiceForm()

        self.assertEqual(form.fields["billing_month"].initial, timezone.localdate().replace(day=1))
        self.assertTrue(form.fields["billing_month"].required)

        contract = Contract.objects.create(
            client=Client.objects.create(
                client_name="Billing Client",
                contact_person="Jane",
                phone_number="0711111199",
                contract_start_date="2026-01-01",
            ),
            contract_number="BILL-001",
            service_type="Manned Guarding",
            start_date="2026-01-01",
            end_date="2026-12-31",
            billing_rate=Decimal("500000.00"),
        )
        form = InvoiceForm(
            data={
                "contract": contract.id,
                "client": contract.client_id,
                "billing_scope": Invoice.BillingScope.CONTRACT,
                "billing_month": "2026-05-18",
                "invoice_date": "2026-05-28",
                "due_date": "2026-06-28",
                "guard_count": 1,
                "rate_per_guard": "500000.00",
                "paid_amount": "0.00",
            }
        )

        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.cleaned_data["billing_month"].isoformat(), "2026-05-01")

    def test_payroll_posts_salary_expense_and_payables(self):
        salary = Salary.objects.create(
            employee=self.employee,
            pay_period_start="2026-05-01",
            pay_period_end="2026-05-31",
            basic_salary=Decimal("500000.00"),
            allowances=Decimal("0.00"),
            deductions=Decimal("25000.00"),
            overtime_pay=Decimal("0.00"),
            bonus=Decimal("0.00"),
        )

        entry = post_salary(salary)

        self.assertTrue(entry.is_balanced)
        self.assertEqual(entry.total_debit, entry.total_credit)
        self.assertTrue(JournalLine.objects.filter(account__account_code="5000", debit=salary.gross_pay).exists())
        self.assertTrue(JournalLine.objects.filter(account__account_code="2100", credit=salary.net_salary).exists())

    def test_accounting_reports_render_from_posted_journals(self):
        client = Client.objects.create(
            client_name="Report Client",
            contact_person="Jane",
            phone_number="0711111113",
            contract_start_date="2026-01-01",
        )
        invoice = Invoice.objects.create(
            client=client,
            invoice_number="INV-REPORT-001",
            due_date="2026-06-15",
            total_amount=Decimal("1000000.00"),
        )
        Payment.objects.create(
            invoice=invoice,
            payment_date="2026-05-20",
            amount=Decimal("250000.00"),
            payment_method="Bank",
            transaction_ref="REPORT-PAY-001",
        )
        post_all_accounting()

        for path in [
            "/accounting/general-ledger/",
            "/accounting/trial-balance/",
            "/accounting/balance-sheet/",
            "/accounting/income-statement/",
        ]:
            with self.subTest(path=path):
                response = self.client.get(path)
                self.assertEqual(response.status_code, 200)


class GuardSchedulingTests(TestCase):
    def setUp(self):
        self.hr_user = User.objects.create_user(username="hr", password="pass")
        hr_group, _ = Group.objects.get_or_create(name="Human Resources Manager")
        self.hr_user.groups.add(hr_group)
        self.client.login(username="hr", password="pass")
        self.client_record = Client.objects.create(
            client_name="Central Plaza",
            contact_person="Sarah",
            phone_number="0700000001",
            contract_start_date="2026-01-01",
        )
        self.site = Site.objects.create(
            client=self.client_record,
            site_name="Main Gate",
            site_address="Plot 1",
            city="Kampala",
            latitude=Decimal("0.347596"),
            longitude=Decimal("32.582520"),
            geofence_radius_meters=150,
        )
        self.contract = Contract.objects.create(
            client=self.client_record,
            contract_number="TEST-CON-001",
            start_date="2026-01-01",
            status=StatusChoices.ACTIVE,
        )
        ContractSiteRequirement.objects.create(
            contract=self.contract,
            site=self.site,
            required_guards=10,
            start_date="2026-01-01",
            status=StatusChoices.ACTIVE,
        )
        role = Role.objects.create(role_name="Scheduling Guard", department=DepartmentChoices.OPERATIONS)
        employee = Employee.objects.create(
            first_name="Brian",
            last_name="Kato",
            phone_number="0700000002",
            email="brian@example.com",
            national_id="NIN002",
            role=role,
            company_number="G-001",
        )
        self.guard = employee
        self.guard.work_card_uid = "CARD-001"
        self.guard.save(update_fields=["work_card_uid", "updated_at"])
        replacement_employee = Employee.objects.create(
            first_name="Cathy",
            last_name="Namuli",
            phone_number="0700000005",
            email="cathy@example.com",
            national_id="NIN005",
            role=role,
            company_number="G-003",
        )
        self.replacement_guard = replacement_employee
        supervisor_employee = Employee.objects.create(
            first_name="Doreen",
            last_name="Supervisor",
            phone_number="0700000006",
            email="doreen@example.com",
            national_id="NIN006",
            role=role,
        )
        self.supervisor = supervisor_employee
        self.zone = Zone.objects.create(zone_code="TEST-ZN-001", zone_name="Test Zone", supervisor=self.supervisor)
        ZoneSiteAllocation.objects.create(zone=self.zone, site=self.site)
        self.shift, _created = Shift.objects.update_or_create(
            code="D",
            defaults={
                "shift_name": "Day",
                "start_time": "08:00",
                "end_time": "20:00",
            },
        )
        self.deployment = Deployment.objects.create(
            employee=self.guard,
            client=self.client_record,
            site=self.site,
            shift=self.shift,
            start_date="2026-05-01",
        )

    def test_iot_swipe_captures_attendance_inside_geofence(self):
        schedule = GuardSchedule.objects.create(
            deployment=self.deployment,
            employee=self.guard,
            site=self.site,
            shift=self.shift,
            shift_date="2026-05-16",
        )
        device = AttendanceDevice.objects.create(
            device_id="SUP-DEVICE-001",
            name="Supervisor handheld",
            api_key="secret-token",
            assigned_site=self.site,
            assigned_supervisor=self.supervisor,
        )

        response = self.client.post(
            "/api/attendance/swipe/",
            data={
                "card_uid": "CARD-001",
                "device_id": device.device_id,
                "site_id": self.site.id,
                "timestamp": "2026-05-16T08:02:00+03:00",
                "event": "check_in",
                "latitude": "0.347596",
                "longitude": "32.582520",
            },
            content_type="application/json",
            HTTP_X_DEVICE_TOKEN="secret-token",
        )

        self.assertEqual(response.status_code, 200)
        attendance = Attendance.objects.get(employee=self.guard, date="2026-05-16", shift=self.shift)
        schedule.refresh_from_db()
        self.assertEqual(attendance.status, "Present")
        self.assertEqual(attendance.capture_source, Attendance.CaptureSource.IOT)
        self.assertEqual(attendance.device_id, device.device_id)
        self.assertEqual(schedule.status, GuardSchedule.ScheduleStatus.COMPLETED)
        self.assertEqual(AttendanceDeviceEvent.objects.get().status, AttendanceDeviceEvent.EventStatus.ACCEPTED)

    def test_deployment_list_shows_and_searches_employee_number(self):
        response = self.client.get("/records/deployments/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Employee Number")
        self.assertContains(response, self.guard.company_number)
        self.assertContains(response, "Export Excel")
        self.assertContains(response, "Import Excel")

        response = self.client.get("/records/deployments/", {"q": self.guard.company_number})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.guard.full_name)
        self.assertNotContains(response, "No records found.")

    def test_deployment_excel_export_downloads_workbook(self):
        response = self.client.get("/deployments/export/excel/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook.active
        headers = [cell.value for cell in worksheet[1]]
        self.assertEqual(headers[:4], ["employee_number", "employee_name", "client", "site_code"])
        self.assertEqual(worksheet["A2"].value, self.guard.company_number)

    def test_deployment_excel_import_updates_deployment(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["employee_number", "site_code", "shift_code", "start_date", "end_date", "status"])
        worksheet.append(
            [
                self.guard.company_number,
                self.site.site_code,
                self.shift.code,
                "2026-05-01",
                "2026-05-31",
                "inactive",
            ]
        )
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        upload = SimpleUploadedFile(
            "deployments.xlsx",
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post("/deployments/import/excel/", {"deployment_file": upload})

        self.assertRedirects(response, "/records/deployments/")
        self.deployment.refresh_from_db()
        self.assertEqual(self.deployment.end_date.isoformat(), "2026-05-31")
        self.assertEqual(self.deployment.status, StatusChoices.INACTIVE)
        self.assertEqual(Deployment.objects.count(), 1)

    def test_iot_swipe_rejects_attendance_outside_geofence(self):
        GuardSchedule.objects.create(
            deployment=self.deployment,
            employee=self.guard,
            site=self.site,
            shift=self.shift,
            shift_date="2026-05-16",
        )
        device = AttendanceDevice.objects.create(
            device_id="SUP-DEVICE-002",
            name="Supervisor handheld",
            api_key="secret-token-2",
            assigned_site=self.site,
        )

        response = self.client.post(
            "/api/attendance/swipe/",
            data={
                "card_uid": "CARD-001",
                "device_id": device.device_id,
                "site_id": self.site.id,
                "timestamp": "2026-05-16T08:02:00+03:00",
                "event": "check_in",
                "latitude": "0.360000",
                "longitude": "32.600000",
            },
            content_type="application/json",
            HTTP_X_DEVICE_TOKEN="secret-token-2",
        )

        self.assertEqual(response.status_code, 400)
        self.assertFalse(Attendance.objects.filter(employee=self.guard, date="2026-05-16").exists())
        event = AttendanceDeviceEvent.objects.get()
        self.assertEqual(event.status, AttendanceDeviceEvent.EventStatus.REJECTED)
        self.assertIn("outside geofence", event.message)

    def test_attendance_page_generates_guard_schedule(self):
        response = self.client.get(
            "/attendances/",
            {"site": self.site.id, "date": "2026-05-16"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(GuardSchedule.objects.count(), 1)
        schedule = GuardSchedule.objects.get()
        self.assertEqual(schedule.deployment, self.deployment)
        self.assertContains(response, "Brian Kato")

    def test_attendance_page_defaults_to_today_all_sites(self):
        today = timezone.localdate()
        self.deployment.start_date = today
        self.deployment.save(update_fields=["start_date", "updated_at"])

        response = self.client.get("/attendances/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(GuardSchedule.objects.count(), 1)
        self.assertContains(response, "All sites")
        self.assertContains(response, "Brian Kato")

    def test_scheduled_guard_attendance_can_be_saved(self):
        self.client.get("/attendances/", {"site": self.site.id, "date": "2026-05-16"})
        schedule = GuardSchedule.objects.get()

        response = self.client.post(
            "/attendances/",
            {
                "site": str(self.site.id),
                "date": "2026-05-16",
                "schedule_ids": [str(schedule.id)],
                f"present_{schedule.id}": "on",
                f"reason_{schedule.id}": "Reported on time",
            },
        )

        self.assertEqual(response.status_code, 302)
        attendance = Attendance.objects.get(schedule=schedule)
        schedule.refresh_from_db()
        self.assertEqual(attendance.status, "Present")
        self.assertEqual(attendance.remarks, "Reported on time")
        self.assertEqual(schedule.status, GuardSchedule.ScheduleStatus.COMPLETED)

    def test_scheduled_guard_attendance_can_be_reassigned_without_duplicate_schedule(self):
        self.client.get("/attendances/", {"site": self.site.id, "date": "2026-05-16"})
        schedule = GuardSchedule.objects.get()
        self.client.post(
            "/attendances/",
            {
                "site": str(self.site.id),
                "date": "2026-05-16",
                "schedule_ids": [str(schedule.id)],
                f"scheduled_guard_{schedule.id}": str(self.guard.id),
                f"present_{schedule.id}": "yes",
            },
        )

        response = self.client.post(
            "/attendances/",
            {
                "site": str(self.site.id),
                "date": "2026-05-16",
                "schedule_ids": [str(schedule.id)],
                f"scheduled_guard_{schedule.id}": str(self.replacement_guard.id),
                f"present_{schedule.id}": "yes",
                f"reason_{schedule.id}": "Roster correction",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(Attendance.objects.filter(schedule=schedule).count(), 1)
        attendance = Attendance.objects.get(schedule=schedule)
        schedule.refresh_from_db()
        self.assertEqual(attendance.employee, self.replacement_guard)
        self.assertEqual(attendance.remarks, "Roster correction")
        self.assertEqual(schedule.employee, self.replacement_guard)

    def test_absent_scheduled_guard_can_have_replacement(self):
        self.client.get("/attendances/", {"site": self.site.id, "date": "2026-05-16"})
        schedule = GuardSchedule.objects.get()

        response = self.client.post(
            "/attendances/",
            {
                "site": str(self.site.id),
                "date": "2026-05-16",
                "schedule_ids": [str(schedule.id)],
                f"scheduled_guard_{schedule.id}": str(self.guard.id),
                f"replacement_guard_{schedule.id}": str(self.replacement_guard.id),
                f"reason_{schedule.id}": "Scheduled guard called in sick",
            },
        )

        self.assertEqual(response.status_code, 302)
        schedule.refresh_from_db()
        self.assertEqual(schedule.status, GuardSchedule.ScheduleStatus.MISSED)
        self.assertEqual(schedule.replacement_employee, self.replacement_guard)
        self.assertEqual(schedule.replacement_reason, "Scheduled guard called in sick")
        self.assertEqual(Attendance.objects.get(employee=self.guard).status, "Absent")
        replacement_attendance = Attendance.objects.get(employee=self.replacement_guard)
        self.assertEqual(replacement_attendance.status, "Present")

    def test_replacement_guard_attendance_counts_in_payroll(self):
        self.client.get("/attendances/", {"site": self.site.id, "date": "2026-05-16"})
        schedule = GuardSchedule.objects.get()

        self.client.post(
            "/attendances/",
            {
                "site": str(self.site.id),
                "date": "2026-05-16",
                "schedule_ids": [str(schedule.id)],
                f"scheduled_guard_{schedule.id}": str(self.guard.id),
                f"replacement_guard_{schedule.id}": str(self.replacement_guard.id),
                f"reason_{schedule.id}": "Scheduled guard called in sick",
            },
        )

        replacement_salary = Salary.objects.get(
            employee=self.replacement_guard,
            pay_period_start="2026-05-01",
        )
        self.assertEqual(replacement_salary.attendance_days, 1)
        self.assertFalse(
            Salary.objects.filter(employee=self.guard, pay_period_start="2026-05-01").exists()
        )

    def test_attendance_report_filters_replacement_guard_as_attended_employee(self):
        self.client.get("/attendances/", {"site": self.site.id, "date": "2026-05-16"})
        schedule = GuardSchedule.objects.get()
        self.client.post(
            "/attendances/",
            {
                "site": str(self.site.id),
                "date": "2026-05-16",
                "schedule_ids": [str(schedule.id)],
                f"scheduled_guard_{schedule.id}": str(self.guard.id),
                f"replacement_guard_{schedule.id}": str(self.replacement_guard.id),
                f"reason_{schedule.id}": "Scheduled guard called in sick",
            },
        )

        response = self.client.get(
            "/reports/attendance/",
            {
                "employee_number": self.replacement_guard.company_number,
                "start_date": "2026-05-01",
                "end_date": "2026-05-31",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, f"{self.replacement_guard.company_number}-{self.replacement_guard.full_name}")
        self.assertContains(response, f"{self.guard.company_number}-{self.guard.full_name}")
        self.assertNotContains(response, "No attendance records found.")

    def test_guard_schedule_form_routes_back_to_attendance_screen(self):
        response = self.client.get("/records/guard-schedules/add/")

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, "/attendances/")

    def test_excel_duty_roster_upload_creates_schedule(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["company_number", "site_code", "shift_name", "shift_date"])
        worksheet.append([self.guard.company_number, self.site.site_code, self.shift.shift_name, "2026-05-20"])
        content = BytesIO()
        workbook.save(content)
        upload = SimpleUploadedFile(
            "duty-roster.xlsx",
            content.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post("/attendances/upload-roster/", {"roster_file": upload})

        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            GuardSchedule.objects.filter(
                employee=self.guard,
                site=self.site,
                shift=self.shift,
                shift_date="2026-05-20",
            ).exists()
        )
        roster_row = RosterAttendance.objects.get(shift_date="2026-05-20")
        self.assertEqual(roster_row.employee, self.guard)
        self.assertEqual(roster_row.import_status, RosterAttendance.ImportStatus.CREATED)
        self.assertIsNotNone(roster_row.schedule)

    def test_excel_duty_roster_does_not_exceed_contract_requirement(self):
        ContractSiteRequirement.objects.filter(site=self.site).update(required_guards=1)
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["company_number", "site_code", "shift_name", "shift_date"])
        worksheet.append([self.guard.company_number, self.site.site_code, self.shift.shift_name, "2026-05-21"])
        worksheet.append([self.replacement_guard.company_number, self.site.site_code, self.shift.shift_name, "2026-05-21"])
        content = BytesIO()
        workbook.save(content)
        upload = SimpleUploadedFile(
            "limited-duty-roster.xlsx",
            content.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post("/attendances/upload-roster/", {"roster_file": upload})

        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            GuardSchedule.objects.filter(site=self.site, shift=self.shift, shift_date="2026-05-21").count(),
            1,
        )
        self.assertEqual(RosterAttendance.objects.filter(shift_date="2026-05-21").count(), 2)
        self.assertTrue(
            RosterAttendance.objects.filter(
                shift_date="2026-05-21",
                import_status=RosterAttendance.ImportStatus.SKIPPED,
            ).exists()
        )

    def test_duty_roster_template_downloads_excel_workbook(self):
        response = self.client.get("/attendances/roster-template/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook.active
        self.assertEqual(
            [cell.value for cell in worksheet[1]],
            ["site_code", "site_name", "shift", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"],
        )

    def test_roster_attendance_summary_pdf_downloads(self):
        self.client.get("/attendances/", {"site": self.site.id, "date": "2026-05-16"})
        schedule = GuardSchedule.objects.get()
        self.client.post(
            "/attendances/",
            {
                "site": str(self.site.id),
                "date": "2026-05-16",
                "schedule_ids": [str(schedule.id)],
                f"scheduled_guard_{schedule.id}": str(self.guard.id),
                f"present_{schedule.id}": "yes",
            },
        )

        response = self.client.get("/attendances/summary/pdf/", {"month": "2026-05"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertIn("attendance-summary-2026-05.pdf", response["Content-Disposition"])
        self.assertTrue(response.content.startswith(b"%PDF"))

    def test_roster_pages_link_to_attendance_summary_pdf(self):
        attendance_response = self.client.get("/attendances/", {"site": self.site.id, "date": "2026-05-16"})
        upload_response = self.client.get("/attendances/upload-roster/")

        self.assertContains(attendance_response, "PDF Summary")
        self.assertContains(attendance_response, "/attendances/summary/pdf/?month=2026-05")
        self.assertContains(upload_response, "Attendance Summary PDF")
        self.assertContains(upload_response, "/attendances/summary/pdf/")

    def test_wide_monthly_roster_stores_o_as_off_attendance(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append([])
        worksheet.append([])
        worksheet.append(["site code", "site name", "shift", "Mon", "Tue", "Wed"])
        worksheet.append([None, None, "D", 1, 2, 3])
        worksheet.append(["S001", "Wide Monthly Site", "D", "O", "D", "N"])
        content = BytesIO()
        workbook.save(content)
        upload = SimpleUploadedFile(
            "DUTY ROASTER.xlsx",
            content.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post(
            "/attendances/upload-roster/",
            {"roster_file": upload, "roster_month": "2026-05"},
        )

        self.assertEqual(response.status_code, 302)
        off_row = RosterAttendance.objects.get(duty_code="O")
        self.assertEqual(off_row.import_status, RosterAttendance.ImportStatus.OFF)
        self.assertEqual(off_row.shift_date.isoformat(), "2026-05-01")
        self.assertEqual(off_row.site.site_code, "S001")
        self.assertEqual(off_row.site.site_name, "Wide Monthly Site")
        self.assertTrue(
            RosterAttendance.objects.filter(
                duty_code="D",
                import_status=RosterAttendance.ImportStatus.CREATED,
                shift_date="2026-05-02",
            ).exists()
        )
        self.assertEqual(GuardSchedule.objects.filter(shift_date__range=("2026-05-01", "2026-05-03")).count(), 0)

    def test_guard_schedule_model_rejects_more_than_required_guards(self):
        ContractSiteRequirement.objects.filter(site=self.site).update(required_guards=1)
        second_deployment = Deployment.objects.create(
            employee=self.replacement_guard,
            client=self.client_record,
            site=self.site,
            shift=self.shift,
            start_date="2026-05-22",
        )
        GuardSchedule.objects.create(
            deployment=self.deployment,
            employee=self.guard,
            site=self.site,
            shift=self.shift,
            shift_date="2026-05-22",
        )

        with self.assertRaises(ValidationError):
            GuardSchedule.objects.create(
                deployment=second_deployment,
                employee=self.replacement_guard,
                site=self.site,
                shift=self.shift,
                shift_date="2026-05-22",
            )

    def test_cancelled_guard_schedule_does_not_count_against_required_guards(self):
        ContractSiteRequirement.objects.filter(site=self.site).update(required_guards=1)
        second_deployment = Deployment.objects.create(
            employee=self.replacement_guard,
            client=self.client_record,
            site=self.site,
            shift=self.shift,
            start_date="2026-05-23",
        )
        GuardSchedule.objects.create(
            deployment=self.deployment,
            employee=self.guard,
            site=self.site,
            shift=self.shift,
            shift_date="2026-05-23",
            status=GuardSchedule.ScheduleStatus.CANCELLED,
        )

        GuardSchedule.objects.create(
            deployment=second_deployment,
            employee=self.replacement_guard,
            site=self.site,
            shift=self.shift,
            shift_date="2026-05-23",
        )

        active_schedules = GuardSchedule.objects.filter(
            site=self.site,
            shift=self.shift,
            shift_date="2026-05-23",
        ).exclude(status=GuardSchedule.ScheduleStatus.CANCELLED)
        self.assertEqual(active_schedules.count(), 1)

    def test_contract_requirement_is_created_from_existing_site_data(self):
        self.assertTrue(ContractSiteRequirement.objects.filter(site=self.site, required_guards=10).exists())

    def test_contract_list_displays_required_guards(self):
        response = self.client.get("/records/contracts/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Required Guards")
        self.assertContains(response, "10")

    def test_employee_list_displays_employee_identifiers_without_extra_guard_columns(self):
        response = self.client.get("/records/employees/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Company Number")
        self.assertContains(response, "NSSF Number")
        self.assertContains(response, "Full Name")
        self.assertContains(response, "National ID")
        content = response.content.decode()
        self.assertLess(content.index("Phone Number"), content.index("Date Of Birth"))
        self.assertLess(content.index("Date Of Birth"), content.index("Gender"))
        self.assertLess(content.index("Gender"), content.index("Email"))
        self.assertLess(content.index("National ID"), content.index("NSSF Number"))
        self.assertLess(content.index("NSSF Number"), content.index("Hire Date"))
        self.assertNotContains(response, "Uniform Size")
        self.assertNotContains(response, "Armed Status")
        self.assertNotContains(response, "Licence Number")
        self.assertNotContains(response, "Authority Level")
        self.assertNotContains(response, "First Name")
        self.assertNotContains(response, "Last Name")

    def test_employee_form_hides_removed_guard_fields(self):
        response = self.client.get("/records/employees/add/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "NSSF number")
        self.assertNotContains(response, "Uniform size")
        self.assertNotContains(response, "Armed status")
        self.assertNotContains(response, "License no")
        self.assertNotContains(response, "Authority level")

    def test_payroll_is_generated_from_present_attendance(self):
        position = Position.objects.create(
            position_title="Payroll Guard",
            department=DepartmentChoices.OPERATIONS,
            salary_range_min=Decimal("416000.00"),
            salary_range_max=Decimal("416000.00"),
        )
        self.guard.position = position
        self.guard.bank_account = "010000123456"
        self.guard.save(update_fields=["position", "bank_account", "updated_at"])
        Attendance.objects.create(
            employee=self.guard,
            shift=self.shift,
            date="2026-05-16",
            status="Present",
        )

        response = self.client.post("/payroll/", {"month": "2026-05"})

        self.assertEqual(response.status_code, 302)
        salary = Salary.objects.get(employee=self.guard, pay_period_start="2026-05-01")
        self.assertEqual(salary.attendance_days, 1)
        self.assertEqual(salary.basic_hours, Decimal("8.00"))
        self.assertEqual(salary.overtime_hours, Decimal("4.00"))
        self.assertEqual(salary.basic_salary, Decimal("16000.00"))
        self.assertEqual(salary.overtime_pay, Decimal("12000.00"))
        self.assertEqual(salary.gross_pay, Decimal("28000.00"))
        self.assertEqual(salary.nssf_employee, Decimal("1400.00"))
        self.assertEqual(salary.nssf_employer, Decimal("2800.00"))
        self.assertEqual(salary.total_deductions, Decimal("1400.00"))
        self.assertEqual(salary.net_salary, Decimal("26600.00"))

    def test_payroll_refreshes_when_attendance_increases(self):
        position = Position.objects.create(
            position_title="Attendance Linked Payroll Guard",
            department=DepartmentChoices.OPERATIONS,
            salary_range_min=Decimal("416000.00"),
            salary_range_max=Decimal("416000.00"),
        )
        self.guard.position = position
        self.guard.save(update_fields=["position", "updated_at"])
        Attendance.objects.create(employee=self.guard, shift=self.shift, date="2026-05-16", status="Present")
        self.client.get("/payroll/", {"month": "2026-05"})
        salary = Salary.objects.get(employee=self.guard, pay_period_start="2026-05-01")
        self.assertEqual(salary.attendance_days, 1)
        self.assertEqual(salary.gross_pay, Decimal("28000.00"))

        Attendance.objects.create(employee=self.guard, shift=self.shift, date="2026-05-17", status="Present")
        self.client.get("/payroll/", {"month": "2026-05"})
        salary.refresh_from_db()

        self.assertEqual(salary.attendance_days, 2)
        self.assertEqual(salary.basic_hours, Decimal("16.00"))
        self.assertEqual(salary.overtime_hours, Decimal("8.00"))
        self.assertEqual(salary.gross_pay, Decimal("56000.00"))
        self.assertEqual(salary.net_salary, Decimal("53200.00"))

    def test_payroll_refresh_resets_days_when_attendance_is_no_longer_present(self):
        position = Position.objects.create(
            position_title="Attendance Reset Payroll Guard",
            department=DepartmentChoices.OPERATIONS,
            salary_range_min=Decimal("416000.00"),
            salary_range_max=Decimal("416000.00"),
        )
        self.guard.position = position
        self.guard.save(update_fields=["position", "updated_at"])
        attendance = Attendance.objects.create(
            employee=self.guard,
            site=self.site,
            shift=self.shift,
            date="2026-05-16",
            status="Present",
        )
        self.client.get("/payroll/", {"month": "2026-05"})
        salary = Salary.objects.get(employee=self.guard, pay_period_start="2026-05-01")
        self.assertEqual(salary.attendance_days, 1)

        attendance.status = "Absent"
        attendance.save(update_fields=["status", "updated_at"])
        self.client.get("/payroll/", {"month": "2026-05"})
        salary.refresh_from_db()

        self.assertEqual(salary.attendance_days, 0)
        self.assertEqual(salary.basic_hours, Decimal("0.00"))
        self.assertEqual(salary.overtime_hours, Decimal("0.00"))
        self.assertEqual(salary.gross_pay, Decimal("0.00"))

    def test_marking_attendance_increments_payroll_days(self):
        position = Position.objects.create(
            position_title="Marked Attendance Payroll Guard",
            department=DepartmentChoices.OPERATIONS,
            salary_range_min=Decimal("416000.00"),
            salary_range_max=Decimal("416000.00"),
        )
        self.guard.position = position
        self.guard.save(update_fields=["position", "updated_at"])

        for day in ("2026-05-16", "2026-05-17"):
            self.client.get("/attendances/", {"site": self.site.id, "date": day})
            schedule = GuardSchedule.objects.get(shift_date=day)
            self.client.post(
                "/attendances/",
                {
                    "site": str(self.site.id),
                    "date": day,
                    "schedule_ids": [str(schedule.id)],
                    f"scheduled_guard_{schedule.id}": str(self.guard.id),
                    f"present_{schedule.id}": "yes",
                },
            )

        salary = Salary.objects.get(employee=self.guard, pay_period_start="2026-05-01")
        self.assertEqual(salary.attendance_days, 2)
        self.assertEqual(salary.basic_hours, Decimal("16.00"))
        self.assertEqual(salary.overtime_hours, Decimal("8.00"))

    def test_marking_attendance_counts_full_month_not_only_first_two_days(self):
        position = Position.objects.create(
            position_title="Full Month Attendance Payroll Guard",
            department=DepartmentChoices.OPERATIONS,
            salary_range_min=Decimal("416000.00"),
            salary_range_max=Decimal("416000.00"),
        )
        self.guard.position = position
        self.guard.save(update_fields=["position", "updated_at"])

        for day in ("2026-05-16", "2026-05-17", "2026-05-18", "2026-05-19"):
            self.client.get("/attendances/", {"site": self.site.id, "date": day})
            schedule = GuardSchedule.objects.get(shift_date=day)
            self.client.post(
                "/attendances/",
                {
                    "site": str(self.site.id),
                    "date": day,
                    "schedule_ids": [str(schedule.id)],
                    f"scheduled_guard_{schedule.id}": str(self.guard.id),
                    f"present_{schedule.id}": "yes",
                },
            )

        salary = Salary.objects.get(employee=self.guard, pay_period_start="2026-05-01")
        self.assertEqual(salary.attendance_days, 4)
        self.assertEqual(salary.basic_hours, Decimal("32.00"))
        self.assertEqual(salary.overtime_hours, Decimal("16.00"))

    def test_payroll_days_count_each_marked_shift_on_same_date(self):
        position = Position.objects.create(
            position_title="Same Date Shift Payroll Guard",
            department=DepartmentChoices.OPERATIONS,
            salary_range_min=Decimal("416000.00"),
            salary_range_max=Decimal("416000.00"),
        )
        self.guard.position = position
        self.guard.save(update_fields=["position", "updated_at"])
        night_shift, _created = Shift.objects.update_or_create(
            code="N",
            defaults={
                "shift_name": "Night",
                "start_time": "20:00",
                "end_time": "08:00",
            },
        )
        night_deployment = Deployment.objects.create(
            employee=self.guard,
            client=self.client_record,
            site=self.site,
            shift=night_shift,
            start_date="2026-05-01",
        )
        first_schedule = GuardSchedule.objects.create(
            deployment=self.deployment,
            employee=self.guard,
            site=self.site,
            shift=self.shift,
            shift_date="2026-05-16",
        )
        second_schedule = GuardSchedule.objects.create(
            deployment=night_deployment,
            employee=self.guard,
            site=self.site,
            shift=night_shift,
            shift_date="2026-05-16",
        )

        self.client.post(
            "/attendances/",
            {
                "site": str(self.site.id),
                "date": "2026-05-16",
                "schedule_ids": [str(first_schedule.id), str(second_schedule.id)],
                f"scheduled_guard_{first_schedule.id}": str(self.guard.id),
                f"present_{first_schedule.id}": "yes",
                f"scheduled_guard_{second_schedule.id}": str(self.guard.id),
                f"present_{second_schedule.id}": "yes",
            },
        )

        salary = Salary.objects.get(employee=self.guard, pay_period_start="2026-05-01")
        self.assertEqual(salary.attendance_days, 2)
        self.assertEqual(salary.basic_hours, Decimal("16.00"))
        self.assertEqual(salary.overtime_hours, Decimal("8.00"))

    def test_payroll_days_count_same_guard_shift_at_different_sites(self):
        position = Position.objects.create(
            position_title="Multi Site Payroll Guard",
            department=DepartmentChoices.OPERATIONS,
            salary_range_min=Decimal("416000.00"),
            salary_range_max=Decimal("416000.00"),
        )
        self.guard.position = position
        self.guard.save(update_fields=["position", "updated_at"])
        second_site = Site.objects.create(
            client=self.client_record,
            site_name="Warehouse Gate",
            site_address="Plot 2",
            city="Kampala",
        )
        second_deployment = Deployment.objects.create(
            employee=self.guard,
            client=self.client_record,
            site=second_site,
            shift=self.shift,
            start_date="2026-05-01",
        )
        first_schedule = GuardSchedule.objects.create(
            deployment=self.deployment,
            employee=self.guard,
            site=self.site,
            shift=self.shift,
            shift_date="2026-05-16",
        )
        second_schedule = GuardSchedule.objects.create(
            deployment=second_deployment,
            employee=self.guard,
            site=second_site,
            shift=self.shift,
            shift_date="2026-05-16",
        )

        self.client.post(
            "/attendances/",
            {
                "site": "",
                "date": "2026-05-16",
                "schedule_ids": [str(first_schedule.id), str(second_schedule.id)],
                f"scheduled_guard_{first_schedule.id}": str(self.guard.id),
                f"present_{first_schedule.id}": "yes",
                f"scheduled_guard_{second_schedule.id}": str(self.guard.id),
                f"present_{second_schedule.id}": "yes",
            },
        )

        salary = Salary.objects.get(employee=self.guard, pay_period_start="2026-05-01")
        self.assertEqual(Attendance.objects.filter(employee=self.guard, date="2026-05-16", shift=self.shift).count(), 2)
        self.assertEqual(salary.attendance_days, 2)
        self.assertEqual(salary.basic_hours, Decimal("16.00"))
        self.assertEqual(salary.overtime_hours, Decimal("8.00"))

    def test_payroll_exports_and_payslip_are_downloadable(self):
        position = Position.objects.create(
            position_title="Payroll Export Guard",
            department=DepartmentChoices.OPERATIONS,
            salary_range_min=Decimal("416000.00"),
            salary_range_max=Decimal("416000.00"),
        )
        self.guard.position = position
        self.guard.bank_account = "010000123456"
        self.guard.save(update_fields=["position", "bank_account", "updated_at"])
        Attendance.objects.create(
            employee=self.guard,
            shift=self.shift,
            date="2026-05-16",
            status="Present",
        )
        self.client.post("/payroll/", {"month": "2026-05"})
        salary = Salary.objects.get(employee=self.guard, pay_period_start="2026-05-01")

        excel_response = self.client.get("/payroll/export/excel/", {"month": "2026-05"})
        self.assertEqual(excel_response.status_code, 200)
        self.assertEqual(
            excel_response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(excel_response.content))
        worksheet = workbook.active
        self.assertEqual(worksheet["A1"].value, "Payroll Register: 2026-05-01 to 2026-05-31")
        self.assertEqual(worksheet["A3"].value, "Employee")
        self.assertEqual(worksheet["D3"].value, "Bank Account")
        self.assertEqual(worksheet["D4"].value, "010000123456")

        pdf_response = self.client.get("/payroll/export/pdf/", {"month": "2026-05"})
        self.assertEqual(pdf_response.status_code, 200)
        self.assertEqual(pdf_response["Content-Type"], "application/pdf")
        self.assertTrue(pdf_response.content.startswith(b"%PDF"))

        payslip_response = self.client.get(f"/payroll/{salary.pk}/payslip/")
        self.assertEqual(payslip_response.status_code, 200)
        self.assertEqual(payslip_response["Content-Type"], "application/pdf")
        self.assertTrue(payslip_response.content.startswith(b"%PDF"))

        page_response = self.client.get("/payroll/", {"month": "2026-05"})
        self.assertContains(page_response, "Bank Account")
        self.assertContains(page_response, "Total Deductions")
        self.assertContains(page_response, "010000123456")

    def test_training_register_tracks_professional_certification_details(self):
        training = Training.objects.create(
            employee=self.guard,
            training_name="Basic Guarding and Client Conduct",
            course_code="SEC-101",
            training_type=Training.TrainingType.INDUCTION,
            training_objective="Confirm guard readiness before site deployment.",
            provider="Sentinel Training Academy",
            trainer_name="Senior Instructor",
            trainer_contact="0700000999",
            venue="Head Office Training Room",
            start_date="2026-05-01",
            end_date="2026-05-03",
            duration_hours=Decimal("18.00"),
            training_cost=Decimal("250000.00"),
            pass_mark=70,
            score=Decimal("82.50"),
            result=Training.TrainingResult.PASSED,
            certificate_no="CERT-SEC-101-001",
            expiry_date="2027-05-03",
            next_refresh_date="2027-04-03",
            status=StatusChoices.APPROVED,
            action_notes="Cleared for deployment.",
        )

        training.refresh_from_db()
        self.assertTrue(training.is_certificate_current)
        response = self.client.get("/records/trainings/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Training Type")
        self.assertContains(response, "Senior Instructor")
        self.assertContains(response, "CERT-SEC-101-001")

    def test_hr_recruitment_tracks_physical_and_online_applications(self):
        requisition = RecruitmentRequisition.objects.create(
            requisition_number="REQ-2026-001",
            vacancy_title="Security Guard",
            position=self.guard.position,
            department=DepartmentChoices.OPERATIONS,
            requested_by=self.supervisor,
            number_of_openings=5,
            employment_type=RecruitmentRequisition.EmploymentType.FULL_TIME,
            work_location="Kampala",
            opening_date="2026-05-01",
            closing_date="2026-05-31",
            minimum_qualification="UCE",
            experience_required="One year security experience",
            status=RecruitmentRequisition.RequisitionStatus.OPEN,
        )
        physical_application = RecruitmentApplication.objects.create(
            requisition=requisition,
            first_name="Physical",
            last_name="Applicant",
            phone_number="0777000001",
            application_source=RecruitmentApplication.ApplicationSource.PHYSICAL,
            date_received="2026-05-02",
            highest_qualification="UCE",
            screening_score=76,
            status=RecruitmentApplication.ApplicationStatus.SHORTLISTED,
        )
        online_application = RecruitmentApplication.objects.create(
            requisition=requisition,
            first_name="Online",
            last_name="Applicant",
            phone_number="0777000002",
            email="online.applicant@example.com",
            application_source=RecruitmentApplication.ApplicationSource.ONLINE,
            online_profile_url="https://jobs.example.test/applications/1",
            date_received="2026-05-03",
            highest_qualification="UACE",
            screening_score=82,
            status=RecruitmentApplication.ApplicationStatus.INTERVIEW,
        )
        RecruitmentInterview.objects.create(
            application=online_application,
            interview_type=RecruitmentInterview.InterviewType.ONLINE,
            scheduled_at=timezone.now(),
            venue_or_link="https://meet.example.test/interview",
            interviewer=self.supervisor,
            score=84,
            recommendation=RecruitmentInterview.InterviewRecommendation.RECOMMENDED,
            status=StatusChoices.APPROVED,
        )

        self.assertEqual(requisition.applications_count, 2)
        self.assertEqual(physical_application.full_name, "Physical Applicant")
        response = self.client.get("/records/recruitment-applications/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Application Source")
        self.assertContains(response, "Physical Applicant")
        self.assertContains(response, "Online Applicant")
        self.assertContains(response, "Online Application")

    def test_audit_report_and_dashboard_include_recruitment_and_training_value(self):
        requisition = RecruitmentRequisition.objects.create(
            requisition_number="REQ-2026-AUDIT",
            vacancy_title="Audit Guard",
            department=DepartmentChoices.OPERATIONS,
            requested_by=self.supervisor,
            number_of_openings=2,
            opening_date="2026-05-01",
            closing_date="2026-05-31",
            recruitment_budget=Decimal("1000000.00"),
            actual_recruitment_cost=Decimal("850000.00"),
            status=RecruitmentRequisition.RequisitionStatus.OPEN,
        )
        application = RecruitmentApplication.objects.create(
            requisition=requisition,
            first_name="Audit",
            last_name="Hire",
            phone_number="0777000003",
            application_source=RecruitmentApplication.ApplicationSource.ONLINE,
            date_received="2026-05-04",
            screening_score=88,
            status=RecruitmentApplication.ApplicationStatus.HIRED,
        )
        JobOffer.objects.create(
            application=application,
            offer_date="2026-05-05",
            salary_offer=Decimal("600000.00"),
            status=JobOffer.OfferStatus.ACCEPTED,
            accepted_date="2026-05-06",
        )
        Training.objects.create(
            employee=self.guard,
            training_name="Audit Training",
            provider="Sentinel Academy",
            start_date="2026-05-07",
            end_date="2026-05-08",
            duration_hours=Decimal("12.00"),
            budgeted_cost=Decimal("300000.00"),
            training_cost=Decimal("250000.00"),
            result=Training.TrainingResult.PASSED,
            status=StatusChoices.APPROVED,
        )

        audit_response = self.client.get("/audit/", {"month": "2026-05"})

        self.assertEqual(audit_response.status_code, 200)
        self.assertContains(audit_response, "Recruitment")
        self.assertContains(audit_response, "Training")
        self.assertContains(audit_response, "Benefiting")
        self.assertContains(audit_response, "Recruitment Spend")
        self.assertContains(audit_response, "Training Spend")

        dashboard_response = self.client.get("/dashboard/")

        self.assertEqual(dashboard_response.status_code, 200)
        self.assertContains(dashboard_response, "Recruitment applications")
        self.assertContains(dashboard_response, "Successful trainings")

    def test_saracen_style_duty_roster_upload_creates_schedules(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["Saracen Uganda Limited Schedule for period : "])
        worksheet.append([])
        worksheet.append(["2022-12-12 07:04 Saracen International Site Roster for ST03956:Hima Factory"])
        worksheet.append([])
        worksheet.append(["Deployment Area: Fort Portal"])
        worksheet.append([])
        worksheet.append(["Scheduled Period: 2022-12-26 to 2023-01-25"])
        worksheet.append([])
        worksheet.append(["Pers No ", "Grade ", "Name ", "Contact ", "Worked days ", "Mo/26", "Tu/27", "We/28"])
        worksheet.append(["1001", "", "Test Guard", "", "", "D", "O", "N"])
        content = BytesIO()
        workbook.save(content)
        upload = SimpleUploadedFile(
            "saracen-roster.xlsx",
            content.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post("/attendances/upload-roster/", {"roster_file": upload})

        self.assertEqual(response.status_code, 302)
        self.assertTrue(Site.objects.filter(site_code="ST03956", site_name="Hima Factory").exists())
        self.assertEqual(GuardSchedule.objects.filter(notes__icontains="Saracen").count(), 2)

    def test_zone_shift_summary_shows_marked_guards(self):
        self.client.get("/attendances/", {"site": self.site.id, "date": "2026-05-16"})
        schedule = GuardSchedule.objects.get()
        self.client.post(
            "/attendances/",
            {
                "site": str(self.site.id),
                "date": "2026-05-16",
                "schedule_ids": [str(schedule.id)],
                f"scheduled_guard_{schedule.id}": str(self.guard.id),
                f"present_{schedule.id}": "yes",
            },
        )

        response = self.client.get("/reports/zone-shift-summary/", {"date": "2026-05-16"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Test Zone")
        self.assertContains(response, "Day")
        self.assertContains(response, "Brian Kato")

    def test_attendance_report_matches_query_layout(self):
        self.client.get("/attendances/", {"site": self.site.id, "date": "2026-05-16"})
        schedule = GuardSchedule.objects.get()
        self.client.post(
            "/attendances/",
            {
                "site": str(self.site.id),
                "date": "2026-05-16",
                "schedule_ids": [str(schedule.id)],
                f"scheduled_guard_{schedule.id}": str(self.guard.id),
                f"present_{schedule.id}": "yes",
            },
        )

        response = self.client.get(
            "/reports/attendance/",
            {"start_date": "2026-05-01", "end_date": "2026-05-31"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "attendance report")
        self.assertContains(response, "Employee Number")
        self.assertContains(response, "Site scheduled")
        self.assertContains(response, "Brian Kato")


class SiteCodeTests(TestCase):
    def test_site_code_is_generated_from_client_name(self):
        client = Client.objects.create(
            client_name="Central Plaza",
            contact_person="Sarah",
            phone_number="0700000001",
            contract_start_date="2026-01-01",
        )

        first_site = Site.objects.create(
            client=client,
            site_name="Main Gate",
            site_address="Plot 1",
            city="Kampala",
        )
        second_site = Site.objects.create(
            client=client,
            site_name="Parking Yard",
            site_address="Plot 1",
            city="Kampala",
        )

        self.assertEqual(first_site.site_code, "CPXXS0001")
        self.assertEqual(second_site.site_code, "CPXXS0002")


class ResponsiveCrudPageTests(TestCase):
    @override_settings(ALLOWED_HOSTS=["testserver"])
    def test_site_add_page_loads(self):
        user = User.objects.create_user(username="hr", password="pass")
        group, _ = Group.objects.get_or_create(name="Human Resources Manager")
        user.groups.add(group)
        self.client.login(username="hr", password="pass")

        response = self.client.get("/records/sites/add/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Add Sites")


class ZoneAuthorizationTests(TestCase):
    def test_guard_cannot_have_two_active_zone_allocations(self):
        client_record = Client.objects.create(
            client_name="Zone Client",
            contact_person="Sarah",
            phone_number="0700000001",
            contract_start_date="2026-01-01",
        )
        role = Role.objects.create(role_name="Zone Role", department=DepartmentChoices.OPERATIONS)
        supervisor_employee = Employee.objects.create(
            first_name="Susan",
            last_name="Nabirye",
            phone_number="0700000003",
            email="susan@example.com",
            national_id="NIN003",
            role=role,
        )
        supervisor = supervisor_employee
        guard_employee = Employee.objects.create(
            first_name="Paul",
            last_name="Mugerwa",
            phone_number="0700000004",
            email="paul@example.com",
            national_id="NIN004",
            role=role,
            company_number="G-002",
        )
        guard = guard_employee
        first_zone = Zone.objects.create(zone_code="ZN-001", zone_name="North", supervisor=supervisor)
        second_zone = Zone.objects.create(zone_code="ZN-002", zone_name="South", supervisor=supervisor)

        ZoneEmployeeAllocation.objects.create(zone=first_zone, employee=guard)

        with self.assertRaises(IntegrityError):
            ZoneEmployeeAllocation.objects.create(zone=second_zone, employee=guard)

    def test_supervisor_cannot_access_client_management(self):
        user = User.objects.create_user(username="supervisor", password="pass")
        group, _ = Group.objects.get_or_create(name="Supervisor")
        user.groups.add(group)
        self.client.login(username="supervisor", password="pass")

        response = self.client.get("/records/clients/")

        self.assertEqual(response.status_code, 403)

    def test_hr_manager_can_access_client_management(self):
        user = User.objects.create_user(username="hr2", password="pass")
        group, _ = Group.objects.get_or_create(name="Human Resources Manager")
        user.groups.add(group)
        self.client.login(username="hr2", password="pass")

        response = self.client.get("/records/clients/")

        self.assertEqual(response.status_code, 200)

    def test_user_with_direct_permission_can_login_and_view_permitted_module(self):
        user = User.objects.create_user(username="client-viewer", password="pass")
        permission = Permission.objects.get(codename="view_client")
        user.user_permissions.add(permission)
        self.client.login(username="client-viewer", password="pass")

        dashboard_response = self.client.get("/dashboard/")
        records_response = self.client.get("/records/clients/")

        self.assertEqual(dashboard_response.status_code, 200)
        self.assertEqual(records_response.status_code, 200)

    def test_manager_can_access_every_module(self):
        user = User.objects.create_user(username="manager", password="pass")
        group, _ = Group.objects.get_or_create(name="Manager")
        user.groups.add(group)
        self.client.login(username="manager", password="pass")

        for slug in MODEL_REGISTRY:
            with self.subTest(slug=slug):
                response = self.client.get(f"/records/{slug}/")
                self.assertEqual(response.status_code, 200)

    def test_manager_sidebar_displays_every_module(self):
        user = User.objects.create_user(username="sidebar-manager", password="pass")
        group, _ = Group.objects.get_or_create(name="Manager")
        user.groups.add(group)
        self.client.login(username="sidebar-manager", password="pass")

        response = self.client.get("/dashboard/")

        self.assertEqual(response.status_code, 200)
        for config in MODEL_REGISTRY.values():
            with self.subTest(title=config.title):
                self.assertContains(response, config.title)

    def test_sidebar_hides_redundant_workflow_tables(self):
        user = User.objects.create_user(username="clean-sidebar-manager", password="pass")
        group, _ = Group.objects.get_or_create(name="Manager")
        user.groups.add(group)
        self.client.login(username="clean-sidebar-manager", password="pass")

        response = self.client.get("/dashboard/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Attendances")
        self.assertContains(response, "Attendance Report")
        self.assertContains(response, "Zonal Employees")
        self.assertContains(response, "Asset Report")
        self.assertNotContains(response, '<span class="menu-title">Assets</span>', html=True)
        self.assertNotContains(response, '<span class="menu-title">Guard Schedules</span>', html=True)
        self.assertNotContains(response, '<span class="menu-title">Attendance Records</span>', html=True)
        self.assertNotContains(response, '<span class="menu-title">Zone Employee Allocations</span>', html=True)
        self.assertNotContains(response, '<span class="menu-title">Zone Site Allocations</span>', html=True)

    def test_supervisor_sidebar_hides_restricted_modules(self):
        user = User.objects.create_user(username="sidebar-supervisor", password="pass")
        group, _ = Group.objects.get_or_create(name="Supervisor")
        user.groups.add(group)
        self.client.login(username="sidebar-supervisor", password="pass")

        response = self.client.get("/dashboard/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Sites")
        self.assertNotContains(response, 'href="/records/clients/"')

    def test_attendance_supervisor_is_limited_to_attendance_workflows(self):
        user = User.objects.create_user(username="attendance-supervisor", password="pass")
        group, _ = Group.objects.get_or_create(name="Attendance Supervisor")
        user.groups.add(group)
        self.client.login(username="attendance-supervisor", password="pass")

        allowed_paths = [
            "/attendances/",
            "/attendances/upload-roster/",
            "/attendances/summary/pdf/?month=2026-05",
            "/reports/attendance/",
            "/records/attendance-records/",
            "/records/roster-attendances/",
        ]
        blocked_paths = [
            "/records/assets/",
            "/records/incidents/",
            "/records/clients/",
            "/reports/assets/",
            "/payroll/",
        ]

        for path in allowed_paths:
            with self.subTest(path=path):
                response = self.client.get(path)
                self.assertNotEqual(response.status_code, 403)
        for path in blocked_paths:
            with self.subTest(path=path):
                response = self.client.get(path)
                self.assertIn(response.status_code, (302, 403))

    def test_limited_operations_manager_can_access_asset_attendance_and_incident_workflows(self):
        user = User.objects.create_user(username="limited-ops-manager", password="pass")
        group, _ = Group.objects.get_or_create(name="Operations Limited Manager")
        user.groups.add(group)
        self.client.login(username="limited-ops-manager", password="pass")

        allowed_paths = [
            "/attendances/",
            "/records/assets/",
            "/records/incidents/",
            "/records/attendance-records/",
            "/reports/assets/",
            "/reports/attendance/",
        ]
        blocked_paths = [
            "/records/clients/",
            "/records/payments/",
            "/payroll/",
            "/audit/",
        ]

        for path in allowed_paths:
            with self.subTest(path=path):
                response = self.client.get(path)
                self.assertNotEqual(response.status_code, 403)
        for path in blocked_paths:
            with self.subTest(path=path):
                response = self.client.get(path)
                self.assertIn(response.status_code, (302, 403))

    def test_sidebar_system_admin_link_is_visible_only_to_superusers(self):
        limited_user = User.objects.create_user(username="limited-sidebar", password="pass", is_staff=True)
        group, _ = Group.objects.get_or_create(name="Operations Limited Manager")
        limited_user.groups.add(group)
        self.client.login(username="limited-sidebar", password="pass")

        limited_response = self.client.get("/dashboard/")

        self.assertEqual(limited_response.status_code, 200)
        self.assertNotContains(limited_response, "System Admin")
        self.assertNotContains(limited_response, 'href="/admin/"')

        self.client.logout()
        admin_user = User.objects.create_superuser(username="system-admin", password="pass", email="admin@example.com")
        self.client.login(username="system-admin", password="pass")

        admin_response = self.client.get("/dashboard/")

        self.assertEqual(admin_response.status_code, 200)
        self.assertContains(admin_response, "System Admin")
        self.assertContains(admin_response, 'href="/admin/"')
