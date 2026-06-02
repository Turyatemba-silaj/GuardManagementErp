import calendar
from collections import defaultdict
import csv
from datetime import timedelta
from decimal import Decimal
from io import BytesIO
import json
import os
import re
import uuid
import zipfile
from xml.sax.saxutils import escape as xml_escape

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.exceptions import ValidationError
from django.db import DatabaseError, connection
from django.db import transaction
from django.db.models import Count, ProtectedError, Sum
from django.db.models.functions import Coalesce
from django.forms import DateInput, DateTimeInput, FileInput, TimeInput, modelform_factory
from django.http import Http404, HttpResponse, HttpResponseBadRequest, HttpResponseForbidden, JsonResponse
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.dateparse import parse_date, parse_datetime
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook, load_workbook
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from . import models
from .access import can_access_internal, can_manage_attendance, can_manage_slug, has_model_perm, is_manager, is_supervisor
from .accounting import ensure_default_accounts, post_all_accounting
from .crud import MODEL_REGISTRY, visible_grouped_registry
from .forms import ContractForm, ContractSiteRequirementForm, InvoiceForm, SecureModelForm
from .security import file_extension, validate_excel_upload, validate_schedule_upload


MODEL_FORM_EXCLUDES = {
    models.Employee: (
        "created_at",
        "updated_at",
        "uniform_size",
        "authority_level",
    ),
}


def get_model_config(slug):
    config = MODEL_REGISTRY.get(slug)
    if not config:
        raise Http404("Page not found")
    return config


def build_model_form(model):
    if model is models.Contract:
        return ContractForm
    if model is models.ContractSiteRequirement:
        return ContractSiteRequirementForm
    if model is models.Invoice:
        return InvoiceForm
    form = modelform_factory(
        model,
        form=SecureModelForm,
        exclude=MODEL_FORM_EXCLUDES.get(model, ("created_at", "updated_at")),
    )
    for field_name, field in form.base_fields.items():
        field.widget.attrs.setdefault("class", "form-control")
        if field_name == "passport_photo":
            field.widget.attrs.setdefault("accept", "image/*")
        if getattr(field.widget, "input_type", "") == "text":
            field.widget.attrs.setdefault("placeholder", field.label)
        if field.widget.__class__.__name__ == "DateInput":
            field.widget = DateInput(attrs={"class": "form-control"}, format="%Y-%m-%d")
            field.input_formats = ["%Y-%m-%d"]
        elif field.widget.__class__.__name__ == "DateTimeInput":
            field.widget = DateTimeInput(attrs={"class": "form-control"}, format="%Y-%m-%dT%H:%M")
            field.input_formats = ["%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M:%S"]
        elif field.widget.__class__.__name__ == "TimeInput":
            field.widget = TimeInput(attrs={"class": "form-control"}, format="%H:%M")
            field.input_formats = ["%H:%M", "%H:%M:%S"]
        elif isinstance(field.widget, FileInput):
            field.widget.attrs["class"] = "form-control"
    return form


def column_label(column):
    labels = {
        "company_number": "Company Number",
        "employee_number": "Employee Number",
        "nssf_number": "NSSF Number",
        "full_name": "Full Name",
        "phone_number": "Phone Number",
        "bank_account": "Bank Account",
        "date_of_birth": "Date Of Birth",
        "national_id": "National ID",
        "hire_date": "Hire Date",
        "uniform_size": "Uniform Size",
        "training_level": "Training Level",
        "assigned_zone": "Assigned Zone",
        "experience_years": "Experience Years",
        "authority_level": "Authority Level",
        "created_at": "Created At",
        "updated_at": "Updated At",
    }
    return labels.get(column, column.replace("_", " ").title())


def column_value(obj, column):
    display_method = getattr(obj, f"get_{column}_display", None)
    if callable(display_method):
        return display_method()
    return getattr(obj, column, "")


def stringify_cell(value):
    if value in (None, ""):
        return "-"
    if isinstance(value, Decimal):
        return f"{value:,.2f}"
    if hasattr(value, "isoformat") and not isinstance(value, str):
        try:
            return value.isoformat()
        except TypeError:
            pass
    return str(value)


def record_queryset_for_request(request, slug, config):
    queryset = scoped_queryset(request.user, slug, config.model.objects.all())
    search_query = request.GET.get("q", "").strip()
    if not search_query:
        return queryset, search_query

    search_text = search_query.lower()
    return [
        obj
        for obj in queryset
        if search_text in str(obj).lower()
        or any(search_text in str(column_value(obj, column)).lower() for column in config.columns)
    ], search_query


def record_rows(queryset, columns, limit=200):
    rows = []
    for obj in queryset[:limit]:
        rows.append(
            {
                "object": obj,
                "values": [column_value(obj, column) for column in columns],
            }
        )
    return rows


def scoped_queryset(user, slug, queryset):
    if is_manager(user) or is_supervisor(user):
        return queryset
    return queryset.none()


def contract_schedule_limit_reached(site, shift, shift_date, deployment=None):
    limit = contract_required_guards(site, shift, shift_date)
    if not limit:
        return False
    schedules = models.GuardSchedule.objects.filter(site=site, shift=shift, shift_date=shift_date).exclude(
        status=models.GuardSchedule.ScheduleStatus.CANCELLED
    )
    if deployment and deployment.pk:
        schedules = schedules.exclude(deployment=deployment)
    return schedules.count() >= limit


def contract_limit_message(site, shift, shift_date):
    shift_name = shift.shift_name
    return (
        f"{site.site_name} already has the contracted {contract_required_guards(site, shift, shift_date)} "
        f"guard(s) for {shift_name} on {shift_date}."
    )


def contract_required_guards(site, shift, shift_date):
    requirement = (
        models.ContractSiteRequirement.objects.select_related("contract")
        .filter(
            site=site,
            status=models.StatusChoices.ACTIVE,
            contract__status=models.StatusChoices.ACTIVE,
            start_date__lte=shift_date,
        )
        .filter(Q(end_date__gte=shift_date) | Q(end_date__isnull=True))
        .filter(Q(contract__end_date__gte=shift_date) | Q(contract__end_date__isnull=True))
        .filter(Q(shift=shift) | Q(shift__isnull=True))
        .order_by("-shift_id", "-start_date")
        .first()
    )
    if requirement:
        return requirement.required_guards
    return site.required_guards_per_shift


@login_required
def home(request):
    return render(
        request,
        "core/home.html",
        {
            "profile_stats": [
                ("24/7", "Control room monitoring"),
                ("3", "Departments integrated"),
                ("100%", "Roster-based deployment"),
                ("Rapid", "Incident response workflow"),
            ],
            "services": [
                {
                    "title": "Manned Guarding",
                    "text": "Trained guards for offices, estates, schools, retail sites, warehouses, and gated communities.",
                },
                {
                    "title": "Site Supervision",
                    "text": "Supervisor-led patrols, deployment checks, shift handovers, and accountability reports.",
                },
                {
                    "title": "Incident Management",
                    "text": "Structured reporting for incident type, severity, location, guard, and follow-up status.",
                },
                {
                    "title": "Workforce Control",
                    "text": "Attendance, leave, training, disciplinary action, documents, and performance tracking.",
                },
            ],
        },
    )


@login_required
@user_passes_test(can_access_internal)
def dashboard(request):
    try:
        visible_groups = visible_grouped_registry(request.user)
        operations = {
            "clients": models.Client.objects.count(),
            "sites": models.Site.objects.count(),
            "deployments": models.Deployment.objects.count(),
            "incidents": models.Incident.objects.count(),
            "patrol_logs": models.PatrolLog.objects.count(),
            "assets": models.Asset.objects.count(),
        }
        human_resource = {
            "employees": models.Employee.objects.count(),
            "recruitment_applications": models.RecruitmentApplication.objects.count(),
            "attendance": models.Attendance.objects.count(),
            "leave_requests": models.Leave.objects.count(),
            "trainings": models.Training.objects.count(),
            "salaries": models.Salary.objects.count(),
        }
        finance = {
            "advances": models.Advance.objects.count(),
            "invoices": models.Invoice.objects.count(),
            "payments": models.Payment.objects.count(),
            "budgets": models.Budget.objects.count(),
            "expenses": models.Expense.objects.count(),
        }
        recruitment_dashboard = {
            "open_requisitions": models.RecruitmentRequisition.objects.filter(
                status__in=[
                    models.RecruitmentRequisition.RequisitionStatus.OPEN,
                    models.RecruitmentRequisition.RequisitionStatus.SHORTLISTING,
                    models.RecruitmentRequisition.RequisitionStatus.INTERVIEWING,
                    models.RecruitmentRequisition.RequisitionStatus.OFFERING,
                ]
            ).count(),
            "applications": models.RecruitmentApplication.objects.count(),
            "accepted_offers": models.JobOffer.objects.filter(status=models.JobOffer.OfferStatus.ACCEPTED).count(),
        }
        training_dashboard = {
            "records": models.Training.objects.count(),
            "successful": models.Training.objects.filter(
                result__in=[models.Training.TrainingResult.COMPLETED, models.Training.TrainingResult.PASSED]
            ).count(),
            "expired": models.Training.objects.filter(result=models.Training.TrainingResult.EXPIRED).count(),
        }
    except DatabaseError as error:
        return render(
            request,
            "core/system_error.html",
            {
                "title": "Database is not available",
                "message": "The ERP database could not be opened in this deployment.",
                "detail": str(error),
            },
            status=503,
        )
    department_totals = {
        "Operations": sum(operations.values()),
        "Human Resource": sum(human_resource.values()),
        "Finance": sum(finance.values()),
    }
    cumulative_total = sum(department_totals.values()) or 1
    running_total = 0
    cumulative_rows = []
    for label, value in department_totals.items():
        running_total += value
        cumulative_rows.append(
            {
                "label": label,
                "value": value,
                "cumulative": running_total,
                "percent": round((running_total / cumulative_total) * 100, 1),
            }
        )
    chart_data = {
        "departmentLabels": list(department_totals.keys()),
        "departmentValues": list(department_totals.values()),
        "operationsLabels": [column_label(key) for key in operations.keys()],
        "operationsValues": list(operations.values()),
        "hrLabels": [column_label(key) for key in human_resource.keys()],
        "hrValues": list(human_resource.values()),
        "financeLabels": [column_label(key) for key in finance.keys()],
        "financeValues": list(finance.values()),
        "cumulativeLabels": [row["label"] for row in cumulative_rows],
        "cumulativeValues": [row["cumulative"] for row in cumulative_rows],
    }
    context = {
        "operations": operations,
        "human_resource": human_resource,
        "finance": finance,
        "recruitment_dashboard": recruitment_dashboard,
        "training_dashboard": training_dashboard,
        "department_totals": department_totals,
        "cumulative_rows": cumulative_rows,
        "chart_data": chart_data,
        "recent_incidents": models.Incident.objects.select_related("employee", "deployment")[:5],
        "open_invoices": models.Invoice.objects.select_related("client").exclude(status=models.StatusChoices.PAID)[:5],
        "model_groups": visible_groups,
    }
    return render(request, "core/dashboard.html", context)


@login_required
@user_passes_test(can_access_internal)
def contract_invoice_data(request, pk):
    contract = get_object_or_404(models.Contract.objects.select_related("client"), pk=pk)
    billing_month = parse_date(request.GET.get("billing_month") or "") or timezone.localdate()
    requirements = (
        models.ContractSiteRequirement.objects.select_related("site")
        .filter(contract=contract, status=models.StatusChoices.ACTIVE, start_date__lte=billing_month)
        .filter(Q(end_date__gte=billing_month) | Q(end_date__isnull=True))
        .order_by("site__site_name", "shift__start_time", "start_date")
    )
    sites = {}
    for requirement in requirements:
        site = requirement.site
        row = sites.setdefault(
            site.id,
            {
                "id": site.id,
                "name": str(site),
                "address": site.site_address,
                "guards": 0,
                "subtotal": Decimal("0.00"),
            },
        )
        row["guards"] += requirement.required_guards
        row["subtotal"] += requirement.billable_total
    for row in sites.values():
        row["subtotal"] = str(row["subtotal"].quantize(Decimal("0.01")))
    client = contract.client
    return JsonResponse(
        {
            "client": {
                "id": client.id,
                "name": client.client_name,
                "address": client.address,
                "email": client.email,
                "contact": client.contact_person,
                "phone": client.phone_number,
            },
            "contract": {
                "id": contract.id,
                "number": contract.contract_number,
                "billing_rate": str(contract.billing_rate),
            },
            "sites": list(sites.values()),
        }
    )


def contract_payload(contract):
    return {
        "id": contract.id,
        "number": contract.contract_number,
        "billing_rate": str(contract.billing_rate),
        "start_date": contract.start_date.isoformat() if contract.start_date else "",
        "end_date": contract.end_date.isoformat() if contract.end_date else "",
        "dog_count": contract.dog_count,
        "dog_rate": str(contract.dog_rate),
        "metal_detector_count": contract.metal_detector_count,
        "metal_detector_rate": str(contract.metal_detector_rate),
        "walk_through_machine_count": contract.walk_through_detector_count,
        "walk_through_machine_rate": str(contract.walk_through_detector_rate),
        "panic_baton_count": contract.panic_baton_count,
        "panic_baton_rate": str(contract.panic_baton_rate),
        "handcuffs_count": contract.handcuffs_count,
        "handcuffs_rate": str(contract.handcuffs_rate),
    }


@login_required
@user_passes_test(can_access_internal)
def client_contract_requirement_data(request, pk):
    client = get_object_or_404(models.Client, pk=pk)
    contracts = [contract_payload(contract) for contract in client.contracts.order_by("-start_date", "contract_number")]
    sites = [
        {
            "id": site.id,
            "name": str(site),
            "site_name": site.site_name,
            "site_address": site.site_address,
            "city": site.city,
        }
        for site in client.sites.order_by("site_name")
    ]
    return JsonResponse(
        {
            "client": {"id": client.id, "name": client.client_name},
            "next_site_code": ContractSiteRequirementForm.next_site_code(client),
            "contracts": contracts,
            "sites": sites,
        }
    )


def month_bounds(value=None):
    selected = parse_date(value) if value else timezone.localdate()
    if not selected:
        selected = timezone.localdate()
    start = selected.replace(day=1)
    end = selected.replace(day=calendar.monthrange(selected.year, selected.month)[1])
    return start, end


def attendance_shift_hours(attendance):
    shift = attendance.shift
    if not shift and attendance.schedule_id:
        shift = attendance.schedule.shift
    if shift:
        return shift.basic_hours, shift.daily_overtime_hours
    if attendance.time_in and attendance.time_out:
        start_minutes = attendance.time_in.hour * 60 + attendance.time_in.minute
        end_minutes = attendance.time_out.hour * 60 + attendance.time_out.minute
        if end_minutes <= start_minutes:
            end_minutes += 24 * 60
        duration = (Decimal(end_minutes - start_minutes) / Decimal(60)).quantize(Decimal("0.01"))
        return min(duration, Decimal("8.00")), max(duration - Decimal("8.00"), Decimal("0.00"))
    return Decimal("0.00"), Decimal("0.00")


def payroll_hourly_rate(employee):
    if employee.position_id and employee.position.salary_range_min:
        return (employee.position.salary_range_min / Decimal("208.00")).quantize(Decimal("0.01"))
    return Decimal("0.00")


def money(value):
    return value.quantize(Decimal("0.01"))


def approved_advance_total(employee_id, end):
    return money(
        models.Advance.objects.filter(
            employee_id=employee_id,
            approval_status=models.StatusChoices.APPROVED,
        )
        .filter(Q(disbursement_date__lte=end) | Q(disbursement_date__isnull=True, request_date__lte=end))
        .exclude(repayment_status__in=[models.StatusChoices.PAID, models.StatusChoices.CLOSED, models.StatusChoices.REJECTED])
        .aggregate(total=Coalesce(Sum("amount_requested"), Decimal("0.00")))["total"]
    )


def previous_advance_deductions(employee_id, start):
    return money(
        models.Salary.objects.filter(employee_id=employee_id, pay_period_start__lt=start).aggregate(
            total=Coalesce(Sum("advance_deduction"), Decimal("0.00"))
        )["total"]
    )


def payroll_advance_values(employee_id, start, end, gross_pay, nssf_employee, other_deductions):
    advance_opening_balance = max(
        approved_advance_total(employee_id, end) - previous_advance_deductions(employee_id, start),
        Decimal("0.00"),
    )
    net_before_advance = max(gross_pay - nssf_employee - other_deductions, Decimal("0.00"))
    advance_deduction = money(min(advance_opening_balance, net_before_advance))
    advance_balance = money(max(advance_opening_balance - advance_deduction, Decimal("0.00")))
    return advance_deduction, advance_balance


def generate_payroll_from_attendance(start, end):
    totals = defaultdict(lambda: {"days": 0, "basic_hours": Decimal("0.00"), "overtime_hours": Decimal("0.00")})
    attendances = (
        models.Attendance.objects.select_related("employee__position", "shift", "schedule__shift")
        .filter(date__range=(start, end), status__iexact="Present")
        .order_by("employee__first_name", "employee__last_name", "date")
    )
    employees = {}
    for attendance in attendances:
        employee = attendance.employee
        employees[employee.id] = employee
        basic_hours, overtime_hours = attendance_shift_hours(attendance)
        totals[employee.id]["days"] += 1
        totals[employee.id]["basic_hours"] += basic_hours
        totals[employee.id]["overtime_hours"] += overtime_hours

    salaries = []
    existing_salaries = {
        salary.employee_id: salary
        for salary in models.Salary.objects.select_related("employee").filter(pay_period_start=start)
    }
    employee_ids = set(totals) | set(existing_salaries)
    for employee_id in employee_ids:
        total = totals[employee_id]
        employee = employees.get(employee_id) or existing_salaries[employee_id].employee
        hourly_rate = payroll_hourly_rate(employee)
        basic_salary = money(total["basic_hours"] * hourly_rate)
        overtime_pay = money(total["overtime_hours"] * hourly_rate * Decimal("1.50"))
        existing = existing_salaries.get(employee_id)
        allowances = existing.allowances if existing else Decimal("0.00")
        other_deductions = existing.deductions if existing else Decimal("0.00")
        bonus = existing.bonus if existing else Decimal("0.00")
        gross_pay = money(basic_salary + allowances + overtime_pay + bonus)
        nssf_employee = money(gross_pay * models.Salary.NSSF_EMPLOYEE_RATE)
        advance_deduction, advance_balance = payroll_advance_values(
            employee_id,
            start,
            end,
            gross_pay,
            nssf_employee,
            other_deductions,
        )
        salary, _created = models.Salary.objects.update_or_create(
            employee=employee,
            pay_period_start=start,
            defaults={
                "pay_period_end": end,
                "attendance_days": total["days"],
                "basic_hours": total["basic_hours"],
                "overtime_hours": total["overtime_hours"],
                "basic_salary": basic_salary,
                "allowances": allowances,
                "deductions": other_deductions,
                "advance_deduction": advance_deduction,
                "advance_balance": advance_balance,
                "overtime_pay": overtime_pay,
                "bonus": bonus,
                "payment_date": existing.payment_date if existing else None,
                "payment_method": existing.payment_method if existing else "",
                "status": existing.status if existing else models.StatusChoices.UNPAID,
            },
        )
        salaries.append(salary)
    return salaries


def refresh_payroll_for_date(value):
    start, end = month_bounds(value.isoformat())
    return generate_payroll_from_attendance(start, end)


def refresh_payroll_for_dates(*values):
    refreshed = {}
    for value in values:
        if not value:
            continue
        start, end = month_bounds(value.isoformat())
        refreshed[start] = generate_payroll_from_attendance(start, end)
    return refreshed


def decimal_from_payload(value):
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value))
    except Exception:
        return None


def shift_contains_datetime(shift, captured_at):
    current_minutes = captured_at.hour * 60 + captured_at.minute
    start_minutes = shift.start_time.hour * 60 + shift.start_time.minute
    end_minutes = shift.end_time.hour * 60 + shift.end_time.minute
    if end_minutes <= start_minutes:
        return current_minutes >= start_minutes or current_minutes <= end_minutes
    return start_minutes <= current_minutes <= end_minutes


def matching_schedule(employee, site, captured_at, payload):
    shift_date = timezone.localtime(captured_at).date()
    schedules = models.GuardSchedule.objects.select_related("shift", "site", "employee").filter(
        employee=employee,
        site=site,
        shift_date=shift_date,
    )
    shift_value = str(payload.get("shift") or payload.get("shift_code") or "").strip()
    shift_id = payload.get("shift_id")
    if shift_id:
        schedules = schedules.filter(shift_id=shift_id)
    elif shift_value:
        schedules = schedules.filter(Q(shift__code__iexact=shift_value) | Q(shift__shift_name__iexact=shift_value))

    schedule_list = list(schedules.order_by("shift__start_time"))
    if len(schedule_list) == 1:
        return schedule_list[0]
    for schedule in schedule_list:
        if shift_contains_datetime(schedule.shift, timezone.localtime(captured_at)):
            return schedule
    return schedule_list[0] if schedule_list else None


def device_token_from_request(request):
    auth_header = request.headers.get("Authorization", "")
    if auth_header.lower().startswith("bearer "):
        return auth_header.split(" ", 1)[1].strip()
    return request.headers.get("X-Device-Token", "").strip()


def json_error(message, status=400, **extra):
    payload = {"status": "error", "message": message}
    payload.update(extra)
    return JsonResponse(payload, status=status)


def healthz(request):
    db_ok = False
    db_error = ""
    db_writable = False
    db_write_error = ""
    env_username = os.environ.get("DJANGO_SUPERUSER_USERNAME", "")
    env_user_exists = False
    env_user_active = False
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT 1")
            db_ok = cursor.fetchone()[0] == 1
            try:
                cursor.execute("CREATE TEMP TABLE healthcheck_write_probe (checked_at text)")
                cursor.execute("INSERT INTO healthcheck_write_probe (checked_at) VALUES (%s)", [timezone.now().isoformat()])
                cursor.execute("DROP TABLE healthcheck_write_probe")
                db_writable = True
            except Exception as error:
                db_write_error = str(error)
        if env_username:
            User = models.User if hasattr(models, "User") else None
            if User is None:
                from django.contrib.auth import get_user_model

                User = get_user_model()
            env_user = User.objects.filter(username=env_username).only("is_active").first()
            env_user_exists = env_user is not None
            env_user_active = bool(env_user and env_user.is_active)
    except Exception as error:
        db_error = str(error)
    return JsonResponse(
        {
            "status": "ok" if db_ok else "degraded",
            "debug": settings.DEBUG,
            "database": "ok" if db_ok else "error",
            "database_error": db_error,
            "database_writable": db_writable,
            "database_write_error": db_write_error,
            "database_engine": settings.DATABASES["default"].get("ENGINE", ""),
            "database_name": settings.DATABASES["default"].get("NAME", ""),
            "database_runtime_note": getattr(settings, "DATABASE_RUNTIME_NOTE", ""),
            "session_engine": getattr(settings, "SESSION_ENGINE", "django.contrib.sessions.backends.db"),
            "permanent_login": getattr(settings, "ERP_PERMANENT_LOGIN", False),
            "permanent_login_username": getattr(settings, "ERP_PERMANENT_LOGIN_USERNAME", ""),
            "saas_platform_name": getattr(settings, "SAAS_PLATFORM_NAME", ""),
            "saas_enforce_tenant_access": getattr(settings, "SAAS_ENFORCE_TENANT_ACCESS", False),
            "tenant_count": models.TenantOrganization.objects.count(),
            "env_superuser_username_configured": bool(env_username),
            "env_superuser_password_configured": bool(os.environ.get("DJANGO_SUPERUSER_PASSWORD")),
            "env_superuser_username": env_username,
            "env_superuser_user_exists": env_user_exists,
            "env_superuser_user_active": env_user_active,
            "allowed_hosts": settings.ALLOWED_HOSTS,
        },
        status=200 if db_ok else 503,
    )


@csrf_exempt
def attendance_swipe_api(request):
    if request.method != "POST":
        return json_error("Only POST is allowed.", status=405)
    if len(request.body) > settings.ATTENDANCE_API_MAX_BODY_BYTES:
        return json_error("Attendance payload is too large.", status=413)
    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return HttpResponseBadRequest("Invalid JSON payload.")

    token = device_token_from_request(request)
    device_identifier = str(payload.get("device_id") or "").strip()
    card_uid = str(payload.get("card_uid") or "").strip()
    event_type = str(payload.get("event") or models.AttendanceDeviceEvent.EventType.CHECK_IN).strip().lower()
    if event_type not in models.AttendanceDeviceEvent.EventType.values:
        event_type = models.AttendanceDeviceEvent.EventType.CHECK_IN

    timestamp_value = payload.get("timestamp") or payload.get("captured_at")
    captured_at = parse_datetime(str(timestamp_value)) if timestamp_value else timezone.now()
    if captured_at is None:
        captured_at = timezone.now()
    if timezone.is_naive(captured_at):
        captured_at = timezone.make_aware(captured_at, timezone.get_current_timezone())

    latitude = decimal_from_payload(payload.get("latitude"))
    longitude = decimal_from_payload(payload.get("longitude"))

    device = None
    if token and device_identifier:
        device = models.AttendanceDevice.objects.filter(device_id=device_identifier, api_key=token, is_active=True).first()
    site = None
    employee = None
    schedule = None
    attendance = None
    status = models.AttendanceDeviceEvent.EventStatus.REJECTED
    message = ""
    distance = None

    site_value = payload.get("site_id") or payload.get("site_code")
    if site_value:
        site_filter = Q(site_code__iexact=str(site_value))
        if str(site_value).isdigit():
            site_filter |= Q(id=int(site_value))
        site = models.Site.objects.filter(site_filter).first()

    if device and device.assigned_site_id:
        if site and site.id != device.assigned_site_id:
            message = "Device is not registered for the submitted site."
        site = device.assigned_site

    if not message:
        if not device:
            message = "Unknown or inactive attendance device."
        elif not card_uid:
            message = "Missing card UID."
        elif not site:
            message = "Unknown attendance site."
        elif latitude is None or longitude is None:
            message = "GPS latitude and longitude are required."
        else:
            within_geofence, distance = site.is_within_geofence(latitude, longitude)
            if distance is None:
                message = "Site geofence is not configured."
            elif not within_geofence:
                message = f"Attendance rejected outside geofence. Distance: {distance:.2f}m; allowed: {site.geofence_radius_meters}m."

    if not message:
        employee = models.Employee.objects.filter(work_card_uid__iexact=card_uid, status=models.StatusChoices.ACTIVE).first()
        if not employee:
            message = "Unknown or inactive guard card."

    if not message:
        schedule = matching_schedule(employee, site, captured_at, payload)
        if not schedule:
            message = "No matching scheduled shift found for this guard, site, and date."

    with transaction.atomic():
        event = models.AttendanceDeviceEvent.objects.create(
            device=device,
            device_identifier=device_identifier,
            card_uid=card_uid,
            employee=employee,
            site=site,
            schedule=schedule,
            event_type=event_type,
            event_timestamp=captured_at,
            latitude=latitude,
            longitude=longitude,
            geofence_distance_meters=Decimal(str(round(distance, 2))) if distance is not None else None,
            status=status,
            message=message,
            payload=payload,
        )

        if not message:
            local_captured_at = timezone.localtime(captured_at)
            defaults = {
                "schedule": schedule,
                "shift": schedule.shift,
                "status": "Present",
                "capture_source": models.Attendance.CaptureSource.IOT,
                "card_uid": card_uid,
                "device_id": device.device_id,
                "captured_by": device.assigned_supervisor.user if device.assigned_supervisor_id and device.assigned_supervisor.user_id else None,
                "captured_at": captured_at,
                "latitude": latitude,
                "longitude": longitude,
                "geofence_distance_meters": Decimal(str(round(distance, 2))) if distance is not None else None,
                "remarks": f"IoT {event.get_event_type_display()} captured by {device.device_id}.",
            }
            if event_type == models.AttendanceDeviceEvent.EventType.CHECK_OUT:
                defaults["time_out"] = local_captured_at.time()
            else:
                defaults["time_in"] = local_captured_at.time()
            attendance, _created = models.Attendance.objects.update_or_create(
                employee=employee,
                date=schedule.shift_date,
                shift=schedule.shift,
                site=site,
                defaults=defaults,
            )
            schedule.status = models.GuardSchedule.ScheduleStatus.COMPLETED
            schedule.notes = f"IoT attendance captured by {device.device_id}."
            schedule.save(update_fields=["status", "notes", "updated_at"])
            event.attendance = attendance
            event.status = models.AttendanceDeviceEvent.EventStatus.ACCEPTED
            event.message = "Attendance captured successfully."
            event.save(update_fields=["attendance", "status", "message", "updated_at"])

    if attendance:
        refresh_payroll_for_date(attendance.date)
        return JsonResponse(
            {
                "status": "ok",
                "message": "Attendance captured successfully.",
                "employee": employee.full_name,
                "company_number": employee.company_number,
                "site": site.site_name,
                "shift": schedule.shift.shift_name,
                "date": schedule.shift_date.isoformat(),
                "event": event_type,
                "distance_meters": round(distance, 2) if distance is not None else None,
                "attendance_id": attendance.id,
                "event_id": event.id,
            }
        )
    return json_error(message, event_id=event.id)


def payroll_queryset(start):
    return models.Salary.objects.select_related("employee").filter(pay_period_start=start).order_by(
        "employee__first_name", "employee__last_name"
    )


def payroll_headers():
    return [
        "Employee",
        "Company Number",
        "NSSF Number",
        "Bank Account",
        "Days",
        "Basic Hours",
        "Overtime Hours",
        "Basic Pay",
        "Overtime Pay",
        "Gross Pay",
        "NSSF Employee",
        "NSSF Employer",
        "Other Deductions",
        "Advance Deduction",
        "Advance Balance",
        "Total Deductions",
        "Net Salary",
        "Status",
    ]


def payroll_row(salary):
    return [
        salary.employee.full_name,
        salary.employee.company_number or "",
        salary.employee.nssf_number or "",
        salary.employee.bank_account or "",
        salary.attendance_days,
        salary.basic_hours,
        salary.overtime_hours,
        salary.basic_salary,
        salary.overtime_pay,
        salary.gross_pay,
        salary.nssf_employee,
        salary.nssf_employer,
        salary.deductions,
        salary.advance_deduction,
        salary.advance_balance,
        salary.total_deductions,
        salary.net_salary,
        salary.get_status_display(),
    ]


PDF_NAVY = colors.HexColor("#102033")
PDF_GREEN = colors.HexColor("#16824a")
PDF_BLUE = colors.HexColor("#2563eb")
PDF_LINE = colors.HexColor("#d9dee4")
PDF_SOFT = colors.HexColor("#f4f7fb")
PDF_TEXT = colors.HexColor("#243244")
PDF_MUTED = colors.HexColor("#667085")


def money_display(value):
    return f"UGX {Decimal(value or 0):,.2f}"


def response_pdf_bytes(response):
    return bytes(response.content)


def safe_filename(value):
    value = re.sub(r"[^A-Za-z0-9_.-]+", "-", str(value or "").strip()).strip("-")
    return value or "document"


def pdf_styles():
    styles = getSampleStyleSheet()
    styles.add(
        ParagraphStyle(
            name="DocTitle",
            parent=styles["Title"],
            fontName="Helvetica-Bold",
            fontSize=22,
            leading=26,
            textColor=PDF_NAVY,
            spaceAfter=4,
        )
    )
    styles.add(
        ParagraphStyle(
            name="SectionTitle",
            parent=styles["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=10,
            leading=13,
            textColor=PDF_NAVY,
            spaceBefore=8,
            spaceAfter=6,
        )
    )
    styles.add(
        ParagraphStyle(
            name="SmallMuted",
            parent=styles["Normal"],
            fontSize=8,
            leading=10,
            textColor=PDF_MUTED,
        )
    )
    styles.add(
        ParagraphStyle(
            name="Right",
            parent=styles["Normal"],
            alignment=TA_RIGHT,
            fontSize=9,
            leading=11,
            textColor=PDF_TEXT,
        )
    )
    styles.add(
        ParagraphStyle(
            name="CenterSmall",
            parent=styles["Normal"],
            alignment=TA_CENTER,
            fontSize=8,
            leading=10,
            textColor=PDF_MUTED,
        )
    )
    return styles


def pdf_header_footer(title):
    def draw(canvas, document):
        canvas.saveState()
        width, height = document.pagesize
        canvas.setFillColor(PDF_NAVY)
        canvas.rect(0, height - 54, width, 54, stroke=0, fill=1)
        canvas.setFillColor(colors.white)
        canvas.setFont("Helvetica-Bold", 12)
        canvas.drawString(document.leftMargin, height - 30, "Security Company Management")
        canvas.setFont("Helvetica", 8)
        canvas.drawString(document.leftMargin, height - 43, title)
        canvas.setFillColor(PDF_GREEN)
        canvas.rect(width - document.rightMargin - 86, height - 38, 86, 14, stroke=0, fill=1)
        canvas.setFillColor(colors.white)
        canvas.setFont("Helvetica-Bold", 7)
        canvas.drawCentredString(width - document.rightMargin - 43, height - 34, "OFFICIAL DOCUMENT")
        canvas.setFillColor(PDF_MUTED)
        canvas.setFont("Helvetica", 8)
        canvas.drawString(document.leftMargin, 24, "Generated by Security Company Management System")
        canvas.drawRightString(width - document.rightMargin, 24, f"Page {document.page}")
        canvas.restoreState()

    return draw


def key_value_table(rows, col_widths):
    table = Table(rows, colWidths=col_widths, hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), PDF_SOFT),
                ("TEXTCOLOR", (0, 0), (0, -1), PDF_MUTED),
                ("TEXTCOLOR", (1, 0), (1, -1), PDF_TEXT),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                ("LEADING", (0, 0), (-1, -1), 11),
                ("GRID", (0, 0), (-1, -1), 0.35, PDF_LINE),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 7),
                ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    return table


def styled_table(data, col_widths=None, right_columns=()):
    table = Table(data, colWidths=col_widths, repeatRows=1, hAlign="LEFT")
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), PDF_NAVY),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.35, PDF_LINE),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#fbfcfd")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]
    for column in right_columns:
        style.append(("ALIGN", (column, 1), (column, -1), "RIGHT"))
    table.setStyle(TableStyle(style))
    return table


def signature_table(left_label="Prepared By", right_label="Approved By"):
    data = [
        ["", ""],
        [left_label, right_label],
    ]
    table = Table(data, colWidths=[240, 240], hAlign="CENTER")
    table.setStyle(
        TableStyle(
            [
                ("LINEABOVE", (0, 1), (-1, 1), 0.6, PDF_LINE),
                ("TEXTCOLOR", (0, 1), (-1, 1), PDF_MUTED),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("TOPPADDING", (0, 0), (-1, 0), 24),
                ("TOPPADDING", (0, 1), (-1, 1), 6),
            ]
        )
    )
    return table


def attendance_summary_rows(start, end):
    today = timezone.localdate()
    is_current_month = start <= today <= end
    elapsed_days = today.day if is_current_month else calendar.monthrange(start.year, start.month)[1]
    month_days = calendar.monthrange(start.year, start.month)[1]
    attendances = (
        models.Attendance.objects.select_related("employee", "site", "schedule__site")
        .filter(date__range=(start, end), status__iexact="Present")
        .order_by("employee__first_name", "employee__last_name", "date")
    )
    totals = defaultdict(lambda: {"count": 0, "sites": set()})
    for attendance in attendances:
        totals[attendance.employee_id]["count"] += 1
        site = attendance.site or (attendance.schedule.site if attendance.schedule_id else None)
        if site:
            site_label = f"{site.site_code or 'SITE'} - {site.site_name}"
            totals[attendance.employee_id]["sites"].add(site_label)

    employees = models.Employee.objects.filter(company_number__isnull=False).exclude(company_number="").order_by(
        "first_name", "last_name", "company_number"
    )
    rows = []
    for employee in employees:
        total = totals[employee.id]
        attended = total["count"]
        projected = attended
        if is_current_month and attended:
            projected = max(attended, (attended * month_days + elapsed_days - 1) // elapsed_days)
        rows.append(
            {
                "employee_number": employee.company_number,
                "name": employee.full_name,
                "sites": ", ".join(sorted(total["sites"])) or "-",
                "attendance": attended,
                "expected_growth": max(projected - attended, 0),
                "projected_total": projected,
            }
        )
    return rows


@login_required
@user_passes_test(can_manage_attendance)
def attendance_summary_pdf(request):
    selected_month = request.GET.get("month") or timezone.localdate().isoformat()[:7]
    start, end = month_bounds(f"{selected_month}-01")
    rows = attendance_summary_rows(start, end)
    total_attendance = sum(row["attendance"] for row in rows)
    total_growth = sum(row["expected_growth"] for row in rows)

    output = BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=24,
        leftMargin=24,
        topMargin=72,
        bottomMargin=42,
    )
    styles = pdf_styles()
    elements = [
        Paragraph("Roster Attendance Summary", styles["DocTitle"]),
        Paragraph(f"Period: {start:%Y-%m-%d} to {end:%Y-%m-%d}", styles["SmallMuted"]),
        Spacer(1, 10),
        key_value_table(
            [
                ["Employees", len(rows)],
                ["Total attendance", total_attendance],
                ["Expected growth", total_growth],
            ],
            [110, 120],
        ),
        Spacer(1, 12),
        Paragraph("Employee Attendance Summary", styles["SectionTitle"]),
    ]
    data = [["Emp No.", "Name", "Sites Worked", "Attendance", "Expected Growth", "Projected Total"]]
    for row in rows:
        data.append(
            [
                row["employee_number"],
                row["name"],
                Paragraph(row["sites"], styles["SmallMuted"]),
                row["attendance"],
                row["expected_growth"],
                row["projected_total"],
            ]
        )
    elements.append(styled_table(data, col_widths=[64, 120, 300, 68, 82, 82], right_columns=(3, 4, 5)))
    elements.extend([Spacer(1, 18), signature_table("Prepared By", "Operations Manager")])
    document.build(elements, onFirstPage=pdf_header_footer("Roster Attendance Summary"), onLaterPages=pdf_header_footer("Roster Attendance Summary"))

    response = HttpResponse(output.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="attendance-summary-{selected_month}.pdf"'
    return response


@login_required
@user_passes_test(is_manager)
def payroll(request):
    selected_month = request.POST.get("month") or request.GET.get("month") or timezone.localdate().isoformat()[:7]
    start, end = month_bounds(f"{selected_month}-01")
    if request.method == "POST":
        generated = generate_payroll_from_attendance(start, end)
        messages.success(request, f"Payroll generated for {len(generated)} employee(s) from attendance.")
        return redirect(f"/payroll/?month={selected_month}")
    generate_payroll_from_attendance(start, end)
    salaries = payroll_queryset(start)
    return render(
        request,
        "core/payroll.html",
        {
            "selected_month": selected_month,
            "pay_period_start": start,
            "pay_period_end": end,
            "salaries": salaries,
        },
    )


def audit_difference(primary, comparison):
    return primary - comparison


def audit_verdict(spent, budget, success_count, total_count, success_label):
    if not total_count and not spent:
        return "No activity recorded for this month."
    if budget and spent > budget and not success_count:
        return f"Overspending: costs are above budget and no {success_label} recorded."
    if budget and spent > budget:
        return f"Over budget, but producing {success_label}; review cost control."
    if success_count:
        return f"Benefiting: {success_count} {success_label} recorded within the tracked activity."
    if spent:
        return f"Spending recorded, but no {success_label} yet."
    return f"No {success_label} recorded yet."


@login_required
@user_passes_test(is_manager)
def audit_report(request):
    selected_month = request.GET.get("month") or timezone.localdate().isoformat()[:7]
    start, end = month_bounds(f"{selected_month}-01")

    attendances = models.Attendance.objects.filter(date__range=(start, end), status__iexact="Present")
    salaries = models.Salary.objects.filter(pay_period_start=start)
    invoices = models.Invoice.objects.filter(billing_month__range=(start, end))
    active_contracts = models.Contract.objects.filter(
        status=models.StatusChoices.ACTIVE,
        start_date__lte=end,
    ).filter(Q(end_date__gte=start) | Q(end_date__isnull=True))
    active_clients = models.Client.objects.filter(
        contract_status=models.StatusChoices.ACTIVE,
        contract_start_date__lte=end,
    ).filter(Q(contract_end_date__gte=start) | Q(contract_end_date__isnull=True))
    active_deployments = models.Deployment.objects.filter(
        status=models.StatusChoices.ACTIVE,
        start_date__lte=end,
    ).filter(Q(end_date__gte=start) | Q(end_date__isnull=True))
    recruitment_requisitions = models.RecruitmentRequisition.objects.filter(opening_date__range=(start, end))
    recruitment_applications = models.RecruitmentApplication.objects.filter(date_received__range=(start, end))
    job_offers = models.JobOffer.objects.filter(offer_date__range=(start, end))
    trainings = models.Training.objects.filter(start_date__range=(start, end))

    attendance_shift_count = attendances.count()
    payroll_employee_count = salaries.values("employee_id").distinct().count()
    salary_totals = salaries.aggregate(
        gross=Coalesce(Sum("gross_pay"), Decimal("0.00")),
        net=Coalesce(Sum("net_salary"), Decimal("0.00")),
    )
    invoice_totals = invoices.aggregate(
        billed=Coalesce(Sum("total_amount"), Decimal("0.00")),
        paid=Coalesce(Sum("paid_amount"), Decimal("0.00")),
        balance=Coalesce(Sum("balance_amount"), Decimal("0.00")),
    )
    asset_totals = models.Asset.objects.aggregate(purchased=Coalesce(Sum("quantity"), 0))
    assets_at_hand = models.Asset.objects.filter(Q(assigned_to__isnull=True) | Q(return_date__isnull=False)).aggregate(
        total=Coalesce(Sum("quantity"), 0)
    )["total"]
    assets_deployed = models.Asset.objects.filter(assigned_to__isnull=False, return_date__isnull=True).aggregate(
        total=Coalesce(Sum("quantity"), 0)
    )["total"]

    paid_invoices = invoices.filter(status=models.StatusChoices.PAID).count()
    unpaid_invoices = invoices.exclude(status=models.StatusChoices.PAID).count()
    invoiced_client_count = invoices.values("client_id").distinct().count()
    invoiced_contract_count = invoices.filter(contract__isnull=False).values("contract_id").distinct().count()
    active_client_count = active_clients.count()
    active_contract_count = active_contracts.count()
    employee_count = models.Employee.objects.count()
    active_employee_count = models.Employee.objects.filter(status=models.StatusChoices.ACTIVE).count()
    deployed_employee_count = active_deployments.values("employee_id").distinct().count()
    recruitment_totals = recruitment_requisitions.aggregate(
        budget=Coalesce(Sum("recruitment_budget"), Decimal("0.00")),
        spent=Coalesce(Sum("actual_recruitment_cost"), Decimal("0.00")),
        openings=Coalesce(Sum("number_of_openings"), 0),
    )
    accepted_offers = job_offers.filter(status=models.JobOffer.OfferStatus.ACCEPTED).count()
    hired_applications = recruitment_applications.filter(status=models.RecruitmentApplication.ApplicationStatus.HIRED).count()
    recruitment_success_count = accepted_offers + hired_applications
    training_totals = trainings.aggregate(
        budget=Coalesce(Sum("budgeted_cost"), Decimal("0.00")),
        spent=Coalesce(Sum("training_cost"), Decimal("0.00")),
        hours=Coalesce(Sum("duration_hours"), Decimal("0.00")),
    )
    successful_training_count = trainings.filter(
        result__in=[models.Training.TrainingResult.COMPLETED, models.Training.TrainingResult.PASSED],
    ).count()
    expired_training_count = trainings.filter(result=models.Training.TrainingResult.EXPIRED).count()

    audit_rows = [
        {
            "area": "Attendance vs Payroll",
            "metric": "Present shifts recorded / employees paid",
            "expected": attendance_shift_count,
            "actual": payroll_employee_count,
            "difference": audit_difference(attendance_shift_count, payroll_employee_count),
            "amount": salary_totals["net"],
            "note": "Net payroll generated for the month.",
        },
        {
            "area": "Payroll",
            "metric": "Gross salary generated",
            "expected": salary_totals["gross"],
            "actual": salary_totals["net"],
            "difference": salary_totals["gross"] - salary_totals["net"],
            "amount": salary_totals["net"],
            "note": "Gross less net shows payroll deductions.",
        },
        {
            "area": "Invoices",
            "metric": "Paid / unpaid invoices",
            "expected": paid_invoices,
            "actual": unpaid_invoices,
            "difference": paid_invoices - unpaid_invoices,
            "amount": invoice_totals["balance"],
            "note": "Amount shows unpaid invoice balance.",
        },
        {
            "area": "Invoice Coverage",
            "metric": "Invoiced / uninvoiced clients",
            "expected": invoiced_client_count,
            "actual": max(active_client_count - invoiced_client_count, 0),
            "difference": invoiced_client_count - max(active_client_count - invoiced_client_count, 0),
            "amount": invoice_totals["billed"],
            "note": "Based on active clients in the selected month.",
        },
        {
            "area": "Contract Coverage",
            "metric": "Invoiced / uninvoiced contracts",
            "expected": invoiced_contract_count,
            "actual": max(active_contract_count - invoiced_contract_count, 0),
            "difference": invoiced_contract_count - max(active_contract_count - invoiced_contract_count, 0),
            "amount": invoice_totals["billed"],
            "note": "Based on active contracts in the selected month.",
        },
        {
            "area": "Equipment",
            "metric": "Purchased / at hand",
            "expected": asset_totals["purchased"],
            "actual": assets_at_hand,
            "difference": asset_totals["purchased"] - assets_at_hand,
            "amount": None,
            "note": f"{assets_deployed} equipment item(s) currently issued.",
        },
        {
            "area": "Employees",
            "metric": "Database / deployed",
            "expected": employee_count,
            "actual": deployed_employee_count,
            "difference": employee_count - deployed_employee_count,
            "amount": None,
            "note": f"{active_employee_count} active employee(s) in the database.",
        },
        {
            "area": "Recruitment",
            "metric": "Budget / actual recruitment spend",
            "expected": recruitment_totals["budget"],
            "actual": recruitment_totals["spent"],
            "difference": recruitment_totals["budget"] - recruitment_totals["spent"],
            "amount": recruitment_totals["spent"],
            "note": audit_verdict(
                recruitment_totals["spent"],
                recruitment_totals["budget"],
                recruitment_success_count,
                recruitment_applications.count(),
                "hire(s) or accepted offer(s)",
            ),
        },
        {
            "area": "Recruitment Output",
            "metric": "Openings / applications received",
            "expected": recruitment_totals["openings"],
            "actual": recruitment_applications.count(),
            "difference": recruitment_applications.count() - recruitment_totals["openings"],
            "amount": None,
            "note": f"{accepted_offers} accepted offer(s), {hired_applications} hired application(s).",
        },
        {
            "area": "Training",
            "metric": "Budget / actual training spend",
            "expected": training_totals["budget"],
            "actual": training_totals["spent"],
            "difference": training_totals["budget"] - training_totals["spent"],
            "amount": training_totals["spent"],
            "note": audit_verdict(
                training_totals["spent"],
                training_totals["budget"],
                successful_training_count,
                trainings.count(),
                "completed or passed training(s)",
            ),
        },
        {
            "area": "Training Output",
            "metric": "Training records / successful trainings",
            "expected": trainings.count(),
            "actual": successful_training_count,
            "difference": successful_training_count - trainings.count(),
            "amount": training_totals["hours"],
            "note": f"{expired_training_count} expired training record(s); amount shows total training hours.",
        },
    ]

    summary_cards = [
        ("Attendance Shifts", attendance_shift_count, "fa-calendar-check"),
        ("Net Payroll", salary_totals["net"], "fa-money-check-dollar"),
        ("Invoice Balance", invoice_totals["balance"], "fa-file-invoice-dollar"),
        ("Employees Deployed", deployed_employee_count, "fa-user-shield"),
        ("Recruitment Spend", recruitment_totals["spent"], "fa-user-plus"),
        ("Training Spend", training_totals["spent"], "fa-graduation-cap"),
    ]

    return render(
        request,
        "core/audit_report.html",
        {
            "selected_month": selected_month,
            "pay_period_start": start,
            "pay_period_end": end,
            "audit_rows": audit_rows,
            "summary_cards": summary_cards,
            "invoice_totals": invoice_totals,
        },
    )


def posted_lines_until(end_date=None):
    lines = models.JournalLine.objects.select_related("account", "journal_entry").filter(
        journal_entry__status=models.JournalEntry.EntryStatus.POSTED
    )
    if end_date:
        lines = lines.filter(journal_entry__entry_date__lte=end_date)
    return lines


def account_balances(account_types=None, end_date=None):
    ensure_default_accounts()
    accounts = models.Account.objects.filter(is_active=True).order_by("account_code")
    if account_types:
        accounts = accounts.filter(account_type__in=account_types)
    rows = []
    for account in accounts:
        lines = posted_lines_until(end_date).filter(account=account)
        debit = sum(line.debit for line in lines)
        credit = sum(line.credit for line in lines)
        if account.account_type in {models.Account.AccountType.ASSET, models.Account.AccountType.EXPENSE}:
            balance = debit - credit
        else:
            balance = credit - debit
        rows.append({"account": account, "debit": debit, "credit": credit, "balance": balance})
    return rows


def journal_entry_totals(reference):
    entry = (
        models.JournalEntry.objects.prefetch_related("lines")
        .filter(reference=reference, status=models.JournalEntry.EntryStatus.POSTED)
        .first()
    )
    if not entry:
        return None
    debit = sum(line.debit for line in entry.lines.all())
    credit = sum(line.credit for line in entry.lines.all())
    return {"entry": entry, "debit": debit, "credit": credit, "balanced": debit == credit}


def reconciliation_status(expected_amount, posted):
    if posted is None:
        return "Missing", expected_amount
    if not posted["balanced"]:
        return "Unbalanced", posted["debit"] - posted["credit"]
    if posted["debit"] != expected_amount or posted["credit"] != expected_amount:
        return "Difference", expected_amount - posted["debit"]
    return "Matched", Decimal("0.00")


def reconciliation_row(module, reference, source_label, expected_amount, posted, source_date=None):
    status, difference = reconciliation_status(expected_amount, posted)
    return {
        "module": module,
        "reference": reference,
        "source_label": source_label,
        "source_date": source_date,
        "expected_amount": expected_amount,
        "posted_debit": posted["debit"] if posted else Decimal("0.00"),
        "posted_credit": posted["credit"] if posted else Decimal("0.00"),
        "difference": difference,
        "status": status,
        "entry": posted["entry"] if posted else None,
    }


def build_reconciliation_rows():
    rows = []
    for invoice in models.Invoice.objects.select_related("client").order_by("-invoice_date", "invoice_number"):
        reference = f"INV-{invoice.id}"
        rows.append(
            reconciliation_row(
                "Invoice",
                reference,
                invoice.invoice_number,
                invoice.total_amount,
                journal_entry_totals(reference),
                invoice.invoice_date,
            )
        )
    for payment in models.Payment.objects.select_related("invoice", "employee").order_by("-payment_date", "id"):
        reference = f"PAY-{payment.id}"
        target = payment.invoice.invoice_number if payment.invoice_id else payment.employee.full_name if payment.employee_id else "General"
        rows.append(
            reconciliation_row(
                "Payment",
                reference,
                target,
                payment.amount,
                journal_entry_totals(reference),
                payment.payment_date,
            )
        )
    for expense in models.Expense.objects.order_by("-expense_date", "id"):
        reference = f"EXP-{expense.id}"
        rows.append(
            reconciliation_row(
                "Expense",
                reference,
                expense.category,
                expense.amount,
                journal_entry_totals(reference),
                expense.expense_date,
            )
        )
    for salary in models.Salary.objects.select_related("employee").order_by("-pay_period_start", "employee__first_name"):
        reference = f"PAYROLL-{salary.id}"
        expected = salary.gross_pay + salary.nssf_employer
        rows.append(
            reconciliation_row(
                "Payroll",
                reference,
                f"{salary.employee.full_name} - {salary.pay_period_start:%b %Y}",
                expected,
                journal_entry_totals(reference),
                salary.pay_period_start,
            )
        )
    return rows


def customer_code(client):
    prefix = models.Site.client_code_prefix(client.client_name)
    return f"C{prefix}{client.id:04d}"[:12]


def client_area(client):
    site = client.sites.order_by("site_name").first()
    if not site:
        return "-"
    return site.city or site.state or "-"


def client_manager_and_collector(client):
    deployment = (
        models.Deployment.objects.select_related("supervisor")
        .filter(client=client, supervisor__isnull=False)
        .order_by("-start_date")
        .first()
    )
    manager = deployment.supervisor.full_name if deployment and deployment.supervisor_id else "-"
    return manager, manager


def months_overdue(invoice, as_of):
    if invoice.balance_amount <= 0:
        return "-"
    days = (as_of - invoice.due_date).days
    if days <= 0:
        return "0"
    return str(max(1, (days + 29) // 30))


def aging_rows(as_of):
    rows = []
    clients = models.Client.objects.prefetch_related("sites").order_by("client_name")
    for client in clients:
        invoices = models.Invoice.objects.filter(client=client, invoice_date__lte=as_of)
        if not invoices.exists():
            continue
        payments = models.Payment.objects.filter(invoice__client=client, payment_date__lte=as_of)
        invoice_total = invoices.aggregate(total=Coalesce(Sum("total_amount"), Decimal("0.00")))["total"]
        receipts_total = payments.aggregate(total=Coalesce(Sum("amount"), Decimal("0.00")))["total"]
        balance_due = invoices.aggregate(total=Coalesce(Sum("balance_amount"), Decimal("0.00")))["total"]
        if invoice_total == 0 and receipts_total == 0 and balance_due == 0:
            continue
        area = client_area(client)
        manager, debt_collector = client_manager_and_collector(client)
        open_invoices = invoices.exclude(status=models.StatusChoices.PAID).order_by("due_date")
        month_values = [months_overdue(invoice, as_of) for invoice in open_invoices]
        numeric_months = [int(value) for value in month_values if str(value).isdigit()]
        rows.append(
            {
                "customer_code": customer_code(client),
                "customer_name": client.client_name,
                "area": area,
                "manager": manager,
                "debt_collector": debt_collector,
                "receipts": receipts_total,
                "invoices": invoice_total,
                "balance_due": balance_due,
                "months": max(numeric_months) if numeric_months else "-",
                "group_key": (area, manager, debt_collector),
            }
        )
    groups = []
    for key in sorted({row["group_key"] for row in rows}):
        group_rows = [row for row in rows if row["group_key"] == key]
        groups.append(
            {
                "key": key,
                "rows": group_rows,
                "receipts": sum(row["receipts"] for row in group_rows),
                "invoices": sum(row["invoices"] for row in group_rows),
                "balance_due": sum(row["balance_due"] for row in group_rows),
            }
        )
    return groups


def reconciliation_rows_for(module):
    return [row for row in build_reconciliation_rows() if row["module"] == module]


def reconciliation_summary(rows):
    return {
        "matched": sum(1 for row in rows if row["status"] == "Matched"),
        "missing": sum(1 for row in rows if row["status"] == "Missing"),
        "unbalanced": sum(1 for row in rows if row["status"] == "Unbalanced"),
        "difference": sum(1 for row in rows if row["status"] == "Difference"),
        "expected": sum(row["expected_amount"] for row in rows),
        "posted_debit": sum(row["posted_debit"] for row in rows),
        "posted_credit": sum(row["posted_credit"] for row in rows),
        "variance": sum(row["difference"] for row in rows),
    }


def render_reconciliation_module(request, module, template_title):
    rows = reconciliation_rows_for(module)
    selected_status = request.GET.get("status", "").strip()
    selected_month = request.GET.get("month", "").strip()
    if selected_month:
        rows = [
            row for row in rows
            if row.get("source_date") and row["source_date"].strftime("%Y-%m") == selected_month
        ]
    if selected_status:
        rows = [row for row in rows if row["status"] == selected_status]
    template_name = "core/payroll_reconciliation_report.html" if module == "Payroll" else "core/reconciliation_module_report.html"
    return render(
        request,
        template_name,
        {
            "title": template_title,
            "module": module,
            "rows": rows,
            "summary": reconciliation_summary(rows),
            "selected_status": selected_status,
            "selected_month": selected_month,
            "status_options": ["Matched", "Missing", "Difference", "Unbalanced"],
        },
    )


@login_required
@user_passes_test(is_manager)
def post_accounting_entries(request):
    entries = post_all_accounting(posted_by=request.user)
    messages.success(request, f"Posted {len(entries)} accounting journal entries.")
    return redirect("core:trial_balance")


@login_required
@user_passes_test(is_manager)
def general_ledger(request):
    account_id = request.GET.get("account")
    accounts = models.Account.objects.order_by("account_code")
    lines = posted_lines_until().order_by("journal_entry__entry_date", "journal_entry__reference", "id")
    selected_account = None
    if account_id:
        selected_account = get_object_or_404(models.Account, id=account_id)
        lines = lines.filter(account=selected_account)
    return render(
        request,
        "core/general_ledger.html",
        {"accounts": accounts, "selected_account": selected_account, "selected_account_id": account_id or "", "lines": lines},
    )


@login_required
@user_passes_test(is_manager)
def trial_balance(request):
    rows = account_balances()
    return render(
        request,
        "core/trial_balance.html",
        {
            "rows": rows,
            "total_debit": sum(row["debit"] for row in rows),
            "total_credit": sum(row["credit"] for row in rows),
        },
    )


@login_required
@user_passes_test(is_manager)
def balance_sheet(request):
    asset_rows = account_balances([models.Account.AccountType.ASSET])
    liability_rows = account_balances([models.Account.AccountType.LIABILITY])
    equity_rows = account_balances([models.Account.AccountType.EQUITY])
    return render(
        request,
        "core/balance_sheet.html",
        {
            "asset_rows": asset_rows,
            "liability_rows": liability_rows,
            "equity_rows": equity_rows,
            "total_assets": sum(row["balance"] for row in asset_rows),
            "total_liabilities": sum(row["balance"] for row in liability_rows),
            "total_equity": sum(row["balance"] for row in equity_rows),
        },
    )


@login_required
@user_passes_test(is_manager)
def income_statement(request):
    income_rows = account_balances([models.Account.AccountType.INCOME])
    expense_rows = account_balances([models.Account.AccountType.EXPENSE])
    total_income = sum(row["balance"] for row in income_rows)
    total_expenses = sum(row["balance"] for row in expense_rows)
    return render(
        request,
        "core/income_statement.html",
        {
            "income_rows": income_rows,
            "expense_rows": expense_rows,
            "total_income": total_income,
            "total_expenses": total_expenses,
            "net_income": total_income - total_expenses,
        },
    )


@login_required
@user_passes_test(is_manager)
def receivables_aging(request):
    as_of = parse_date(request.GET.get("as_of") or "") or timezone.localdate()
    groups = aging_rows(as_of)
    return render(
        request,
        "core/receivables_aging.html",
        {
            "as_of": as_of,
            "groups": groups,
            "total_receipts": sum(group["receipts"] for group in groups),
            "total_invoices": sum(group["invoices"] for group in groups),
            "grand_total": sum(group["balance_due"] for group in groups),
        },
    )


@login_required
@user_passes_test(is_manager)
def reconciliation_report(request):
    rows = build_reconciliation_rows()
    selected_status = request.GET.get("status", "").strip()
    if selected_status:
        rows = [row for row in rows if row["status"] == selected_status]
    return render(
        request,
        "core/reconciliation_report.html",
        {
            "rows": rows,
            "summary": reconciliation_summary(rows),
            "selected_status": selected_status,
            "status_options": ["Matched", "Missing", "Difference", "Unbalanced"],
        },
    )


@login_required
@user_passes_test(is_manager)
def payroll_reconciliation_report(request):
    return render_reconciliation_module(request, "Payroll", "Payroll Reconciliation")


@login_required
@user_passes_test(is_manager)
def expense_reconciliation_report(request):
    return render_reconciliation_module(request, "Expense", "Expense Reconciliation")


@login_required
@user_passes_test(is_manager)
def payment_reconciliation_report(request):
    return render_reconciliation_module(request, "Payment", "Payment Reconciliation")


@login_required
@user_passes_test(can_access_internal)
def reports_center(request):
    manager = is_manager(request.user)
    report_groups = [
        (
            "Operations",
            [
                {
                    "title": "Attendance Report",
                    "description": "Query employee attendance by number and period.",
                    "url_name": "core:attendance_report",
                    "icon": "fa-table-list",
                },
                {
                    "title": "Zonal Employees",
                    "description": "View employee allocation by zone and supervisor.",
                    "url_name": "core:zonal_guard_list",
                    "icon": "fa-map-location-dot",
                },
                {
                    "title": "Zone Shift Summary",
                    "description": "Review scheduled guards by zone, site, date, and shift.",
                    "url_name": "core:zone_shift_summary",
                    "icon": "fa-chart-column",
                },
                {
                    "title": "Asset Report",
                    "description": "Track security equipment, condition, and assignments.",
                    "url_name": "core:asset_report",
                    "icon": "fa-boxes-stacked",
                },
            ],
        ),
        (
            "Human Resources",
            [
                {
                    "title": "Payroll",
                    "description": "Generate monthly payroll from attendance and export summaries.",
                    "url_name": "core:payroll",
                    "icon": "fa-money-check-dollar",
                    "manager_only": True,
                },
                {
                    "title": "Employee Records",
                    "description": "Review workforce bio-data, roles, zones, and status.",
                    "url_name": "core:record_list",
                    "url_args": ["employees"],
                    "icon": "fa-user-shield",
                },
                {
                    "title": "Training Records",
                    "description": "Review training completion, costs, certificates, and refresh dates.",
                    "url_name": "core:record_list",
                    "url_args": ["trainings"],
                    "icon": "fa-graduation-cap",
                },
                {
                    "title": "Recruitment Applications",
                    "description": "Track candidate applications, screening, interviews, and offers.",
                    "url_name": "core:record_list",
                    "url_args": ["recruitment-applications"],
                    "icon": "fa-user-plus",
                },
            ],
        ),
        (
            "Finance",
            [
                {
                    "title": "General Ledger",
                    "description": "Review posted journal entries by account.",
                    "url_name": "core:general_ledger",
                    "icon": "fa-book",
                    "manager_only": True,
                },
                {
                    "title": "Trial Balance",
                    "description": "Confirm debit and credit balances across accounts.",
                    "url_name": "core:trial_balance",
                    "icon": "fa-scale-balanced",
                    "manager_only": True,
                },
                {
                    "title": "Balance Sheet",
                    "description": "Summarize assets, liabilities, and equity.",
                    "url_name": "core:balance_sheet",
                    "icon": "fa-table-columns",
                    "manager_only": True,
                },
                {
                    "title": "Income Statement",
                    "description": "Review revenue, expenses, and net income.",
                    "url_name": "core:income_statement",
                    "icon": "fa-chart-line",
                    "manager_only": True,
                },
                {
                    "title": "Receivables Aging",
                    "description": "Group unpaid invoice balances by overdue age.",
                    "url_name": "core:receivables_aging",
                    "icon": "fa-clock-rotate-left",
                    "manager_only": True,
                },
                {
                    "title": "Automated Reconciliation",
                    "description": "Compare invoices, payments, expenses, and payroll against posted journals.",
                    "url_name": "core:reconciliation_report",
                    "icon": "fa-code-compare",
                    "manager_only": True,
                },
                {
                    "title": "Payroll Reconciliation",
                    "description": "Review payroll records against payroll journal postings.",
                    "url_name": "core:payroll_reconciliation_report",
                    "icon": "fa-money-check-dollar",
                    "manager_only": True,
                },
                {
                    "title": "Expense Reconciliation",
                    "description": "Review operating expenses against expense journal postings.",
                    "url_name": "core:expense_reconciliation_report",
                    "icon": "fa-receipt",
                    "manager_only": True,
                },
                {
                    "title": "Payment Reconciliation",
                    "description": "Review received and paid amounts against payment journal postings.",
                    "url_name": "core:payment_reconciliation_report",
                    "icon": "fa-money-bill-transfer",
                    "manager_only": True,
                },
                {
                    "title": "Invoices",
                    "description": "Review billing, VAT, totals, balances, and invoice status.",
                    "url_name": "core:record_list",
                    "url_args": ["invoices"],
                    "icon": "fa-file-invoice-dollar",
                },
            ],
        ),
        (
            "Admin",
            [
                {
                    "title": "Audit Report",
                    "description": "Compare attendance, payroll, invoices, assets, recruitment, and training.",
                    "url_name": "core:audit_report",
                    "icon": "fa-clipboard-check",
                    "manager_only": True,
                },
            ],
        ),
    ]
    visible_groups = []
    for title, reports in report_groups:
        visible_reports = [report for report in reports if manager or not report.get("manager_only")]
        for report in visible_reports:
            report["url"] = reverse(report["url_name"], args=report.get("url_args", []))
        if visible_reports:
            visible_groups.append((title, visible_reports))

    summary_cards = [
        ("Employees", models.Employee.objects.count(), "fa-user-shield"),
        ("Sites", models.Site.objects.count(), "fa-building-shield"),
        ("Attendance Records", models.Attendance.objects.count(), "fa-calendar-check"),
        ("Invoices", models.Invoice.objects.count(), "fa-file-invoice-dollar"),
        ("Assets", models.Asset.objects.count(), "fa-boxes-stacked"),
        ("Incidents", models.Incident.objects.count(), "fa-triangle-exclamation"),
    ]

    return render(
        request,
        "core/reports/index.html",
        {
            "report_groups": visible_groups,
            "summary_cards": summary_cards,
        },
    )


@login_required
@user_passes_test(is_manager)
def payroll_export_excel(request):
    selected_month = request.GET.get("month") or timezone.localdate().isoformat()[:7]
    start, end = month_bounds(f"{selected_month}-01")
    generate_payroll_from_attendance(start, end)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Payroll"
    worksheet.append([f"Payroll Register: {start} to {end}"])
    worksheet.append([])
    worksheet.append(payroll_headers())
    for salary in payroll_queryset(start):
        worksheet.append(payroll_row(salary))
    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 24)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="payroll-{selected_month}.xlsx"'
    return response


@login_required
@user_passes_test(is_manager)
def payroll_export_pdf(request):
    selected_month = request.GET.get("month") or timezone.localdate().isoformat()[:7]
    start, end = month_bounds(f"{selected_month}-01")
    generate_payroll_from_attendance(start, end)
    output = BytesIO()
    document = SimpleDocTemplate(output, pagesize=landscape(A4), rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    elements = [Paragraph(f"Payroll Register: {start} to {end}", styles["Title"]), Spacer(1, 12)]
    data = [payroll_headers()]
    data.extend(payroll_row(salary) for salary in payroll_queryset(start))
    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2a3f54")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d9dee4")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    elements.append(table)
    document.build(elements)
    response = HttpResponse(output.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="payroll-{selected_month}.pdf"'
    return response


DEPLOYMENT_EXPORT_HEADERS = [
    "employee_number",
    "employee_name",
    "client",
    "site_code",
    "site_name",
    "shift_code",
    "shift_name",
    "supervisor_number",
    "supervisor_name",
    "start_date",
    "end_date",
    "status",
]


@login_required
@user_passes_test(lambda user: is_manager(user) or is_supervisor(user))
def deployments_export_excel(request):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Deployments"
    worksheet.append(DEPLOYMENT_EXPORT_HEADERS)
    deployments = models.Deployment.objects.select_related(
        "employee",
        "client",
        "site",
        "shift",
        "supervisor",
    ).order_by("-start_date", "employee__first_name", "employee__last_name")
    for deployment in deployments:
        worksheet.append(
            [
                deployment.employee.company_number,
                deployment.employee.full_name,
                deployment.client.client_name,
                deployment.site.site_code,
                deployment.site.site_name,
                deployment.shift.code,
                deployment.shift.shift_name,
                deployment.supervisor.company_number if deployment.supervisor else "",
                deployment.supervisor.full_name if deployment.supervisor else "",
                deployment.start_date,
                deployment.end_date,
                deployment.status,
            ]
        )
    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 28)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="deployments.xlsx"'
    return response


@login_required
@user_passes_test(lambda user: is_manager(user) or is_supervisor(user))
def deployments_import_excel(request):
    if request.method == "POST":
        deployment_file = request.FILES.get("deployment_file")
        if not deployment_file:
            messages.error(request, "Please choose an Excel deployment file to upload.")
            return redirect("core:deployments_import_excel")
        try:
            validate_excel_upload(deployment_file)
        except ValidationError as error:
            messages.error(request, " ".join(error.messages))
            return redirect("core:deployments_import_excel")
        try:
            workbook = load_workbook(deployment_file, data_only=True)
            worksheet = workbook.active
        except Exception:
            messages.error(request, "The uploaded file could not be read as an Excel workbook.")
            return redirect("core:deployments_import_excel")

        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), [])
        headers = {normalized_header(value): index for index, value in enumerate(header_row)}
        required_groups = [
            ("employee_number", "company_number", "guard_company_number"),
            ("site_code", "site", "site_name"),
            ("shift_code", "shift", "shift_name"),
            ("start_date", "deployment_date"),
        ]
        missing = [group[0] for group in required_groups if not any(name in headers for name in group)]
        if missing:
            messages.error(request, "Missing required deployment columns: " + ", ".join(missing).replace("_", " "))
            return redirect("core:deployments_import_excel")

        status_values = {choice.value for choice in models.StatusChoices}
        status_labels = {choice.label.lower(): choice.value for choice in models.StatusChoices}
        created_count = 0
        updated_count = 0
        skipped_rows = []

        for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            employee_value = value_from_row(row, headers, "employee_number", "company_number", "guard_company_number")
            site_value = value_from_row(row, headers, "site_code", "site", "site_name")
            shift_value = value_from_row(row, headers, "shift_code", "shift", "shift_name")
            supervisor_value = value_from_row(row, headers, "supervisor_number", "supervisor_company_number")
            status_value = value_from_row(row, headers, "status").lower()
            start_date = date_from_row(row, headers, "start_date", "deployment_date")
            end_date = date_from_row(row, headers, "end_date")

            employee = models.Employee.objects.filter(
                Q(company_number__iexact=employee_value)
                | Q(national_id__iexact=employee_value)
                | Q(first_name__iexact=employee_value)
                | Q(last_name__iexact=employee_value)
            ).first()
            site = models.Site.objects.select_related("client").filter(
                Q(site_code__iexact=site_value) | Q(site_name__iexact=site_value)
            ).first()
            shift = models.Shift.objects.filter(Q(code__iexact=shift_value) | Q(shift_name__iexact=shift_value)).first()
            supervisor = None
            if supervisor_value:
                supervisor = models.Employee.objects.filter(
                    Q(company_number__iexact=supervisor_value)
                    | Q(national_id__iexact=supervisor_value)
                    | Q(first_name__iexact=supervisor_value)
                    | Q(last_name__iexact=supervisor_value)
                ).first()
                if not supervisor:
                    skipped_rows.append(f"Row {row_number}: supervisor not found")
                    continue
            status = status_labels.get(status_value, status_value or models.StatusChoices.ACTIVE)
            if status not in status_values:
                skipped_rows.append(f"Row {row_number}: invalid status")
                continue
            if not employee or not site or not shift or not start_date:
                skipped_rows.append(
                    f"Row {row_number}: "
                    f"{'employee not found' if not employee else ''} "
                    f"{'site not found' if not site else ''} "
                    f"{'shift not found' if not shift else ''} "
                    f"{'invalid start date' if not start_date else ''}".strip()
                )
                continue

            _deployment, created = models.Deployment.objects.update_or_create(
                employee=employee,
                site=site,
                shift=shift,
                start_date=start_date,
                defaults={
                    "client": site.client,
                    "supervisor": supervisor,
                    "end_date": end_date,
                    "status": status,
                },
            )
            if created:
                created_count += 1
            else:
                updated_count += 1

        if created_count or updated_count:
            messages.success(
                request,
                f"Deployments imported: {created_count} created, {updated_count} updated.",
            )
        if skipped_rows:
            messages.error(request, "Skipped rows: " + "; ".join(skipped_rows[:8]))
        return redirect("core:record_list", slug="deployments")

    return render(request, "core/import_deployments.html")


@login_required
@user_passes_test(is_manager)
def invoice_pdf(request, pk):
    invoice = get_object_or_404(
        models.Invoice.objects.select_related("client", "contract", "site"),
        pk=pk,
    )
    output = BytesIO()
    document = SimpleDocTemplate(output, pagesize=A4, rightMargin=34, leftMargin=34, topMargin=72, bottomMargin=42)
    styles = pdf_styles()

    billing_label = "All contract sites" if invoice.billing_scope == models.Invoice.BillingScope.CONTRACT else (
        ", ".join(invoice.selected_sites.order_by("site_name").values_list("site_name", flat=True))
        if invoice.billing_scope == models.Invoice.BillingScope.MULTIPLE_SITES
        else invoice.site.site_name if invoice.site_id else "One site"
    )
    if not billing_label:
        billing_label = "Selected sites"
    status_color = PDF_GREEN if invoice.status == models.StatusChoices.PAID else colors.HexColor("#dc2626")
    status_table = Table(
        [[invoice.get_status_display().upper()]],
        colWidths=[96],
        hAlign="RIGHT",
    )
    status_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), status_color),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ]
        )
    )

    supplier_details = [
        ["Supplier", "Security Company Management"],
        ["Address", "Kampala, Uganda"],
        ["Email", "accounts@security-company.local"],
        ["Phone", "+256 700 000 000"],
        ["TIN", "-"],
    ]
    title = Table(
        [
            [
                Paragraph("TAX INVOICE", styles["DocTitle"]),
                status_table,
            ],
            [
                Paragraph(f"Invoice No: <b>{invoice.invoice_number}</b>", styles["SmallMuted"]),
                Paragraph(f"Balance Due: <b>{money_display(invoice.balance_amount)}</b>", styles["Right"]),
            ],
        ],
        colWidths=[330, 190],
    )
    title.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))

    client_details = [
        ["Customer", invoice.client_name or invoice.client.client_name],
        ["Contact", invoice.client_contact_person or "-"],
        ["Phone", invoice.client_phone_number or "-"],
        ["Email", invoice.client_email or "-"],
        ["Address", invoice.client_address or "-"],
    ]
    invoice_details = [
        ["Invoice Date", invoice.invoice_date],
        ["Due Date", invoice.due_date],
        ["Billing Month", invoice.billing_month or "-"],
        ["Contract", invoice.contract.contract_number if invoice.contract_id else "-"],
        ["Service Scope", billing_label],
    ]
    details = Table(
        [
            [
                Paragraph("From", styles["SectionTitle"]),
                Paragraph("Bill To", styles["SectionTitle"]),
                Paragraph("Invoice Details", styles["SectionTitle"]),
            ],
            [
                key_value_table(supplier_details, [58, 106]),
                key_value_table(client_details, [58, 106]),
                key_value_table(invoice_details, [72, 92]),
            ],
        ],
        colWidths=[172, 172, 172],
    )
    details.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))

    service_period = invoice.billing_month.strftime("%B %Y") if invoice.billing_month else "-"
    line_items = [["No.", "Description", "Service Period", "Qty", "Unit Rate", "Amount"]]
    item_map = [
        ("Guarding services", invoice.guard_count, invoice.rate_per_guard),
        ("Guns", invoice.gun_count, invoice.gun_rate),
        ("Radios", invoice.radio_count, invoice.radio_rate),
        ("Metal detectors", invoice.metal_detector_count, invoice.metal_detector_rate),
        ("Walk through machines", invoice.walk_through_machine_count, invoice.walk_through_machine_rate),
        ("Dogs", invoice.dog_count, invoice.dog_rate),
    ]
    item_number = 1
    for description, count, rate in item_map:
        if count:
            line_items.append(
                [
                    item_number,
                    description,
                    service_period,
                    count,
                    money_display(rate),
                    money_display(Decimal(count) * Decimal(rate or 0)),
                ]
            )
            item_number += 1
    if len(line_items) == 1:
        line_items.append(
            [1, "Security services", service_period, 1, money_display(invoice.subtotal_amount), money_display(invoice.subtotal_amount)]
        )

    totals = [
        ["Subtotal", money_display(invoice.subtotal_amount)],
        [f"VAT ({invoice.vat_rate * 100:.0f}%)", money_display(invoice.vat_amount)],
        ["Total", money_display(invoice.total_amount)],
        ["Paid", money_display(invoice.paid_amount)],
        ["Balance Due", money_display(invoice.balance_amount)],
    ]
    totals_table = Table(totals, colWidths=[120, 130], hAlign="RIGHT")
    totals_table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                ("TEXTCOLOR", (0, 0), (-1, -1), PDF_TEXT),
                ("GRID", (0, 0), (-1, -1), 0.35, PDF_LINE),
                ("BACKGROUND", (0, -1), (-1, -1), PDF_NAVY),
                ("TEXTCOLOR", (0, -1), (-1, -1), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )

    payment_terms = Table(
        [
            [
                Paragraph("Payment Instructions", styles["SectionTitle"]),
                Paragraph("Terms & Notes", styles["SectionTitle"]),
            ],
            [
                Paragraph(
                    f"Please pay the balance due by <b>{invoice.due_date}</b> and quote invoice "
                    f"<b>{invoice.invoice_number}</b> on all payments.",
                    styles["SmallMuted"],
                ),
                Paragraph(
                    "This invoice is generated from approved contract billing details. Amounts are stated in Uganda Shillings and include VAT where applicable.",
                    styles["SmallMuted"],
                ),
            ],
        ],
        colWidths=[250, 250],
    )
    payment_terms.setStyle(
        TableStyle(
            [
                ("BOX", (0, 0), (-1, -1), 0.35, PDF_LINE),
                ("INNERGRID", (0, 0), (-1, -1), 0.35, PDF_LINE),
                ("BACKGROUND", (0, 0), (-1, 0), PDF_SOFT),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ]
        )
    )

    elements = [
        title,
        Spacer(1, 14),
        details,
        Spacer(1, 14),
        Paragraph("Invoice Line Items", styles["SectionTitle"]),
        styled_table(line_items, col_widths=[32, 168, 90, 42, 86, 86], right_columns=(3, 4, 5)),
        Spacer(1, 12),
        totals_table,
        Spacer(1, 18),
        payment_terms,
        Spacer(1, 20),
        signature_table("Issued By", "Customer Acknowledgement"),
    ]
    document.build(
        elements,
        onFirstPage=pdf_header_footer("Invoice"),
        onLaterPages=pdf_header_footer("Invoice"),
    )
    response = HttpResponse(output.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="invoice-{invoice.invoice_number}.pdf"'
    return response


@login_required
@user_passes_test(is_manager)
def payslip_pdf(request, pk):
    salary = get_object_or_404(models.Salary.objects.select_related("employee", "employee__position"), pk=pk)
    refresh_payroll_for_date(salary.pay_period_start)
    salary.refresh_from_db()
    output = BytesIO()
    document = SimpleDocTemplate(output, pagesize=A4, rightMargin=36, leftMargin=36, topMargin=76, bottomMargin=42)
    styles = pdf_styles()
    employee = salary.employee
    title = Table(
        [
            [
                Paragraph("EMPLOYEE PAYSLIP", styles["DocTitle"]),
                Paragraph(f"Net Pay<br/><b>{money_display(salary.net_salary)}</b>", styles["Right"]),
            ],
            [
                Paragraph(f"Pay period: <b>{salary.pay_period_start} to {salary.pay_period_end}</b>", styles["SmallMuted"]),
                Paragraph(f"Status: <b>{salary.get_status_display()}</b>", styles["Right"]),
            ],
        ],
        colWidths=[330, 190],
    )
    title.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))

    employee_rows = [
        ["Employee", employee.full_name],
        ["Company Number", employee.company_number or "-"],
        ["NSSF Number", employee.nssf_number or "-"],
        ["Bank Account", employee.bank_account or "-"],
        ["Position", employee.position.position_title if employee.position else "-"],
        ["Payment Date", salary.payment_date or "-"],
        ["Payment Method", salary.payment_method or "-"],
    ]
    attendance_rows = [
        ["Pay Period", f"{salary.pay_period_start} to {salary.pay_period_end}"],
        ["Attendance Days", salary.attendance_days],
        ["Basic Hours", salary.basic_hours],
        ["Overtime Hours", salary.overtime_hours],
    ]
    earnings = [
        ["Earnings", "Amount"],
        ["Basic Pay", money_display(salary.basic_salary)],
        ["Overtime Pay", money_display(salary.overtime_pay)],
        ["Allowances", money_display(salary.allowances)],
        ["Bonus", money_display(salary.bonus)],
        ["Gross Pay", money_display(salary.gross_pay)],
    ]
    deductions = [
        ["Deductions", "Amount"],
        ["NSSF Employee", money_display(salary.nssf_employee)],
        ["Other Deductions", money_display(salary.deductions)],
        ["Advance Deduction", money_display(salary.advance_deduction)],
        ["Advance Balance", money_display(salary.advance_balance)],
        ["Total Deductions", money_display(salary.total_deductions)],
        ["Employer NSSF", money_display(salary.nssf_employer)],
        ["Net Pay", money_display(salary.net_salary)],
    ]
    details = Table(
        [
            [
                Paragraph("Employee Details", styles["SectionTitle"]),
                Paragraph("Attendance Summary", styles["SectionTitle"]),
            ],
            [
                key_value_table(employee_rows, [90, 160]),
                key_value_table(attendance_rows, [100, 150]),
            ],
        ],
        colWidths=[260, 260],
    )
    details.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
    pay_tables = Table(
        [[styled_table(earnings, col_widths=[160, 90], right_columns=(1,)), styled_table(deductions, col_widths=[160, 90], right_columns=(1,))]],
        colWidths=[260, 260],
    )
    pay_tables.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
    net_summary = Table(
        [["NET SALARY PAYABLE", money_display(salary.net_salary)]],
        colWidths=[320, 200],
    )
    net_summary.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), PDF_GREEN),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 12),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                ("TOPPADDING", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
            ]
        )
    )
    elements = [
        title,
        Spacer(1, 14),
        details,
        Spacer(1, 14),
        pay_tables,
        Spacer(1, 14),
        net_summary,
        Spacer(1, 18),
        Paragraph("This payslip is system generated and confidential to the employee named above.", styles["CenterSmall"]),
        Spacer(1, 22),
        signature_table("Prepared By", "Employee Signature"),
    ]
    document.build(
        elements,
        onFirstPage=pdf_header_footer("Payslip"),
        onLaterPages=pdf_header_footer("Payslip"),
    )
    response = HttpResponse(output.getvalue(), content_type="application/pdf")
    filename = f"payslip-{employee.company_number or employee.id}-{salary.pay_period_start}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
@user_passes_test(is_manager)
def payment_receipt_pdf(request, pk):
    payment = get_object_or_404(
        models.Payment.objects.select_related("invoice__client", "employee"),
        pk=pk,
    )
    target = payment.invoice.client.client_name if payment.invoice_id else payment.employee.full_name if payment.employee_id else "General Payment"
    reference = payment.transaction_ref or f"PAY-{payment.id}"
    output = BytesIO()
    document = SimpleDocTemplate(output, pagesize=A4, rightMargin=36, leftMargin=36, topMargin=76, bottomMargin=42)
    styles = pdf_styles()
    title = Table(
        [
            [
                Paragraph("PAYMENT RECEIPT", styles["DocTitle"]),
                Paragraph(f"Amount Received<br/><b>{money_display(payment.amount)}</b>", styles["Right"]),
            ],
            [
                Paragraph(f"Receipt Ref: <b>{reference}</b>", styles["SmallMuted"]),
                Paragraph(f"Date: <b>{payment.payment_date}</b>", styles["Right"]),
            ],
        ],
        colWidths=[330, 190],
    )
    title.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
    receipt_rows = [
        ["Received From", target],
        ["Payment Date", payment.payment_date],
        ["Payment Method", payment.payment_method],
        ["Transaction Ref", payment.transaction_ref or "-"],
        ["Invoice", payment.invoice.invoice_number if payment.invoice_id else "-"],
        ["Employee", payment.employee.full_name if payment.employee_id else "-"],
    ]
    allocation_rows = [
        ["Amount", money_display(payment.amount)],
        ["Invoice Balance", money_display(payment.invoice.balance_amount) if payment.invoice_id else "-"],
        ["Remarks", payment.remarks or "-"],
    ]
    details = Table(
        [
            [
                Paragraph("Receipt Details", styles["SectionTitle"]),
                Paragraph("Allocation", styles["SectionTitle"]),
            ],
            [
                key_value_table(receipt_rows, [92, 158]),
                key_value_table(allocation_rows, [92, 158]),
            ],
        ],
        colWidths=[260, 260],
    )
    details.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
    elements = [
        title,
        Spacer(1, 16),
        details,
        Spacer(1, 20),
        Paragraph("This receipt confirms that the amount above was recorded in the ERP payment register.", styles["SmallMuted"]),
        Spacer(1, 24),
        signature_table("Received By", "Customer / Employee"),
    ]
    document.build(
        elements,
        onFirstPage=pdf_header_footer("Payment Receipt"),
        onLaterPages=pdf_header_footer("Payment Receipt"),
    )
    response = HttpResponse(output.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="receipt-{safe_filename(reference)}.pdf"'
    return response


@login_required
@user_passes_test(is_manager)
def bulk_document_download(request):
    selected_month = request.POST.get("month") or request.GET.get("month") or timezone.localdate().isoformat()[:7]
    selected_types = request.POST.getlist("document_types") or request.GET.getlist("document_types")
    if not selected_types:
        selected_types = ["payslips", "invoices", "receipts"]
    selected_types = [doc_type for doc_type in selected_types if doc_type in {"payslips", "invoices", "receipts"}]
    start, end = month_bounds(f"{selected_month}-01")

    counts = {
        "payslips": models.Salary.objects.filter(pay_period_start=start).count(),
        "invoices": models.Invoice.objects.filter(invoice_date__range=(start, end)).count(),
        "receipts": models.Payment.objects.filter(payment_date__range=(start, end)).count(),
    }
    if request.method != "POST":
        return render(
            request,
            "core/bulk_document_download.html",
            {
                "selected_month": selected_month,
                "selected_types": selected_types,
                "counts": counts,
            },
        )

    output = BytesIO()
    added = 0
    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        if "payslips" in selected_types:
            salaries = models.Salary.objects.select_related("employee").filter(pay_period_start=start).order_by("employee__first_name", "employee__last_name")[:200]
            for salary in salaries:
                filename = f"payslips/{safe_filename(salary.employee.company_number or salary.employee_id)}-{salary.pay_period_start}.pdf"
                archive.writestr(filename, response_pdf_bytes(payslip_pdf(request, salary.pk)))
                added += 1
        if "invoices" in selected_types:
            invoices = models.Invoice.objects.filter(invoice_date__range=(start, end)).order_by("invoice_number")[:200]
            for invoice in invoices:
                filename = f"invoices/invoice-{safe_filename(invoice.invoice_number)}.pdf"
                archive.writestr(filename, response_pdf_bytes(invoice_pdf(request, invoice.pk)))
                added += 1
        if "receipts" in selected_types:
            payments = models.Payment.objects.filter(payment_date__range=(start, end)).order_by("payment_date", "id")[:200]
            for payment in payments:
                filename = f"receipts/receipt-{safe_filename(payment.transaction_ref or f'PAY-{payment.id}')}.pdf"
                archive.writestr(filename, response_pdf_bytes(payment_receipt_pdf(request, payment.pk)))
                added += 1

    if not added:
        messages.error(request, "No documents matched the selected period and document types.")
        return redirect(f"{reverse('core:bulk_document_download')}?month={selected_month}")

    response = HttpResponse(output.getvalue(), content_type="application/zip")
    response["Content-Disposition"] = f'attachment; filename="documents-{selected_month}.zip"'
    return response


@login_required
def record_list(request, slug):
    config = get_model_config(slug)
    if not can_manage_slug(request.user, slug):
        return HttpResponseForbidden("You do not have permission to access this page.")
    queryset, search_query = record_queryset_for_request(request, slug, config)
    rows = record_rows(queryset, config.columns)

    return render(
        request,
        "core/record_list.html",
        {
            "slug": slug,
            "config": config,
            "columns": config.columns,
            "column_labels": [column_label(column) for column in config.columns],
            "rows": rows,
            "search_query": search_query,
            "can_add_record": is_manager(request.user) or is_supervisor(request.user) or has_model_perm(request.user, slug, "add"),
            "can_edit_record": is_manager(request.user) or is_supervisor(request.user) or has_model_perm(request.user, slug, "change"),
            "can_delete_record": is_manager(request.user) or is_supervisor(request.user) or has_model_perm(request.user, slug, "delete"),
        },
    )


@login_required
def record_list_pdf(request, slug):
    config = get_model_config(slug)
    if not can_manage_slug(request.user, slug):
        return HttpResponseForbidden("You do not have permission to export this report.")

    queryset, search_query = record_queryset_for_request(request, slug, config)
    rows = record_rows(queryset, config.columns)
    styles = pdf_styles()
    output = BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=24,
        leftMargin=24,
        topMargin=72,
        bottomMargin=42,
    )
    elements = [
        Paragraph(f"{xml_escape(config.title)} Report", styles["DocTitle"]),
        Paragraph(f"Department: {xml_escape(config.department)}", styles["SmallMuted"]),
        Paragraph(f"Generated: {timezone.localtime(timezone.now()):%Y-%m-%d %H:%M}", styles["SmallMuted"]),
    ]
    if search_query:
        elements.append(Paragraph(f"Search filter: {xml_escape(search_query)}", styles["SmallMuted"]))
    elements.extend(
        [
            Spacer(1, 10),
            key_value_table(
                [
                    ["Records shown", len(rows)],
                    ["Report limit", "200 rows"],
                ],
                [110, 160],
            ),
            Spacer(1, 12),
        ]
    )

    headers = [column_label(column) for column in config.columns]
    data = [[Paragraph(xml_escape(header), styles["SmallMuted"]) for header in headers]]
    max_text_length = 90 if len(headers) <= 8 else 55
    for row in rows:
        data.append(
            [
                Paragraph(xml_escape(stringify_cell(value)[:max_text_length]), styles["SmallMuted"])
                for value in row["values"]
            ]
        )
    if not rows:
        data.append([Paragraph("No records found.", styles["SmallMuted"])] + ["" for _ in headers[1:]])

    available_width = landscape(A4)[0] - document.leftMargin - document.rightMargin
    col_widths = [available_width / max(len(headers), 1)] * max(len(headers), 1)
    elements.append(styled_table(data, col_widths=col_widths))
    elements.extend([Spacer(1, 18), signature_table()])
    document.build(
        elements,
        onFirstPage=pdf_header_footer(f"{config.title} Report"),
        onLaterPages=pdf_header_footer(f"{config.title} Report"),
    )
    response = HttpResponse(output.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{slug}-report.pdf"'
    return response


@login_required
def record_create(request, slug):
    if not (is_manager(request.user) or is_supervisor(request.user) or has_model_perm(request.user, slug, "add")):
        return HttpResponseForbidden("You do not have permission to add this record.")
    if slug == "guard-schedules":
        messages.info(request, "Guard schedules are managed from the attendance screen.")
        return redirect("core:attendances")
    config = get_model_config(slug)
    form_class = build_model_form(config.model)
    form = form_class(request.POST or None, request.FILES or None)
    if request.method == "POST" and form.is_valid():
        if config.model is models.Invoice:
            instance = form.save()
        else:
            instance = form.save(commit=False)
            if hasattr(instance, "allocated_by_id") and not instance.allocated_by_id:
                instance.allocated_by = request.user
            instance.save()
            form.save_m2m()
        if config.model is models.Attendance:
            refresh_payroll_for_date(instance.date)
        messages.success(request, f"{config.title} record added successfully.")
        return redirect("core:record_list", slug=slug)
    template = "core/invoice_form.html" if config.model is models.Invoice else "core/record_form.html"
    return render(request, template, {"slug": slug, "config": config, "form": form, "mode": "Add"})


@login_required
def record_update(request, slug, pk):
    if not (is_manager(request.user) or is_supervisor(request.user) or has_model_perm(request.user, slug, "change")):
        return HttpResponseForbidden("You do not have permission to edit this record.")
    if slug == "guard-schedules":
        messages.info(request, "Guard schedules are managed from the attendance screen.")
        return redirect("core:attendances")
    config = get_model_config(slug)
    instance = get_object_or_404(scoped_queryset(request.user, slug, config.model.objects.all()), pk=pk)
    form_class = build_model_form(config.model)
    previous_attendance_date = instance.date if config.model is models.Attendance else None
    form = form_class(request.POST or None, request.FILES or None, instance=instance)
    if request.method == "POST" and form.is_valid():
        if config.model is models.Invoice:
            form.save()
        else:
            updated_instance = form.save(commit=False)
            updated_instance.save()
            form.save_m2m()
        if config.model is models.Attendance:
            refresh_payroll_for_dates(previous_attendance_date, updated_instance.date)
        messages.success(request, f"{config.title} record updated successfully.")
        return redirect("core:record_list", slug=slug)
    template = "core/invoice_form.html" if config.model is models.Invoice else "core/record_form.html"
    return render(
        request,
        template,
        {"slug": slug, "config": config, "form": form, "mode": "Edit", "record": instance},
    )


@login_required
def record_delete(request, slug, pk):
    if not (is_manager(request.user) or is_supervisor(request.user) or has_model_perm(request.user, slug, "delete")):
        return HttpResponseForbidden("You do not have permission to delete this record.")
    if slug == "guard-schedules":
        messages.info(request, "Guard schedules are managed from the attendance screen.")
        return redirect("core:attendances")
    config = get_model_config(slug)
    instance = get_object_or_404(scoped_queryset(request.user, slug, config.model.objects.all()), pk=pk)
    attendance_date = instance.date if config.model is models.Attendance else None
    if request.method == "POST":
        try:
            instance.delete()
            if config.model is models.Attendance:
                refresh_payroll_for_date(attendance_date)
            messages.success(request, f"{config.title} record deleted successfully.")
        except ProtectedError:
            messages.error(request, "This record is linked to other records and cannot be deleted.")
        return redirect("core:record_list", slug=slug)
    return render(
        request,
        "core/record_confirm_delete.html",
        {"slug": slug, "config": config, "record": instance},
    )


@login_required
@user_passes_test(can_manage_attendance)
def attendances(request):
    selected_site_id = request.GET.get("site") or request.POST.get("site")
    selected_date = request.GET.get("date") or request.POST.get("date") or timezone.localdate().isoformat()
    selected_site = None
    schedules = models.GuardSchedule.objects.none()

    site_queryset = models.Site.objects.select_related("client").order_by("site_name")
    if selected_site_id:
        selected_site = site_queryset.filter(id=selected_site_id).first()
        if not selected_site:
            return HttpResponseForbidden("You do not have permission to access this site.")

    parsed_date = parse_date(selected_date) if selected_date else None

    if parsed_date:
        deployments = models.Deployment.objects.select_related("employee", "site", "shift").filter(
            status=models.StatusChoices.ACTIVE,
            start_date__lte=parsed_date,
        ).filter(
            Q(end_date__gte=parsed_date) | Q(end_date__isnull=True)
        )
        if selected_site:
            deployments = deployments.filter(site=selected_site)

        skipped_contract_limits = []
        for deployment in deployments:
            existing_schedule = models.GuardSchedule.objects.filter(
                deployment=deployment,
                shift_date=parsed_date,
            ).first()
            if not existing_schedule and contract_schedule_limit_reached(
                deployment.site,
                deployment.shift,
                parsed_date,
                deployment=deployment,
            ):
                skipped_contract_limits.append(contract_limit_message(deployment.site, deployment.shift, parsed_date))
                continue
            models.GuardSchedule.objects.get_or_create(
                deployment=deployment,
                shift_date=parsed_date,
                defaults={
                    "employee": deployment.employee,
                    "site": deployment.site,
                    "shift": deployment.shift,
                },
            )
        if skipped_contract_limits:
            messages.error(request, "Skipped schedules beyond contract: " + "; ".join(sorted(set(skipped_contract_limits))[:5]))

        schedules = scoped_queryset(
            request.user,
            "guard-schedules",
            models.GuardSchedule.objects.select_related(
                "employee",
                "replacement_employee",
                "site",
                "shift",
                "attendance",
            ),
        ).filter(shift_date=parsed_date)
        if selected_site:
            schedules = schedules.filter(site=selected_site)
        schedules = schedules.order_by(
            "shift__start_time", "employee__first_name", "employee__last_name"
        )

    if request.method == "POST":
        schedule_ids = request.POST.getlist("schedule_ids")
        allowed_schedules = scoped_queryset(
            request.user,
            "guard-schedules",
            models.GuardSchedule.objects.select_related("employee", "shift"),
        )
        allowed_employees = models.Employee.objects.filter(company_number__isnull=False).exclude(company_number="").order_by(
            "first_name", "last_name", "company_number"
        )

        for schedule_id in schedule_ids:
            schedule = get_object_or_404(allowed_schedules.select_related("employee"), id=schedule_id)
            selected_employee_id = request.POST.get(f"scheduled_guard_{schedule_id}") or schedule.employee_id
            selected_employee = get_object_or_404(allowed_employees, id=selected_employee_id)
            selected_deployment = schedule.deployment
            if selected_employee.id != schedule.deployment.employee_id:
                selected_deployment = deployment_for_attendance_employee(
                    selected_employee,
                    schedule.site,
                    schedule.shift,
                    schedule.shift_date,
                )
                schedule.deployment = selected_deployment
            schedule.employee = selected_employee
            present_value = request.POST.get(f"present_{schedule_id}", "")
            is_present = present_value.lower() in {"on", "yes", "true", "1"}
            replacement_employee_id = request.POST.get(f"replacement_guard_{schedule_id}") or ""
            reason = request.POST.get(f"reason_{schedule_id}", "").strip()
            attendance = models.Attendance.objects.filter(schedule=schedule).first()
            if not attendance:
                attendance = models.Attendance.objects.filter(
                    employee=selected_employee,
                    date=schedule.shift_date,
                    shift=schedule.shift,
                    site=schedule.site,
                ).first()
            attendance_values = {
                "employee": selected_employee,
                "site": schedule.site,
                "schedule": schedule,
                "shift": schedule.shift,
                "date": schedule.shift_date,
                "status": "Present" if is_present else "Absent",
                "capture_source": models.Attendance.CaptureSource.MANUAL,
                "captured_by": request.user,
                "captured_at": timezone.now(),
                "remarks": reason,
            }
            if attendance:
                conflict = (
                    models.Attendance.objects.filter(
                        employee=selected_employee,
                        date=schedule.shift_date,
                        shift=schedule.shift,
                        site=schedule.site,
                    )
                    .exclude(pk=attendance.pk)
                    .first()
                )
                if conflict and conflict.schedule_id:
                    messages.error(
                        request,
                        f"{selected_employee.full_name} already has scheduled attendance for this date and shift.",
                    )
                    continue
                if conflict:
                    conflict.delete()
                for field, value in attendance_values.items():
                    setattr(attendance, field, value)
                attendance.save(
                    update_fields=[
                        "employee",
                        "site",
                        "schedule",
                        "shift",
                        "date",
                        "status",
                        "capture_source",
                        "captured_by",
                        "captured_at",
                        "remarks",
                        "updated_at",
                    ]
                )
            else:
                models.Attendance.objects.create(
                    **attendance_values,
                )
            schedule.replacement_employee = None
            schedule.replacement_reason = ""
            if is_present:
                schedule.status = models.GuardSchedule.ScheduleStatus.COMPLETED
                schedule.notes = reason
            else:
                schedule.status = models.GuardSchedule.ScheduleStatus.MISSED
                schedule.replacement_reason = reason
                schedule.notes = f"Absent: {reason}" if reason else "Absent"
                if replacement_employee_id:
                    replacement_employee = get_object_or_404(allowed_employees, id=replacement_employee_id)
                    deployment_for_attendance_employee(
                        replacement_employee,
                        schedule.site,
                        schedule.shift,
                        schedule.shift_date,
                    )
                    schedule.replacement_employee = replacement_employee
                    models.Attendance.objects.update_or_create(
                        employee=replacement_employee,
                        date=schedule.shift_date,
                        shift=schedule.shift,
                        site=schedule.site,
                        defaults={
                            "schedule": None,
                            "status": "Present",
                            "capture_source": models.Attendance.CaptureSource.MANUAL,
                            "captured_by": request.user,
                            "captured_at": timezone.now(),
                            "remarks": f"Replacement for {selected_employee.full_name}. {reason}".strip(),
                        },
                    )
            schedule.save(
                update_fields=[
                    "employee",
                    "deployment",
                    "replacement_employee",
                    "replacement_reason",
                    "status",
                    "notes",
                    "updated_at",
                ]
            )

        if parsed_date:
            refresh_payroll_for_date(parsed_date)
        messages.success(request, "Attendance saved successfully.")
        return redirect(f"/attendances/?site={selected_site_id or ''}&date={selected_date or ''}")

    context = {
        "sites": site_queryset,
        "selected_site_id": selected_site_id or "",
        "selected_date": selected_date or "",
        "selected_site": selected_site,
        "schedules": schedules,
        "guards": models.Employee.objects.filter(company_number__isnull=False).exclude(company_number="").order_by(
            "first_name", "last_name", "company_number"
        ),
    }
    return render(request, "core/attendances.html", context)


def deployment_for_attendance_employee(employee, site, shift, shift_date):
    existing = (
        models.Deployment.objects.filter(
            employee=employee,
            site=site,
            shift=shift,
            status=models.StatusChoices.ACTIVE,
            start_date__lte=shift_date,
        )
        .filter(Q(end_date__gte=shift_date) | Q(end_date__isnull=True))
        .order_by("-start_date")
        .first()
    )
    if existing:
        return existing
    return models.Deployment.objects.create(
        employee=employee,
        client=site.client,
        site=site,
        shift=shift,
        start_date=shift_date,
        end_date=shift_date,
        status=models.StatusChoices.ACTIVE,
    )


def normalized_header(value):
    return str(value or "").strip().lower().replace(" ", "_").replace("-", "_")


def value_from_row(row, headers, *names):
    for name in names:
        index = headers.get(name)
        if index is not None:
            if index >= len(row):
                continue
            value = row[index]
            if value not in (None, ""):
                return str(value).strip()
    return ""


def date_from_row(row, headers, *names):
    for name in names:
        index = headers.get(name)
        if index is not None:
            if index >= len(row):
                continue
            value = row[index]
            if value in (None, ""):
                continue
            if hasattr(value, "date"):
                return value.date()
            parsed = parse_date(str(value).strip())
            if parsed:
                return parsed
    return None


def month_date_range(roster_month):
    month_start = parse_date(f"{roster_month}-01") if roster_month else None
    if not month_start:
        month_start = timezone.localdate().replace(day=1)
    month_end = month_start.replace(day=calendar.monthrange(month_start.year, month_start.month)[1])
    return month_start, month_end


def date_range_days(start, end):
    current = start
    while current <= end:
        yield current
        current += timedelta(days=1)


def schedule_dates_from_row(row, headers, roster_month):
    shift_date = date_from_row(row, headers, "shift_date", "date", "duty_date")
    if shift_date:
        return [shift_date]

    start, month_end = month_date_range(roster_month)
    start_date = date_from_row(row, headers, "start_date", "from_date")
    end_date = date_from_row(row, headers, "end_date", "to_date")
    start = start_date or start
    end = end_date or month_end
    if end < start:
        return []
    return list(date_range_days(start, end))


def rows_from_csv_upload(uploaded_file):
    content = uploaded_file.read().decode("utf-8-sig")
    return list(csv.reader(content.splitlines()))


def worksheet_rows(worksheet):
    return list(worksheet.iter_rows(values_only=True))


def schedule_file_rows(uploaded_file):
    if file_extension(uploaded_file) == ".csv":
        return rows_from_csv_upload(uploaded_file), None
    workbook = load_workbook(uploaded_file, data_only=True)
    worksheet = workbook.active
    return worksheet_rows(worksheet), worksheet


def day_number_from_header(value):
    match = re.match(r"^(\d{1,2})(?:st|nd|rd|th)?$", str(value or "").strip().lower())
    if not match:
        return None
    day = int(match.group(1))
    return day if 1 <= day <= 31 else None


def monthly_grid_day_columns(header_row):
    columns = []
    for index, value in enumerate(header_row):
        day = day_number_from_header(value)
        if day:
            columns.append((index, day))
    return columns


def find_employee_for_schedule_upload(guard_value):
    return models.Employee.objects.filter(
        Q(company_number__iexact=guard_value)
        | Q(work_card_uid__iexact=guard_value)
        | Q(national_id__iexact=guard_value)
        | Q(first_name__iexact=guard_value)
        | Q(last_name__iexact=guard_value)
    ).first()


def find_site_for_schedule_upload(site_value):
    return models.Site.objects.select_related("client").filter(Q(site_code__iexact=site_value)).first()


def get_import_client():
    return models.Client.objects.get_or_create(
        client_name="Imported Duty Roster Client",
        defaults={
            "contact_person": "Imported roster",
            "phone_number": "0000000000",
            "contract_start_date": timezone.localdate(),
            "contract_status": models.StatusChoices.ACTIVE,
        },
    )[0]


def get_import_contract(client, site=None, start_date=None, end_date=None):
    suffix = f"-{site.site_code}" if site and site.site_code else ""
    contract_number = f"IMP-{client.id:05d}{suffix}"
    return models.Contract.objects.get_or_create(
        contract_number=contract_number[:80],
        defaults={
            "client": client,
            "service_type": "Manned Guarding",
            "start_date": start_date or timezone.localdate(),
            "end_date": end_date,
            "status": models.StatusChoices.ACTIVE,
        },
    )[0]


def ensure_contract_requirement(site, shift_date, required_guards=1, shift=None, end_date=None):
    contract = get_import_contract(site.client, site=site, start_date=shift_date, end_date=end_date)
    requirement, created = models.ContractSiteRequirement.objects.get_or_create(
        contract=contract,
        site=site,
        shift=shift,
        start_date=contract.start_date,
        defaults={
            "required_guards": required_guards,
            "end_date": end_date or contract.end_date,
            "status": models.StatusChoices.ACTIVE,
            "notes": "Created from imported duty roster.",
        },
    )
    if requirement.required_guards < required_guards:
        requirement.required_guards = required_guards
        requirement.save(update_fields=["required_guards", "updated_at"])
    return requirement


def get_import_guard(pers_no, guard_name, row_number=None, site_code=""):
    role = models.Role.objects.get_or_create(
        role_name="Imported Guard",
        defaults={"department": models.DepartmentChoices.OPERATIONS},
    )[0]
    position = models.Position.objects.get_or_create(
        position_title="Imported Security Guard",
        defaults={"department": models.DepartmentChoices.OPERATIONS},
    )[0]
    clean_pers_no = str(pers_no or "").strip() or f"ROW-{row_number}"
    company_number = clean_pers_no
    if models.Employee.objects.filter(company_number=company_number).exclude(first_name__iexact=str(guard_name).strip()).exists():
        company_number = f"{site_code}-{clean_pers_no}-{row_number}".strip("-")
    existing_employee = models.Employee.objects.filter(company_number=company_number).first()
    if existing_employee:
        return existing_employee

    parts = str(guard_name or "Imported Guard").strip().split()
    first_name = parts[0] if parts else "Imported"
    last_name = " ".join(parts[1:]) if len(parts) > 1 else "Guard"
    employee, _created = models.Employee.objects.get_or_create(
        national_id=f"IMPORTED-{company_number}",
        defaults={
            "first_name": first_name,
            "last_name": last_name,
            "phone_number": "0000000000",
            "email": f"imported-{company_number}".lower().replace(" ", "-") + "@demo.test",
            "role": role,
            "position": position,
            "status": models.StatusChoices.ACTIVE,
            "company_number": company_number,
        },
    )
    if not employee.company_number:
        employee.company_number = company_number
        employee.save(update_fields=["company_number", "updated_at"])
    return employee


def get_import_shift(shift_code):
    code = str(shift_code or "").strip().upper()
    if code in {"D", "DAY"}:
        return models.Shift.objects.get_or_create(
            code="D",
            defaults={"shift_name": "Day", "start_time": "08:00", "end_time": "20:00"},
        )[0]
    if code in {"N", "NIGHT"}:
        return models.Shift.objects.get_or_create(
            code="N",
            defaults={"shift_name": "Night", "start_time": "18:00", "end_time": "06:00"},
        )[0]
    return models.Shift.objects.get_or_create(
        code=code,
        defaults={"shift_name": code.title(), "start_time": "08:00", "end_time": "17:00"},
    )[0]


def is_off_duty_code(value):
    return str(value or "").strip().upper() in {"O", "OFF"}


def record_roster_attendance(
    *,
    batch_reference,
    file_name,
    source_format,
    source_row,
    uploaded_by,
    import_status,
    message="",
    employee=None,
    site=None,
    shift=None,
    shift_date=None,
    schedule=None,
    duty_code="",
):
    return models.RosterAttendance.objects.create(
        batch_reference=batch_reference,
        file_name=file_name,
        source_format=source_format,
        source_row=source_row,
        employee=employee,
        site=site,
        shift=shift,
        shift_date=shift_date,
        schedule=schedule,
        duty_code=str(duty_code or "")[:40],
        import_status=import_status,
        message=message,
        uploaded_by=uploaded_by if getattr(uploaded_by, "is_authenticated", False) else None,
    )


def supervisor_for_site(site):
    allocation = (
        models.ZoneSiteAllocation.objects.filter(
            site=site,
            status=models.StatusChoices.ACTIVE,
            end_date__isnull=True,
        )
        .select_related("zone__supervisor")
        .first()
    )
    return allocation.zone.supervisor if allocation else None


def upsert_guard_schedule_from_upload(employee, site, shift, shift_date, *, deployment_start, deployment_end, notes):
    existing_schedule = models.GuardSchedule.objects.filter(
        employee=employee,
        site=site,
        shift=shift,
        shift_date=shift_date,
    ).first()
    if not existing_schedule and contract_schedule_limit_reached(site, shift, shift_date):
        return None, False, contract_limit_message(site, shift, shift_date)

    deployment, _created = models.Deployment.objects.update_or_create(
        employee=employee,
        site=site,
        shift=shift,
        start_date=deployment_start,
        defaults={
            "client": site.client,
            "supervisor": supervisor_for_site(site),
            "end_date": deployment_end,
            "status": models.StatusChoices.ACTIVE,
        },
    )
    schedule, created = models.GuardSchedule.objects.update_or_create(
        deployment=deployment,
        shift_date=shift_date,
        defaults={
            "employee": employee,
            "site": site,
            "shift": shift,
            "status": models.GuardSchedule.ScheduleStatus.SCHEDULED,
            "notes": notes,
        },
    )
    return schedule, created, ""


def import_monthly_schedule_grid(rows, headers, *, roster_month, file_name="", uploaded_by=None, batch_reference=None):
    batch_reference = batch_reference or str(uuid.uuid4())
    month_start, month_end = month_date_range(roster_month)
    day_columns = monthly_grid_day_columns(rows[0] if rows else [])
    created_schedules = 0
    updated_schedules = 0
    off_rows = 0
    skipped_rows = []

    for row_number, row in enumerate(rows[1:], start=2):
        if not any(row):
            continue
        guard_value = value_from_row(
            row,
            headers,
            "guard_id",
            "guard_company_number",
            "company_number",
            "employee_number",
            "guard_badge",
            "badge_number",
            "guard",
        )
        site_value = value_from_row(row, headers, "site_code")
        base_shift_value = value_from_row(row, headers, "shift", "shift_name")

        if not guard_value and not site_value:
            continue
        employee = find_employee_for_schedule_upload(guard_value)
        site = find_site_for_schedule_upload(site_value)
        base_shift = (
            models.Shift.objects.filter(Q(shift_name__iexact=base_shift_value) | Q(code__iexact=base_shift_value)).first()
            if base_shift_value
            else None
        )
        if not employee or not site:
            message = f"{'guard not found' if not employee else ''} {'site not found' if not site else ''}".strip()
            skipped_rows.append(f"Row {row_number}: {message}")
            record_roster_attendance(
                batch_reference=batch_reference,
                file_name=file_name,
                source_format=models.RosterAttendance.SourceFormat.SIMPLE,
                source_row=row_number,
                uploaded_by=uploaded_by,
                import_status=models.RosterAttendance.ImportStatus.SKIPPED,
                message=message,
                employee=employee,
                site=site,
                shift=base_shift,
                duty_code=base_shift_value,
            )
            continue

        duty_dates = [
            month_start.replace(day=day)
            for column, day in day_columns
            if column < len(row)
            and str(row[column] or "").strip()
            and not is_off_duty_code(row[column])
            and day <= month_end.day
        ]
        deployment_start = min(duty_dates) if duty_dates else month_start
        deployment_end = max(duty_dates) if duty_dates else month_end

        for column, day in day_columns:
            if column >= len(row) or day > month_end.day:
                continue
            duty_code = str(row[column] or "").strip().upper()
            if not duty_code:
                continue
            shift_date = month_start.replace(day=day)
            if is_off_duty_code(duty_code):
                record_roster_attendance(
                    batch_reference=batch_reference,
                    file_name=file_name,
                    source_format=models.RosterAttendance.SourceFormat.SIMPLE,
                    source_row=row_number,
                    uploaded_by=uploaded_by,
                    import_status=models.RosterAttendance.ImportStatus.OFF,
                    message="Off day from monthly scheduled-guard grid.",
                    employee=employee,
                    site=site,
                    shift=base_shift,
                    shift_date=shift_date,
                    duty_code="O",
                )
                off_rows += 1
                continue

            shift = get_import_shift(duty_code or base_shift_value)
            schedule, created, error_message = upsert_guard_schedule_from_upload(
                employee,
                site,
                shift,
                shift_date,
                deployment_start=deployment_start,
                deployment_end=deployment_end,
                notes="Imported from monthly scheduled-guard grid.",
            )
            if error_message:
                skipped_rows.append(f"Row {row_number} {shift_date}: {error_message}")
                record_roster_attendance(
                    batch_reference=batch_reference,
                    file_name=file_name,
                    source_format=models.RosterAttendance.SourceFormat.SIMPLE,
                    source_row=row_number,
                    uploaded_by=uploaded_by,
                    import_status=models.RosterAttendance.ImportStatus.SKIPPED,
                    message=error_message,
                    employee=employee,
                    site=site,
                    shift=shift,
                    shift_date=shift_date,
                    duty_code=duty_code,
                )
                continue
            if created:
                created_schedules += 1
                import_status = models.RosterAttendance.ImportStatus.CREATED
            else:
                updated_schedules += 1
                import_status = models.RosterAttendance.ImportStatus.UPDATED
            record_roster_attendance(
                batch_reference=batch_reference,
                file_name=file_name,
                source_format=models.RosterAttendance.SourceFormat.SIMPLE,
                source_row=row_number,
                uploaded_by=uploaded_by,
                import_status=import_status,
                message="Imported from monthly scheduled-guard grid.",
                employee=employee,
                site=site,
                shift=shift,
                shift_date=shift_date,
                schedule=schedule,
                duty_code=duty_code,
            )

    return created_schedules, updated_schedules, off_rows, skipped_rows, batch_reference


def roster_header_date(header, period_start, period_end):
    match = re.search(r"/(\d{1,2})$", str(header or "").strip())
    if not match:
        return None
    day = int(match.group(1))
    candidate = period_start.replace(day=day)
    if candidate < period_start:
        if candidate.month == 12:
            candidate = candidate.replace(year=candidate.year + 1, month=1)
        else:
            candidate = candidate.replace(month=candidate.month + 1)
    if period_start <= candidate <= period_end:
        return candidate
    return None


def monthly_date_from_day(day_value, month_start):
    try:
        day = int(day_value)
    except (TypeError, ValueError):
        return None
    try:
        return month_start.replace(day=day)
    except ValueError:
        return None


def find_wide_monthly_roster_header(worksheet):
    rows = list(worksheet.iter_rows(values_only=True))
    for index, row in enumerate(rows):
        normalized = [normalized_header(value) for value in row]
        if {"site_code", "site_name", "shift"}.issubset(set(normalized)) and index + 1 < len(rows):
            header_map = {header: column for column, header in enumerate(normalized) if header}
            date_columns = []
            for column, value in enumerate(rows[index + 1]):
                if monthly_date_from_day(value, timezone.localdate().replace(day=1)):
                    date_columns.append((column, value))
            if date_columns:
                return rows, index, header_map, date_columns
    return rows, None, {}, []


def import_wide_monthly_roster(worksheet, *, roster_month, file_name="", uploaded_by=None, batch_reference=None):
    batch_reference = batch_reference or str(uuid.uuid4())
    month_start = parse_date(f"{roster_month}-01") if roster_month else timezone.localdate().replace(day=1)
    if not month_start:
        month_start = timezone.localdate().replace(day=1)
    client = get_import_client()
    stored_rows = 0
    off_rows = 0
    skipped_rows = []
    rows, header_index, headers, date_columns = find_wide_monthly_roster_header(worksheet)
    if header_index is None:
        return stored_rows, off_rows, skipped_rows, batch_reference

    site_code_index = headers.get("site_code")
    site_name_index = headers.get("site_name")
    shift_index = headers.get("shift")
    current_site = None
    current_shift_code = ""

    for row_number, row in enumerate(rows[header_index + 2 :], start=header_index + 3):
        if not any(row):
            continue
        site_code = str(row[site_code_index] or "").strip() if site_code_index is not None else ""
        site_name = str(row[site_name_index] or "").strip() if site_name_index is not None else ""
        shift_code = str(row[shift_index] or "").strip().upper() if shift_index is not None else ""
        if site_code or site_name:
            lookup_name = site_name or site_code
            current_site = models.Site.objects.filter(Q(site_code__iexact=site_code) | Q(site_name__iexact=lookup_name)).first()
            if not current_site:
                current_site = models.Site.objects.create(
                    client=client,
                    site_code=(site_code or f"IMP{row_number}")[:20],
                    site_name=lookup_name,
                    site_address=lookup_name,
                    city="Imported",
                    security_level="Imported",
                )
        if shift_code:
            current_shift_code = shift_code
        if not current_site:
            skipped_rows.append(f"Row {row_number}: site not found")
            continue

        for column, day_value in date_columns:
            duty_code = str(row[column] or "").strip().upper()
            if not duty_code:
                continue
            shift_date = monthly_date_from_day(day_value, month_start)
            if not shift_date:
                skipped_rows.append(f"Row {row_number}: invalid day {day_value}")
                continue
            if is_off_duty_code(duty_code):
                record_roster_attendance(
                    batch_reference=batch_reference,
                    file_name=file_name,
                    source_format=models.RosterAttendance.SourceFormat.WIDE_MONTHLY,
                    source_row=row_number,
                    uploaded_by=uploaded_by,
                    import_status=models.RosterAttendance.ImportStatus.OFF,
                    message="Off duty from monthly roster.",
                    site=current_site,
                    shift_date=shift_date,
                    duty_code="O",
                )
                off_rows += 1
                continue

            shift = get_import_shift(duty_code or current_shift_code)
            record_roster_attendance(
                batch_reference=batch_reference,
                file_name=file_name,
                source_format=models.RosterAttendance.SourceFormat.WIDE_MONTHLY,
                source_row=row_number,
                uploaded_by=uploaded_by,
                import_status=models.RosterAttendance.ImportStatus.CREATED,
                message="Stored monthly roster duty. Assign a guard from roster attendance scheduling.",
                site=current_site,
                shift=shift,
                shift_date=shift_date,
                duty_code=duty_code,
            )
            stored_rows += 1

    return stored_rows, off_rows, skipped_rows, batch_reference


def import_saracen_roster(worksheet, *, file_name="", uploaded_by=None, batch_reference=None):
    client = get_import_client()
    batch_reference = batch_reference or str(uuid.uuid4())
    created_schedules = 0
    updated_schedules = 0
    skipped_rows = []
    current_site = None
    period_start = None
    period_end = None
    header_row = None
    date_columns = []

    rows = list(worksheet.iter_rows(values_only=True))
    for row_number, row in enumerate(rows, start=1):
        first_cell = str(row[0] or "").strip()
        site_match = re.search(r"Site Roster for\s+([^:]+):\s*(.+)$", first_cell, re.IGNORECASE)
        if site_match:
            site_code = site_match.group(1).strip()
            site_name = site_match.group(2).strip()
            current_site = models.Site.objects.update_or_create(
                client=client,
                site_name=site_name,
                defaults={
                    "site_code": site_code[:20],
                    "site_address": site_name,
                    "city": "Fort Portal",
                    "state": "Western",
                    "security_level": "Imported",
                },
            )[0]
            if period_start:
                ensure_contract_requirement(current_site, period_start, required_guards=1, end_date=period_end)
            header_row = None
            date_columns = []
            continue

        period_match = re.search(r"Scheduled Period:\s*(\d{4}-\d{2}-\d{2})\s+to\s+(\d{4}-\d{2}-\d{2})", first_cell)
        if period_match:
            period_start = parse_date(period_match.group(1))
            period_end = parse_date(period_match.group(2))
            if current_site:
                ensure_contract_requirement(current_site, period_start, required_guards=1, end_date=period_end)
            continue

        normalized_cells = [normalized_header(value) for value in row]
        if "pers_no" in normalized_cells and "name" in normalized_cells:
            header_row = {header: index for index, header in enumerate(normalized_cells)}
            date_columns = []
            for index, value in enumerate(row):
                date_value = roster_header_date(value, period_start, period_end) if period_start and period_end else None
                if date_value:
                    date_columns.append((index, date_value))
            continue

        if not current_site or not header_row or not date_columns:
            continue

        pers_no_index = header_row.get("pers_no")
        name_index = header_row.get("name")
        pers_no = row[pers_no_index] if pers_no_index is not None else ""
        guard_name = row[name_index] if name_index is not None else ""
        if not pers_no and not guard_name:
            continue

        employee = get_import_guard(pers_no, guard_name, row_number=row_number, site_code=current_site.site_code)
        for column_index, shift_code in date_columns:
            code = str(row[column_index] or "").strip().upper()
            if not code:
                continue
            if is_off_duty_code(code):
                record_roster_attendance(
                    batch_reference=batch_reference,
                    file_name=file_name,
                    source_format=models.RosterAttendance.SourceFormat.SARACEN,
                    source_row=row_number,
                    uploaded_by=uploaded_by,
                    import_status=models.RosterAttendance.ImportStatus.OFF,
                    message="Off duty from Saracen duty roster.",
                    employee=employee,
                    site=current_site,
                    shift_date=shift_code,
                    duty_code="O",
                )
                continue
            shift = get_import_shift(code)
            existing_schedule = models.GuardSchedule.objects.filter(
                employee=employee,
                site=current_site,
                shift=shift,
                shift_date=shift_code,
            ).first()
            if not existing_schedule and contract_schedule_limit_reached(current_site, shift, shift_code):
                message = contract_limit_message(current_site, shift, shift_code)
                skipped_rows.append(f"Row {row_number}: {message}")
                record_roster_attendance(
                    batch_reference=batch_reference,
                    file_name=file_name,
                    source_format=models.RosterAttendance.SourceFormat.SARACEN,
                    source_row=row_number,
                    uploaded_by=uploaded_by,
                    import_status=models.RosterAttendance.ImportStatus.SKIPPED,
                    message=message,
                    employee=employee,
                    site=current_site,
                    shift=shift,
                    shift_date=shift_code,
                    duty_code=code,
                )
                continue
            deployment, _created = models.Deployment.objects.update_or_create(
                employee=employee,
                site=current_site,
                start_date=shift_code,
                defaults={
                    "client": current_site.client,
                    "shift": shift,
                    "end_date": None,
                    "status": models.StatusChoices.ACTIVE,
                },
            )
            schedule, created = models.GuardSchedule.objects.update_or_create(
                deployment=deployment,
                shift_date=shift_code,
                defaults={
                    "employee": employee,
                    "site": current_site,
                    "shift": shift,
                    "status": models.GuardSchedule.ScheduleStatus.SCHEDULED,
                    "notes": "Imported from Saracen duty roster.",
                },
            )
            if created:
                created_schedules += 1
                import_status = models.RosterAttendance.ImportStatus.CREATED
            else:
                updated_schedules += 1
                import_status = models.RosterAttendance.ImportStatus.UPDATED
            record_roster_attendance(
                batch_reference=batch_reference,
                file_name=file_name,
                source_format=models.RosterAttendance.SourceFormat.SARACEN,
                source_row=row_number,
                uploaded_by=uploaded_by,
                import_status=import_status,
                message="Imported from Saracen duty roster.",
                employee=employee,
                site=current_site,
                shift=shift,
                shift_date=shift_code,
                schedule=schedule,
                duty_code=code,
            )

    return created_schedules, updated_schedules, skipped_rows, batch_reference


@login_required
@user_passes_test(can_manage_attendance)
def duty_roster_template(request):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Scheduled Guards"
    _, month_end = month_date_range(timezone.localdate().strftime("%Y-%m"))
    day_headers = []
    weekday_row = ["", "", ""]
    for day in range(1, 32):
        suffix = "th"
        if day % 10 == 1 and day != 11:
            suffix = "st"
        elif day % 10 == 2 and day != 12:
            suffix = "nd"
        elif day % 10 == 3 and day != 13:
            suffix = "rd"
        day_headers.append(f"{day}{suffix}")
        if day <= month_end.day:
            weekday_row.append(month_end.replace(day=day).strftime("%a").upper())
        else:
            weekday_row.append("")
    worksheet.append(["guard_id", "site_code", "shift", *day_headers])
    worksheet.append(weekday_row)

    def monthly_duty_row(duty_code, off_day_numbers):
        off_days = set(off_day_numbers)
        return ["O" if day in off_days else duty_code for day in range(1, 32)]

    worksheet.append(["G001", "S001", "D", *monthly_duty_row("D", {2, 8, 14, 20, 26})])
    worksheet.append(["G002", "S001", "N", *monthly_duty_row("N", {5, 11, 17, 23, 29})])
    worksheet.append(["G003", "S002", "D", *monthly_duty_row("D", {1, 12, 18, 24, 30})])
    for column in range(1, 35):
        worksheet.column_dimensions[worksheet.cell(row=1, column=column).column_letter].width = 12
    for column in ("A", "B", "C"):
        worksheet.column_dimensions[column].width = 22
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="duty-roster-template.xlsx"'
    return response


@login_required
@user_passes_test(can_manage_attendance)
def upload_duty_roster(request):
    if request.method == "POST":
        roster_file = request.FILES.get("roster_file")
        if not roster_file:
            messages.error(request, "Please choose an Excel or CSV scheduled-guard file to upload.")
            return redirect("core:upload_duty_roster")
        try:
            validate_schedule_upload(roster_file)
        except ValidationError as error:
            messages.error(request, " ".join(error.messages))
            return redirect("core:upload_duty_roster")
        try:
            uploaded_rows, worksheet = schedule_file_rows(roster_file)
        except Exception:
            messages.error(request, "The uploaded file could not be read as an Excel workbook or CSV file.")
            return redirect("core:upload_duty_roster")

        batch_reference = str(uuid.uuid4())
        file_name = roster_file.name
        roster_month = request.POST.get("roster_month") or timezone.localdate().strftime("%Y-%m")
        header_row = uploaded_rows[0] if uploaded_rows else []
        headers = {normalized_header(value): index for index, value in enumerate(header_row)}
        required_groups = [
            ("guard_id", "guard_company_number", "company_number", "employee_number", "guard_badge", "badge_number", "guard"),
            ("site_code",),
            ("shift", "shift_name"),
        ]
        missing = [group[0] for group in required_groups if not any(name in headers for name in group)]
        if missing and worksheet is not None:
            has_saracen_blocks = any(
                "site roster for" in str(row[0] or "").lower()
                for row in uploaded_rows
                if row
            )
            if has_saracen_blocks:
                created_schedules, updated_schedules, skipped_rows, batch_reference = import_saracen_roster(
                    worksheet,
                    file_name=file_name,
                    uploaded_by=request.user,
                    batch_reference=batch_reference,
                )
                if created_schedules or updated_schedules:
                    messages.success(
                        request,
                        f"Duty roster imported: {created_schedules} schedules created, {updated_schedules} updated. Batch: {batch_reference}",
                    )
                else:
                    messages.error(request, "No duty schedules were imported from this workbook.")
                if skipped_rows:
                    messages.error(request, "Skipped rows: " + "; ".join(skipped_rows[:8]))
                return redirect("core:attendances")
            wide_rows, wide_header_index, _wide_headers, _wide_date_columns = find_wide_monthly_roster_header(worksheet)
            if wide_header_index is not None:
                stored_rows, off_rows, skipped_rows, batch_reference = import_wide_monthly_roster(
                    worksheet,
                    roster_month=roster_month,
                    file_name=file_name,
                    uploaded_by=request.user,
                    batch_reference=batch_reference,
                )
                if stored_rows or off_rows:
                    messages.success(
                        request,
                        f"Monthly roster imported: {stored_rows} duty rows stored, {off_rows} off rows marked O. Batch: {batch_reference}",
                    )
                else:
                    messages.error(request, "No monthly roster rows were imported from this workbook.")
                if skipped_rows:
                    messages.error(request, "Skipped rows: " + "; ".join(skipped_rows[:8]))
                return redirect("core:upload_duty_roster")
            messages.error(
                request,
                "Missing required schedule columns: " + ", ".join(missing).replace("_", " "),
            )
            return redirect("core:upload_duty_roster")
        elif missing:
            messages.error(
                request,
                "Missing required schedule columns: " + ", ".join(missing).replace("_", " "),
            )
            return redirect("core:upload_duty_roster")

        if monthly_grid_day_columns(header_row):
            created_schedules, updated_schedules, off_rows, skipped_rows, batch_reference = import_monthly_schedule_grid(
                uploaded_rows,
                headers,
                roster_month=roster_month,
                file_name=file_name,
                uploaded_by=request.user,
                batch_reference=batch_reference,
            )
            if created_schedules or updated_schedules or off_rows:
                messages.success(
                    request,
                    (
                        f"Monthly guard schedule imported: {created_schedules} schedules created, "
                        f"{updated_schedules} updated, {off_rows} off days recorded. Batch: {batch_reference}"
                    ),
                )
            else:
                messages.error(request, "No monthly guard schedules were imported from this file.")
            if skipped_rows:
                messages.error(request, "Skipped rows: " + "; ".join(skipped_rows[:8]))
            return redirect("core:attendances")

        created_schedules = 0
        updated_schedules = 0
        skipped_rows = []

        for row_number, row in enumerate(uploaded_rows[1:], start=2):
            if not any(row):
                continue
            guard_value = value_from_row(
                row,
                headers,
                "guard_id",
                "guard_company_number",
                "company_number",
                "employee_number",
                "guard_badge",
                "badge_number",
                "guard",
            )
            site_value = value_from_row(row, headers, "site_code", "site", "site_name")
            shift_value = value_from_row(row, headers, "shift", "shift_name")
            schedule_dates = schedule_dates_from_row(row, headers, roster_month)
            if not schedule_dates:
                message = "invalid schedule date range"
                skipped_rows.append(f"Row {row_number}: {message}")
                record_roster_attendance(
                    batch_reference=batch_reference,
                    file_name=file_name,
                    source_format=models.RosterAttendance.SourceFormat.SIMPLE,
                    source_row=row_number,
                    uploaded_by=request.user,
                    import_status=models.RosterAttendance.ImportStatus.SKIPPED,
                    message=message,
                    duty_code=shift_value,
                )
                continue

            employee = models.Employee.objects.filter(
                Q(company_number__iexact=guard_value)
                | Q(work_card_uid__iexact=guard_value)
                | Q(national_id__iexact=guard_value)
                | Q(first_name__iexact=guard_value)
                | Q(last_name__iexact=guard_value)
            ).first()
            site = models.Site.objects.select_related("client").filter(
                Q(site_code__iexact=site_value)
            ).first()
            shift = models.Shift.objects.filter(Q(shift_name__iexact=shift_value) | Q(code__iexact=shift_value)).first()

            if not employee or not site or not shift:
                message = (
                    f"{'guard not found' if not employee else ''} "
                    f"{'site not found' if not site else ''} "
                    f"{'shift not found' if not shift else ''}".strip()
                )
                skipped_rows.append(f"Row {row_number}: {message}")
                record_roster_attendance(
                    batch_reference=batch_reference,
                    file_name=file_name,
                    source_format=models.RosterAttendance.SourceFormat.SIMPLE,
                    source_row=row_number,
                    uploaded_by=request.user,
                    import_status=models.RosterAttendance.ImportStatus.SKIPPED,
                    message=message,
                    employee=employee,
                    site=site,
                    shift=shift,
                    duty_code=shift_value,
                )
                continue

            deployment_start = min(schedule_dates)
            deployment_end = max(schedule_dates)
            for shift_date in schedule_dates:
                schedule, created, error_message = upsert_guard_schedule_from_upload(
                    employee,
                    site,
                    shift,
                    shift_date,
                    deployment_start=deployment_start,
                    deployment_end=deployment_end,
                    notes="Imported from scheduled-guard upload.",
                )
                if error_message:
                    skipped_rows.append(f"Row {row_number} {shift_date}: {error_message}")
                    record_roster_attendance(
                        batch_reference=batch_reference,
                        file_name=file_name,
                        source_format=models.RosterAttendance.SourceFormat.SIMPLE,
                        source_row=row_number,
                        uploaded_by=request.user,
                        import_status=models.RosterAttendance.ImportStatus.SKIPPED,
                        message=error_message,
                        employee=employee,
                        site=site,
                        shift=shift,
                        shift_date=shift_date,
                        duty_code=shift_value,
                    )
                    continue
                if created:
                    created_schedules += 1
                    import_status = models.RosterAttendance.ImportStatus.CREATED
                else:
                    updated_schedules += 1
                    import_status = models.RosterAttendance.ImportStatus.UPDATED
                record_roster_attendance(
                    batch_reference=batch_reference,
                    file_name=file_name,
                    source_format=models.RosterAttendance.SourceFormat.SIMPLE,
                    source_row=row_number,
                    uploaded_by=request.user,
                    import_status=import_status,
                    message="Imported from scheduled-guard upload.",
                    employee=employee,
                    site=site,
                    shift=shift,
                    shift_date=shift_date,
                    schedule=schedule,
                    duty_code=shift_value,
                )

        if created_schedules or updated_schedules:
            messages.success(
                request,
                f"Duty roster imported: {created_schedules} schedules created, {updated_schedules} updated. Batch: {batch_reference}",
            )
        if skipped_rows:
            messages.error(request, "Skipped rows: " + "; ".join(skipped_rows[:8]))
        return redirect("core:attendances")

    recent_roster_rows = models.RosterAttendance.objects.select_related("employee", "site", "shift", "schedule").order_by("-created_at")[:20]
    return render(
        request,
        "core/upload_duty_roster.html",
        {
            "recent_roster_rows": recent_roster_rows,
            "summary_month": timezone.localdate().strftime("%Y-%m"),
        },
    )


@login_required
@user_passes_test(lambda user: is_manager(user) or is_supervisor(user))
def zonal_guard_list(request):
    zones = models.Zone.objects.select_related("supervisor").prefetch_related(
        "employee_allocations__employee", "site_allocations__site"
    )
    return render(request, "core/zonal_guard_list.html", {"zones": zones})


@login_required
@user_passes_test(lambda user: is_manager(user) or is_supervisor(user))
def zone_shift_summary(request):
    selected_date = request.GET.get("date") or timezone.localdate().isoformat()
    selected_zone_id = request.GET.get("zone") or ""
    parsed_date = parse_date(selected_date) or timezone.localdate()

    zones = models.Zone.objects.select_related("supervisor").order_by("zone_name")
    if selected_zone_id:
        zones = zones.filter(id=selected_zone_id)

    summary_rows = []
    total_scheduled = 0
    total_present = 0
    total_absent = 0
    total_replacements = 0

    for zone in zones:
        site_ids = models.ZoneSiteAllocation.objects.filter(
            zone=zone,
            status=models.StatusChoices.ACTIVE,
            end_date__isnull=True,
        ).values_list("site_id", flat=True)
        schedules = models.GuardSchedule.objects.select_related(
            "employee",
            "replacement_employee",
            "site",
            "shift",
            "attendance",
        ).filter(site_id__in=site_ids, shift_date=parsed_date)

        shifts = {}
        for schedule in schedules:
            shift_key = schedule.shift_id
            shift_name = schedule.shift.shift_name
            row = shifts.setdefault(
                shift_key,
                {
                    "zone": zone,
                    "shift": shift_name,
                    "shift_code": schedule.shift.code or shift_name,
                    "basic_hours": schedule.shift.basic_hours,
                    "overtime_hours": schedule.shift.daily_overtime_hours,
                    "scheduled": 0,
                    "present": 0,
                    "absent": 0,
                    "replacements": 0,
                    "guards": [],
                },
            )
            row["scheduled"] += 1
            try:
                attendance = schedule.attendance
            except models.Attendance.DoesNotExist:
                attendance = None
            attendance_status = attendance.status if attendance else ""
            if attendance_status == "Present":
                row["present"] += 1
            elif attendance_status == "Absent" or schedule.status == models.GuardSchedule.ScheduleStatus.MISSED:
                row["absent"] += 1
            if schedule.replacement_employee_id:
                row["replacements"] += 1
            row["guards"].append(
                {
                    "scheduled": schedule.employee.full_name,
                    "site": schedule.site.site_name,
                    "status": attendance_status or schedule.get_status_display(),
                    "replacement": schedule.replacement_employee.full_name if schedule.replacement_employee else "",
                    "reason": schedule.replacement_reason,
                }
            )

        for row in shifts.values():
            total_scheduled += row["scheduled"]
            total_present += row["present"]
            total_absent += row["absent"]
            total_replacements += row["replacements"]
            summary_rows.append(row)

    return render(
        request,
        "core/zone_shift_summary.html",
        {
            "zones": models.Zone.objects.order_by("zone_name"),
            "selected_date": parsed_date.isoformat(),
            "selected_zone_id": selected_zone_id,
            "summary_rows": summary_rows,
            "totals": {
                "scheduled": total_scheduled,
                "present": total_present,
                "absent": total_absent,
                "replacements": total_replacements,
            },
        },
    )


@login_required
@user_passes_test(can_manage_attendance)
def attendance_report(request):
    employee_number = request.GET.get("employee_number", "").strip()
    start_date = request.GET.get("start_date", "").strip()
    end_date = request.GET.get("end_date", "").strip()
    search_query = request.GET.get("q", "").strip()

    start = parse_date(start_date) if start_date else timezone.localdate().replace(day=1)
    end = parse_date(end_date) if end_date else timezone.localdate()

    schedules = models.GuardSchedule.objects.select_related(
        "employee",
        "replacement_employee",
        "site",
        "shift",
        "attendance",
    ).filter(shift_date__range=(start, end))

    if employee_number:
        schedules = schedules.filter(
            Q(employee__company_number__icontains=employee_number)
            | Q(employee__national_id__icontains=employee_number)
            | Q(employee__first_name__icontains=employee_number)
            | Q(employee__last_name__icontains=employee_number)
            | Q(replacement_employee__company_number__icontains=employee_number)
            | Q(replacement_employee__national_id__icontains=employee_number)
            | Q(replacement_employee__first_name__icontains=employee_number)
            | Q(replacement_employee__last_name__icontains=employee_number)
        )

    rows = []
    for schedule in schedules.order_by("shift_date", "site__site_name", "employee__first_name"):
        try:
            attendance = schedule.attendance
        except models.Attendance.DoesNotExist:
            attendance = None

        scheduled_employee = f"{schedule.employee.company_number}-{schedule.employee.full_name}"
        attended_employee = scheduled_employee if attendance and attendance.status == "Present" else "NONE"
        replacement = "NONE"
        replacement_attendance = None
        if schedule.replacement_employee:
            replacement = f"{schedule.replacement_employee.company_number}-{schedule.replacement_employee.full_name}"
            replacement_attendance = models.Attendance.objects.filter(
                employee=schedule.replacement_employee,
                date=schedule.shift_date,
                shift=schedule.shift,
                site=schedule.site,
                status__iexact="Present",
            ).first()
            attended_employee = replacement

        row = {
            "date_scheduled": schedule.shift_date,
            "site_scheduled": f"({schedule.site.site_code}) {schedule.site.site_name}",
            "scheduled_employee": scheduled_employee,
            "shift_code": schedule.shift.code or schedule.shift.shift_name,
            "basic_hours": schedule.shift.basic_hours,
            "overtime_hours": schedule.shift.daily_overtime_hours,
            "attendance": attended_employee,
            "replacement": replacement,
            "recorded_by": "system",
            "date_recorded": (
                replacement_attendance.updated_at
                if replacement_attendance
                else attendance.updated_at if attendance else schedule.updated_at
            ),
        }
        haystack = " ".join(str(value) for value in row.values()).lower()
        if not search_query or search_query.lower() in haystack:
            rows.append(row)

    return render(
        request,
        "core/attendance_report.html",
        {
            "employee_number": employee_number,
            "start_date": start.isoformat(),
            "end_date": end.isoformat(),
            "search_query": search_query,
            "rows": rows,
        },
    )


@login_required
@user_passes_test(lambda user: can_manage_slug(user, "assets"))
def asset_report(request):
    assets = models.Asset.objects.select_related("assigned_to")
    return render(request, "core/asset_report.html", {"assets": assets})
